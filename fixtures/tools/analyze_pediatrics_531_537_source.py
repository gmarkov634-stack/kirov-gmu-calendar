#!/usr/bin/env python3
"""Mechanical structural inventory for KGMU Pediatrics course 5, groups 531-537.

This operator aid preserves workbook evidence needed before semantic normalization.
It does not resolve contradictions or infer schedule semantics.
"""
import hashlib
import io
import json
import urllib.request
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "pediatrics-531-537-source-inventory.json"
SOURCE_URL = "https://kirovgma.ru/sites/default/files/files/2026/08/24/1085/5_ped-24-08-2026-13.xlsx"
EXPECTED_SHA256 = "190d990d2c505490696d04339f13450f03085c85db997ec3ff5b047ac1c27024"
EXPECTED_BYTES = 26428
EXPECTED_SHEET = "2026-2027 осень 5 курс  Пед"
EXPECTED_GROUPS = [str(value) for value in range(531, 538)]
MONTHS = {"Сентябрь": 9, "Октябрь": 10, "Ноябрь": 11, "Декабрь": 12, "Январь": 1}


def fetch() -> bytes:
    req = urllib.request.Request(SOURCE_URL, headers={"User-Agent": "kgmu-calendar-source-inventory/1.0"})
    with urllib.request.urlopen(req, timeout=45) as response:
        data = response.read()
    actual = hashlib.sha256(data).hexdigest()
    if actual != EXPECTED_SHA256:
        raise SystemExit(f"source sha mismatch: {actual}")
    if len(data) != EXPECTED_BYTES:
        raise SystemExit(f"source byte length mismatch: {len(data)}")
    return data


def text(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def main():
    wb = load_workbook(io.BytesIO(fetch()), data_only=False)
    if wb.sheetnames != [EXPECTED_SHEET]:
        raise SystemExit(f"unexpected sheets: {wb.sheetnames}")
    ws = wb[EXPECTED_SHEET]
    merged = list(ws.merged_cells.ranges)
    if (ws.max_row, ws.max_column, len(merged)) != (43, 125, 194):
        raise SystemExit(f"unexpected workbook geometry: {(ws.max_row, ws.max_column, len(merged))}")

    non_empty = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
    if non_empty != 473:
        raise SystemExit(f"unexpected non-empty cell count: {non_empty}")

    month_by_col = {}
    for rng in merged:
        if rng.min_row == 10 and rng.max_row == 10:
            value = text(ws.cell(10, rng.min_col).value)
            if value in MONTHS:
                for col in range(rng.min_col, rng.max_col + 1):
                    month_by_col[col] = MONTHS[value]

    date_by_col = {}
    weekday_by_col = {}
    for col in range(3, ws.max_column + 1):
        raw_day = ws.cell(11, col).value
        month = month_by_col.get(col)
        weekday = text(ws.cell(12, col).value)
        if weekday:
            weekday_by_col[col] = weekday
        if raw_day is None or month is None:
            continue
        day = int(raw_day)
        year = 2027 if month == 1 else 2026
        date_by_col[col] = date(year, month, day).isoformat()

    group_rows = []
    value_usage = defaultdict(lambda: {"groups": set(), "locators": [], "dateCount": 0})
    for row in range(13, 20):
        group = text(ws.cell(row, 2).value)
        if group not in EXPECTED_GROUPS:
            raise SystemExit(f"unexpected group at B{row}: {group!r}")
        blocks = []
        for rng in merged:
            if rng.min_row != row or rng.max_row != row or rng.min_col < 3:
                continue
            value = text(ws.cell(rng.min_row, rng.min_col).value)
            if value is None:
                continue
            dates = [date_by_col[c] for c in range(rng.min_col, rng.max_col + 1) if c in date_by_col]
            locator = f"{get_column_letter(rng.min_col)}{row}"
            cell = ws[locator]
            block = {
                "locator": locator,
                "range": str(rng),
                "value": value,
                "startDate": dates[0] if dates else None,
                "endDate": dates[-1] if dates else None,
                "dateCount": len(dates),
                "dates": dates,
                "styleId": cell.style_id,
                "font": {
                    "bold": bool(cell.font.bold),
                    "italic": bool(cell.font.italic),
                    "superscript": cell.font.vertAlign == "superscript",
                    "subscript": cell.font.vertAlign == "subscript",
                },
            }
            blocks.append(block)
            usage = value_usage[value]
            usage["groups"].add(group)
            usage["locators"].append(locator)
            usage["dateCount"] += len(dates)
        blocks.sort(key=lambda item: ws[item["locator"]].column)
        group_rows.append({"group": group, "row": row, "blocks": blocks})

    # Multi-row merged blocks that intersect the group grid (e.g. service exam period).
    shared_group_blocks = []
    for rng in merged:
        if rng.max_row < 13 or rng.min_row > 19 or rng.min_col < 3:
            continue
        if rng.min_row == rng.max_row:
            continue
        value = text(ws.cell(rng.min_row, rng.min_col).value)
        if value is None:
            continue
        dates = [date_by_col[c] for c in range(rng.min_col, rng.max_col + 1) if c in date_by_col]
        shared_group_blocks.append({
            "locator": f"{get_column_letter(rng.min_col)}{rng.min_row}",
            "range": str(rng),
            "value": value,
            "rows": [rng.min_row, rng.max_row],
            "startDate": dates[0] if dates else None,
            "endDate": dates[-1] if dates else None,
            "dateCount": len(dates),
            "dates": dates,
        })

    reference_summary = []
    for row in range(25, 37):
        reference_summary.append({
            "row": row,
            "number": text(ws[f"B{row}"].value),
            "discipline": text(ws[f"C{row}"].value),
            "assessment": text(ws[f"S{row}"].value),
            "department": text(ws[f"Y{row}"].value),
            "base": text(ws[f"AN{row}"].value),
            "address": text(ws[f"BL{row}"].value),
            "firstShift": text(ws[f"BT{row}"].value),
            "secondShift": text(ws[f"BX{row}"].value),
        })

    star_cells = []
    for row in ws.iter_rows():
        for cell in row:
            value = text(cell.value)
            if value and "*" in value:
                star_cells.append({
                    "coord": cell.coordinate,
                    "value": value,
                    "styleId": cell.style_id,
                    "font": {
                        "bold": bool(cell.font.bold),
                        "italic": bool(cell.font.italic),
                        "superscript": cell.font.vertAlign == "superscript",
                    },
                })

    drawings = []
    for image in getattr(ws, "_images", []):
        drawings.append({"type": "image", "anchor": str(getattr(image, "anchor", ""))})
    for chart in getattr(ws, "_charts", []):
        drawings.append({"type": "chart", "anchor": str(getattr(chart, "anchor", ""))})

    unique_values = []
    for value, usage in sorted(value_usage.items()):
        unique_values.append({
            "value": value,
            "groups": sorted(usage["groups"], key=int),
            "locators": usage["locators"],
            "dateCountAcrossBlocks": usage["dateCount"],
        })

    payload = {
        "schema": "kgmu-source-structural-inventory-v1",
        "semanticParsingPerformed": False,
        "sourceUrl": SOURCE_URL,
        "sourceSha256": EXPECTED_SHA256,
        "sourceByteLength": EXPECTED_BYTES,
        "sheetName": EXPECTED_SHEET,
        "geometry": {
            "maxRow": ws.max_row,
            "maxColumn": ws.max_column,
            "mergedRangeCount": len(merged),
            "nonEmptyCellCount": non_empty,
        },
        "calendarColumns": [
            {"column": get_column_letter(c), "date": d, "weekdayLabel": weekday_by_col.get(c)}
            for c, d in sorted(date_by_col.items())
        ],
        "uniqueBlockValues": unique_values,
        "groupRows": group_rows,
        "sharedGroupBlocks": shared_group_blocks,
        "referenceSummary": reference_summary,
        "starCells": star_cells,
        "drawings": drawings,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "groups": [item["group"] for item in group_rows],
        "blockCounts": {item["group"]: len(item["blocks"]) for item in group_rows},
        "sharedGroupBlocks": shared_group_blocks,
        "starCells": star_cells,
        "referenceTimes": [
            {"row": item["row"], "discipline": item["discipline"], "firstShift": item["firstShift"], "secondShift": item["secondShift"]}
            for item in reference_summary
        ],
        "calendarColumnCount": len(date_by_col),
        "drawings": drawings,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
