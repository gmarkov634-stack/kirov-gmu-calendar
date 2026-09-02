#!/usr/bin/env python3
"""Build a deterministic normalized draft + QA for KGMU Dentistry course 2.

This parser is intentionally source-bound and course-specific. It composes the
existing KGMU mixed profile (G + R + S), never publishes a ScheduleVersion, and
fails closed if the pinned official XLSX changes or a semantic fragment cannot
be explained by the current rules.
"""
from __future__ import annotations

import hashlib
import io
import json
import re
import urllib.request
from collections import Counter, defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
PERIOD = "2026-2027-semester-1"
FIXTURE_DIR = ROOT / "fixtures" / PERIOD
QA_DIR = ROOT / "qa" / PERIOD
NORMALIZED_DIR = FIXTURE_DIR / "normalized"
SOURCE_PATH = FIXTURE_DIR / "dentistry-291-294.source.json"
JOB_PATH = FIXTURE_DIR / "dentistry-291-294.parsing-job.json"
OUT_PATH = NORMALIZED_DIR / "dentistry-291-294.normalized.compact.json"
QA_PATH = QA_DIR / "dentistry-291-294.qa-report.json"
EVIDENCE_PATH = QA_DIR / "dentistry-291-294.evidence.json"

EXPECTED_RULES = "kgmu-2026-08-27-v3"
EXPECTED_PROFILE = "mixed"
SOURCE_ID = "dentistry"
SHEET = "2 стомат."
GROUP_BY_COL = {2: "291", 3: "292", 4: "293", 5: "294"}
WEEKDAY_RANGES = [(8, 12, 0, "ПН"), (13, 18, 1, "ВТ"), (19, 23, 2, "СР"),
                  (24, 30, 3, "ЧТ"), (31, 37, 4, "ПТ"), (38, 41, 5, "СБ")]
BUILDING_ADDRESS = {
    "1": "ул. Владимирская, 137",
    "2": "ул. Пролетарская, 38",
    "3": "ул. Владимирская, 112",
}
WEEKDAY_TOKEN = {"пн": 0, "вт": 1, "ср": 2, "чт": 3, "пт": 4, "сб": 5}

# The output labels are the canonical names used by the existing R/S profile.
ALIASES = [
    ("Топографическая анатомия и оперативная хирургия головы и шеи", "Топографическая анатомия и оперативная хирургия головы и шеи"),
    ("Патологическая анатомия-патологическая анатомия головы и шеи", "Патологическая анатомия-патологическая анатомия головы и шеи"),
    ("Нормальная физиология - физиология челюстно-лицевой области", "Нормальная физиология - физиология челюстно-лицевой области"),
    ("Нормальная физиология-физиология челюстно-лицевой области", "Нормальная физиология - физиология челюстно-лицевой области"),
    ("Биологическая химия-биохимия полости рта", "Биологическая химия-биохимия полости рта"),
    ("Микробиология, вирусология - микробиология полости рта", "Микробиология, вирусология - микробиология полости рта"),
    ("Микробиология, вирусология-микробиология полости рта", "Микробиология, вирусология - микробиология полости рта"),
    ("Патофизиология-патофизиология головы и шеи", "Патофизиология-патофизиология головы и шеи"),
    ("Иммунология-клиническая иммунология", "Иммунология-клиническая иммунология"),
    ("Анатомия человека - анатомия головы и шеи", "Анатомия человека-анатомия головы и шеи"),
    ("Анатомия человека-анатомия головы и шеи", "Анатомия человека-анатомия головы и шеи"),
    ("Элективные дисциплины по физической культуре и спорту", "Элективные дисциплины по физической культуре и спорту"),
    ("Элективная дисциплина по физической культуре и спорту", "Элективные дисциплины по физической культуре и спорту"),
    ("Пропедевтическая стоматология", "Пропедевтическая стоматология"),
    ("Психология и педагогика", "Психология и педагогика"),
    ("Правоведение", "Правоведение"),
    ("Нормальная физиология", "Нормальная физиология - физиология челюстно-лицевой области"),
    ("Микробиология, вирусология", "Микробиология, вирусология - микробиология полости рта"),
    ("Час куратора", "Час куратора"),
]
ALIASES = sorted(ALIASES, key=lambda item: len(item[0]), reverse=True)
ALIAS_OUTPUT = {alias.casefold(): canonical for alias, canonical in ALIASES}
ALT = "|".join(re.escape(alias) for alias, _ in ALIASES)
TIME_TOKEN = r"\d{1,2}[.:]\d{2}"
TIME_PAIR = rf"{TIME_TOKEN}\s*[-–—]\s*{TIME_TOKEN}"
EVENT_HEAD_RE = re.compile(
    rf"(?P<times>{TIME_PAIR}(?:\s*,\s*{TIME_PAIR})*)\s*(?P<lecture>Лекция\s+)?(?P<discipline>{ALT})",
    re.IGNORECASE,
)
COUNT_NOTE_RE = re.compile(
    r"(?P<count>\d+)\s+заняти\w*\s+в[о.]?\s*(?P<weekday>пн|вт|ср|чт|пт|сб)\.?” ,
    re.IGNORECASE,
)
DATE_TIME_TRIPLE_RE = re.compile(
    rf"(?P<date>\d{{1,2}}\.\d{{2}})\s*-\s*(?P<start>{TIME_TOKEN})\s*-\s*(?P<end>{TIME_TOKEN})"
)
DATE_SPACE_TIME_RE = re.compile(
    rf"(?P<date>\d{{1,2}}\.\d{{2}})\s+(?P<start>{TIME_TOKEN})\s*-\s*(?P<end>{TIME_TOKEN})"
)
DATE_RANGE_RE = re.compile(r"(?P<start>\d{1,2}\.\d{2})\s*-\s*(?P<end>\d{1,2}\.\d{2})")
DATE_TOKEN_RE = re.compile(r"\b\d{1,2}\.\d{2}\b")


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def compact_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def fetch(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "kgmu-calendar-dentistry-291-294/1.0"})
    with urllib.request.urlopen(request, timeout=45) as response:
        return response.read()


def parse_time_token(value: str) -> str:
    value = value.replace(":", ".")
    hour, minute = value.split(".", 1)
    if len(minute) != 2:
        raise ValueError(f"invalid time token {value!r}")
    h, m = int(hour), int(minute)
    if h > 23 or m > 59:
        raise ValueError(f"invalid time token {value!r}")
    return f"{h:02d}:{m:02d}"


def parse_time_pairs(value: str):
    pairs = []
    for match in re.finditer(rf"(?P<start>{TIME_TOKEN})\s*[-–—]\s*(?P<end>{TIME_TOKEN})", value):
        pairs.append((parse_time_token(match.group("start")), parse_time_token(match.group("end"))))
    if not pairs:
        raise ValueError(f"no time pairs in {value!r}")
    return pairs


def to_minutes(value: str) -> int:
    hour, minute = map(int, value.split(":"))
    return hour * 60 + minute


def parse_short_date(token: str) -> date:
    day, month = map(int, token.split("."))
    year = 2027 if month == 1 else 2026
    return date(year, month, day)


def iter_dates(start: date, end: date):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def canonicalize_discipline(raw: str) -> str:
    key = compact_text(raw).casefold()
    if key in ALIAS_OUTPUT:
        return ALIAS_OUTPUT[key]
    # Lower reference rows may vary only in spacing around hyphens.
    collapsed = re.sub(r"\s*-\s*", "-", key)
    for alias, canonical in ALIASES:
        if re.sub(r"\s*-\s*", "-", alias.casefold()) == collapsed:
            return canonical
    return compact_text(raw)


def explicit_location(text: str):
    raw = compact_text(text)
    building_aud = re.search(
        r"(?P<building>[123])\s*корпус\s*,?\s*ауд\.?\s*(?P<aud>\d+)", raw, re.IGNORECASE
    )
    if building_aud:
        building = building_aud.group("building")
        return f"{building} корпус, аудитория {building_aud.group('aud')}, {BUILDING_ADDRESS[building]}"
    short = re.search(r"(?<!\d)(?P<building>[123])\s*-\s*(?P<aud>\d{3})(?!\d)", raw)
    if short:
        building = short.group("building")
        return f"{building} корпус, аудитория {short.group('aud')}, {BUILDING_ADDRESS[building]}"
    return None


def address_from_text(text: str):
    raw = compact_text(text)
    match = re.search(r"ул\.?\s*([А-Яа-яЁё-]+)\s*,?\s*(\d+)", raw, re.IGNORECASE)
    if not match:
        return None
    return f"ул. {match.group(1)}, {match.group(2)}"


def lower_reference_location(value: str, *, from_practice_base: bool):
    raw = compact_text(value)
    if not raw:
        return None
    building = re.search(r"(?P<building>[123])\s*корпус", raw, re.IGNORECASE)
    address = address_from_text(raw)
    if "ФОК" in raw.upper() and address:
        if building:
            return f"ФОК, {building.group('building')} корпус, {address}"
        return f"ФОК, {address}"
    if building and address:
        return f"{building.group('building')} корпус, {address}"
    if address:
        if from_practice_base:
            # R62: a named practical base is part of the location; keep source wording.
            prefix = re.split(r"ул\.?", raw, maxsplit=1, flags=re.IGNORECASE)[0].strip(" (,.;")
            if prefix:
                return f"{prefix}, {address}"
        return address
    return raw if from_practice_base else None


def group_scope(ws, row: int, col: int):
    min_col = max_col = col
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            min_col, max_col = rng.min_col, rng.max_col
            break
    groups = [GROUP_BY_COL[c] for c in range(min_col, max_col + 1) if c in GROUP_BY_COL]
    if not groups and col in GROUP_BY_COL:
        groups = [GROUP_BY_COL[col]]
    return groups


def weekday_for_row(row: int):
    for start, end, weekday, _ in WEEKDAY_RANGES:
        if start <= row <= end:
            return weekday
    raise ValueError(f"row outside weekly table: {row}")


def parse_service_dates(value: str):
    text = compact_text(value).replace("–", "-").replace("—", "-")
    parts = re.split(r"(?=\b[12]\s+неделя\s*-)", text)
    result = {1: set(), 2: set()}
    intervals = []
    for part in parts:
        week_match = re.match(r"\s*([12])\s+неделя\s*-\s*(.*)", part)
        if not week_match:
            continue
        week = int(week_match.group(1))
        payload = week_match.group(2)
        for start_token, end_token in re.findall(r"(\d{1,2}\.\d{2})\s*-\s*(\d{1,2}\.\d{2})", payload):
            start, end = parse_short_date(start_token), parse_short_date(end_token)
            intervals.append({"week": week, "start": start.isoformat(), "end": end.isoformat()})
            for current in iter_dates(start, end):
                result[week].add(current.isoformat())
        # A42 has one explicit isolated 1-week date (09.01).
        cleaned = DATE_RANGE_RE.sub(" ", payload)
        for token in DATE_TOKEN_RE.findall(cleaned):
            current = parse_short_date(token)
            result[week].add(current.isoformat())
            intervals.append({"week": week, "start": current.isoformat(), "end": current.isoformat()})
    if not result[1] or not result[2]:
        raise SystemExit("failed to parse service week grid from A42")
    return result, intervals


def normalize_source_text(raw: str, locator: str, corrections: list):
    text = compact_text(raw).replace("–", "-").replace("—", "-")
    if "16.000" in text:
        corrections.append({"locator": locator, "rule": "R06", "from": "16.000", "to": "16.00", "reason": "impossible minute width; unique local time interpretation"})
        text = text.replace("16.000", "16.00")
    if "24.12.12" in text:
        corrections.append({"locator": locator, "rule": "R06/R47", "from": "24.12.12", "to": "24.12", "reason": "duplicated .12 suffix inside 2026 semester; unique valid date"})
        text = text.replace("24.12.12", "24.12")
    return text


def fragment_date_plan(tail: str, weekday: int, service_dates: set[str]):
    work = tail
    overrides = {}
    override_records = []
    for regex in (DATE_TIME_TRIPLE_RE, DATE_SPACE_TIME_RE):
        matches = list(regex.finditer(work))
        for match in matches:
            current = parse_short_date(match.group("date")).isoformat()
            start = parse_time_token(match.group("start"))
            end = parse_time_token(match.group("end"))
            overrides[current] = (start, end)
            override_records.append({"date": current, "startTime": start, "endTime": end})
        work = regex.sub(" ", work)

    ranges = []
    for match in DATE_RANGE_RE.finditer(work):
        start, end = parse_short_date(match.group("start")), parse_short_date(match.group("end"))
        if end < start:
            raise ValueError(f"impossible date range {match.group(0)!r}")
        ranges.append((start, end))
    work = DATE_RANGE_RE.sub(" ", work)
    explicit = sorted({parse_short_date(token).isoformat() for token in DATE_TOKEN_RE.findall(work)})

    dates = {}
    for start, end in ranges:
        for current in iter_dates(start, end):
            iso = current.isoformat()
            if current.weekday() == weekday and iso in service_dates:
                dates.setdefault(iso, {"provenance": "computed-range"})
    for iso in explicit:
        dates[iso] = {"provenance": "explicit-date"}
    for iso, times in overrides.items():
        dates[iso] = {"provenance": "explicit-date-time", "override": times}
    return dates, {
        "ranges": [[start.isoformat(), end.isoformat()] for start, end in ranges],
        "explicitDates": explicit,
        "timeOverrides": override_records,
    }


def event_id(group, day, start, end, discipline, lesson_type, location, locator):
    payload = "|".join([group, day, start, end, discipline, lesson_type, location or "", locator])
    return "kgmu-" + hashlib.sha256(payload.encode("utf-8")).hexdigest()[:24]


def add_event(events, metadata, *, group, day, start, end, discipline, lesson_type,
              location, locator, interval_count=1, provenance="source"):
    if to_minutes(end) <= to_minutes(start):
        raise ValueError(f"non-positive event interval {locator} {start}-{end}")
    identity = (group, day, start, end, discipline, lesson_type, location)
    if identity in metadata["dedupe"]:
        metadata["deduplicated"].append({"groupId": group, "date": day, "discipline": discipline, "sourceLocator": locator})
        return
    metadata["dedupe"].add(identity)
    eid = event_id(group, day, start, end, discipline, lesson_type, location, locator)
    events.append({
        "eventId": eid,
        "groupId": group,
        "date": day,
        "startTime": start,
        "endTime": end,
        "discipline": discipline,
        "lessonType": lesson_type,
        "location": location,
        "sourceLocator": locator,
    })
    metadata["eventMeta"][eid] = {
        "sourceLocator": locator,
        "intervalCount": interval_count,
        "provenance": provenance,
    }


def overlaps(start_a, end_a, start_b, end_b):
    return to_minutes(start_a) < to_minutes(end_b) and to_minutes(start_b) < to_minutes(end_a)


def main():
    source = read_json(SOURCE_PATH)
    job = read_json(JOB_PATH)
    if source["parserProfile"] != EXPECTED_PROFILE:
        raise SystemExit(f"unexpected parser profile: {source['parserProfile']}")
    if source["parserRulesVersion"] != EXPECTED_RULES or job["parserRulesVersion"] != EXPECTED_RULES:
        raise SystemExit("parser rules version drifted from pinned mixed-profile v3")
    if source["expectedGroupIds"] != job["expectedGroupIds"]:
        raise SystemExit("ParsingJob groups differ from source fixture")
    if job["sourceId"] != SOURCE_ID or source["source"]["sourceId"] != SOURCE_ID:
        raise SystemExit("unexpected sourceId")
    if source["source"]["sha256"] not in job["sourceObjectKey"]:
        raise SystemExit("ParsingJob object key is not source-SHA-bound")

    data = fetch(source["source"]["url"])
    digest = hashlib.sha256(data).hexdigest()
    if digest != source["source"]["sha256"]:
        raise SystemExit(f"official source SHA changed: {digest}")
    if len(data) != source["source"]["byteLength"]:
        raise SystemExit("official source byte length changed")
    wb = load_workbook(io.BytesIO(data), data_only=False)
    if wb.sheetnames != source["workbookExpectations"]["sheetNames"]:
        raise SystemExit(f"workbook sheet names changed: {wb.sheetnames}")
    ws = wb[SHEET]
    non_empty = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
    geometry = (ws.max_row, ws.max_column, len(ws.merged_cells.ranges), non_empty)
    expected_geometry = (
        source["workbookExpectations"]["maxRow"],
        source["workbookExpectations"]["maxColumn"],
        source["workbookExpectations"]["mergedRangeCount"],
        source["workbookExpectations"]["nonEmptyCellCount"],
    )
    if geometry != expected_geometry:
        raise SystemExit(f"workbook geometry changed: {geometry} != {expected_geometry}")

    for row, expected in [(8, "ПН"), (13, "ВТ"), (19, "СР"), (24, "ЧТ"), (31, "ПТ"), (38, "СБ")]:
        if compact_text(ws[f"A{row}"].value).upper() != expected:
            raise SystemExit(f"weekday anchor A{row} changed")

    corrections = []
    if "206-2027" in compact_text(ws["B6"].value):
        corrections.append({"locator": f"{SHEET}!B6", "rule": "G04", "from": "206-2027", "to": "2026-2027", "reason": "official page, source filename, semester start and project AcademicYear all identify 2026-2027"})

    week_sets, week_intervals = parse_service_dates(ws["A42"].value)
    service_dates = week_sets[1] | week_sets[2]

    # Lower source table is metadata only (R40/R41/S09), never standalone events.
    lower_refs = {}
    for row in range(50, 63):
        raw_disc = compact_text(ws[f"A{row}"].value)
        if not raw_disc:
            continue
        discipline = canonicalize_discipline(raw_disc)
        department = compact_text(ws[f"C{row}"].value)
        practice_base = compact_text(ws[f"D{row}"].value)
        assessment = compact_text(ws[f"E{row}"].value) or None
        location = None
        source_col = None
        if practice_base:
            location = lower_reference_location(practice_base, from_practice_base=True)
            source_col = "D"
        elif department:
            location = lower_reference_location(department, from_practice_base=False)
            source_col = "C"
        lower_refs[discipline] = {
            "row": row,
            "sourceDiscipline": raw_disc,
            "department": department or None,
            "practiceBase": practice_base or None,
            "location": location,
            "locationSource": f"{SHEET}!{source_col}{row}" if source_col and location else None,
            "assessment": assessment,
            "assessmentSource": f"{SHEET}!E{row}" if assessment else None,
        }

    events = []
    meta = {"dedupe": set(), "deduplicated": [], "eventMeta": {}}
    coverage = []
    fragments_evidence = []
    curator_specs = []
    count_expectations = []
    unmatched_cells = []

    for row in range(8, 42):
        weekday = weekday_for_row(row)
        for col in range(2, 6):
            value = ws.cell(row, col).value
            if value is None:
                continue
            locator = f"{SHEET}!{get_column_letter(col)}{row}"
            raw = normalize_source_text(str(value), locator, corrections)
            heads = list(EVENT_HEAD_RE.finditer(raw))
            if not heads:
                unmatched_cells.append({"locator": locator, "raw": raw})
                continue
            groups = group_scope(ws, row, col)
            if not groups:
                unmatched_cells.append({"locator": locator, "raw": raw, "reason": "no group scope"})
                continue
            coverage.append(locator)
            for index, match in enumerate(heads, start=1):
                tail_start = match.end()
                tail_end = heads[index].start() if index < len(heads) else len(raw)
                tail = raw[tail_start:tail_end]
                source_name = match.group("discipline")
                discipline = canonicalize_discipline(source_name)
                pairs = parse_time_pairs(match.group("times"))
                start_time, end_time = pairs[0][0], pairs[-1][1]
                lesson_type = "lecture" if match.group("lecture") else ("other" if discipline == "Час куратора" else "practice")
                segment_locator = f"{locator}#s{index}"
                location = explicit_location(tail)
                if location is None and discipline != "Час куратора":
                    location = lower_refs.get(discipline, {}).get("location")

                notes = [
                    {"count": int(item.group("count")), "weekday": WEEKDAY_TOKEN[item.group("weekday").casefold()], "token": item.group(0)}
                    for item in COUNT_NOTE_RE.finditer(tail)
                ]
                for group in groups:
                    for note in notes:
                        count_expectations.append({
                            "originLocator": segment_locator,
                            "groupId": group,
                            "discipline": discipline,
                            "expectedCount": note["count"],
                            "targetWeekday": note["weekday"],
                            "sourceStatement": note["token"],
                        })

                if discipline == "Час куратора":
                    exclusions = sorted({parse_short_date(token).isoformat() for token in re.findall(r"кроме\s+([^)]*)", tail, flags=re.IGNORECASE) for token in DATE_TOKEN_RE.findall(token)})
                    for group in groups:
                        curator_specs.append({
                            "groupId": group,
                            "weekday": weekday,
                            "startTime": start_time,
                            "endTime": end_time,
                            "exclusions": exclusions,
                            "sourceLocator": segment_locator,
                        })
                    fragments_evidence.append({
                        "sourceLocator": segment_locator,
                        "groups": groups,
                        "discipline": discipline,
                        "lessonType": lesson_type,
                        "time": [start_time, end_time],
                        "datePlan": {"mode": "first-two-non-overlapping", "exclusions": exclusions},
                        "location": None,
                    })
                    continue

                date_plan, plan_evidence = fragment_date_plan(tail, weekday, service_dates)
                if not date_plan:
                    unmatched_cells.append({"locator": segment_locator, "raw": raw, "reason": "event fragment has no dates/ranges"})
                    continue
                for group in groups:
                    for day, plan in sorted(date_plan.items()):
                        event_start, event_end = start_time, end_time
                        if "override" in plan:
                            event_start, event_end = plan["override"]
                        add_event(
                            events, meta, group=group, day=day, start=event_start, end=event_end,
                            discipline=discipline, lesson_type=lesson_type, location=location,
                            locator=segment_locator, interval_count=len(pairs), provenance=plan["provenance"]
                        )
                fragments_evidence.append({
                    "sourceLocator": segment_locator,
                    "groups": groups,
                    "discipline": discipline,
                    "lessonType": lesson_type,
                    "time": [start_time, end_time],
                    "intervalCount": len(pairs),
                    "datePlan": plan_evidence,
                    "location": location,
                    "countNotes": notes,
                })

    # S01/S08/S09: source-explicit Propedeutic Dentistry cycle by group and period.
    cycle_ref = lower_refs.get("Пропедевтическая стоматология")
    cycle_location = cycle_ref.get("location") if cycle_ref else None
    if not cycle_location:
        unmatched_cells.append({"locator": f"{SHEET}!A43:E48", "reason": "S09 cycle location missing in lower reference table"})
    cycle_time_raw = compact_text(ws["A44"].value)
    cycle_pairs = parse_time_pairs(cycle_time_raw)
    cycle_start, cycle_end = cycle_pairs[0][0], cycle_pairs[-1][1]
    cycle_evidence = []
    for row, group in [(45, "291"), (46, "292"), (47, "293"), (48, "294")]:
        if compact_text(ws[f"A{row}"].value) != group:
            unmatched_cells.append({"locator": f"{SHEET}!A{row}", "reason": "cycle group label changed"})
            continue
        range_text = compact_text(ws[f"B{row}"].value)
        month_names = {"сентября": 9, "октября": 10, "ноября": 11, "декабря": 12}
        range_match = re.search(r"с\s+(\d{1,2})\s+([а-яё]+)\s+по\s+(\d{1,2})\s+([а-яё]+)", range_text, re.IGNORECASE)
        if not range_match:
            unmatched_cells.append({"locator": f"{SHEET}!B{row}", "reason": "cycle range is not recognized"})
            continue
        start = date(2026, month_names[range_match.group(2).casefold()], int(range_match.group(1)))
        end = date(2026, month_names[range_match.group(4).casefold()], int(range_match.group(3)))
        cycle_days = []
        for current in iter_dates(start, end):
            if current.weekday() == 6:
                continue
            if current.isoformat() not in service_dates:
                continue
            day = current.isoformat()
            cycle_days.append(day)
            add_event(
                events, meta, group=group, day=day, start=cycle_start, end=cycle_end,
                discipline="Пропедевтическая стоматология", lesson_type="practice",
                location=cycle_location, locator=f"{SHEET}!B{row}#cycle",
                interval_count=len(cycle_pairs), provenance="cycle-S01"
            )
        cycle_evidence.append({
            "groupId": group,
            "sourceLocator": f"{SHEET}!B{row}",
            "startDate": start.isoformat(),
            "endDate": end.isoformat(),
            "startTime": cycle_start,
            "endTime": cycle_end,
            "eventDays": len(cycle_days),
            "location": cycle_location,
        })

    # R17/S07: first two feasible curator slots after mandatory events + cycles exist.
    mandatory_by_group_date = defaultdict(list)
    for item in events:
        mandatory_by_group_date[(item["groupId"], item["date"])].append(item)
    curator_evidence = []
    for spec in curator_specs:
        candidates = []
        for day in sorted(service_dates):
            current = date.fromisoformat(day)
            if current.weekday() != spec["weekday"] or day in spec["exclusions"]:
                continue
            occupied = mandatory_by_group_date[(spec["groupId"], day)]
            if any(overlaps(spec["startTime"], spec["endTime"], item["startTime"], item["endTime"]) for item in occupied):
                continue
            candidates.append(day)
            if len(candidates) == 2:
                break
        if len(candidates) != 2:
            unmatched_cells.append({"locator": spec["sourceLocator"], "reason": f"only {len(candidates)} feasible curator slots"})
            continue
        for day in candidates:
            add_event(
                events, meta, group=spec["groupId"], day=day, start=spec["startTime"], end=spec["endTime"],
                discipline="Час куратора", lesson_type="other", location=None,
                locator=spec["sourceLocator"], interval_count=1, provenance="curator-R17-S07"
            )
        curator_evidence.append({**spec, "selectedDates": candidates})

    # R07-R09/R67: resolve count/day notes only against other explicit/source-computable blocks.
    cross_day = []
    for expectation in count_expectations:
        candidates = []
        for item in events:
            if item["groupId"] != expectation["groupId"] or item["discipline"] != expectation["discipline"]:
                continue
            if item["sourceLocator"] == expectation["originLocator"]:
                continue
            if date.fromisoformat(item["date"]).weekday() != expectation["targetWeekday"]:
                continue
            emeta = meta["eventMeta"].get(item["eventId"], {})
            if emeta.get("provenance") in {"cycle-S01", "curator-R17-S07"}:
                continue
            candidates.append({
                "date": item["date"],
                "startTime": item["startTime"],
                "endTime": item["endTime"],
                "sourceLocator": item["sourceLocator"],
                "intervalCount": emeta.get("intervalCount", 1),
                "provenance": emeta.get("provenance"),
            })
        # Preserve unique source events; count either calendar occurrences or source intervals.
        unique = {(c["date"], c["startTime"], c["endTime"], c["sourceLocator"]): c for c in candidates}
        candidates = sorted(unique.values(), key=lambda c: (c["date"], c["startTime"], c["sourceLocator"]))
        occurrence_count = len(candidates)
        interval_count = sum(c["intervalCount"] for c in candidates)
        if occurrence_count == expectation["expectedCount"]:
            status, count_mode = "PASS", "occurrences"
        elif interval_count == expectation["expectedCount"]:
            status, count_mode = "PASS", "source-intervals-before-R12"
        else:
            status, count_mode = "REVIEW_REQUIRED", None
        cross_day.append({**expectation, "status": status, "countMode": count_mode, "matchedOccurrences": candidates,
                          "occurrenceCount": occurrence_count, "sourceIntervalCount": interval_count})

    events.sort(key=lambda item: (item["groupId"], item["date"], item["startTime"], item["endTime"], item["discipline"], item["sourceLocator"]))
    group_counts = dict(sorted(Counter(item["groupId"] for item in events).items()))
    tuple_fields = ["eventId", "groupId", "date", "startTime", "endTime", "discipline", "lessonType", "location", "sourceLocator"]
    tuples = [[item[field] for field in tuple_fields] for item in events]
    canonical = json.dumps(tuples, ensure_ascii=False, separators=(",", ":"), sort_keys=False).encode("utf-8")
    candidate_digest = "sha256:" + hashlib.sha256(canonical).hexdigest()

    expected_schedule_cells = []
    for row in range(8, 42):
        for col in range(2, 6):
            if ws.cell(row, col).value is not None:
                expected_schedule_cells.append(f"{SHEET}!{get_column_letter(col)}{row}")
    coverage_ok = set(coverage) == set(expected_schedule_cells)
    cross_day_ok = all(item["status"] == "PASS" for item in cross_day)
    groups_ok = set(group_counts) == set(source["expectedGroupIds"]) and all(group_counts[g] > 0 for g in source["expectedGroupIds"])
    event_ids_unique = len({item["eventId"] for item in events}) == len(events)
    source_locators_ok = all(item["sourceLocator"].startswith(f"{SHEET}!") for item in events)
    cycle_ok = len(cycle_evidence) == 4 and all(item["eventDays"] > 0 and item["location"] for item in cycle_evidence)
    curator_ok = len(curator_evidence) == len(curator_specs) and all(len(item["selectedDates"]) == 2 for item in curator_evidence)

    checks = {
        "sourceIdentityPinned": True,
        "workbookGeometryPinned": True,
        "parserRulesPinned": True,
        "scheduleCellCoverageComplete": coverage_ok,
        "noUnexplainedSourceFragments": len(unmatched_cells) == 0,
        "crossDayCountNotesResolved": cross_day_ok,
        "mixedCycleResolved": cycle_ok,
        "curatorRulesResolved": curator_ok,
        "expectedGroupsPresent": groups_ok,
        "eventIdsUnique": event_ids_unique,
        "sourceLocatorsPresent": source_locators_ok,
    }
    warnings = []
    if unmatched_cells:
        warnings.append(f"REVIEW_REQUIRED: {len(unmatched_cells)} source fragments are unexplained")
    for item in cross_day:
        if item["status"] != "PASS":
            warnings.append(f"REVIEW_REQUIRED: count note mismatch {item['originLocator']} {item['groupId']} {item['discipline']}")
    status = "PASS" if all(checks.values()) and not warnings else "REVIEW_REQUIRED"

    normalized = {
        "fixtureId": "dentistry-291-294-2026-2027-semester-1",
        "sourceFixtureId": source["fixtureId"],
        "parsingJobId": job["jobId"],
        "encoding": "normalized-event-tuples-v1",
        "tupleFields": tuple_fields,
        "constants": {
            "universityId": "kirov-gmu",
            "academicPeriodId": PERIOD,
            "timeSemantics": "floating",
            "sourceId": SOURCE_ID,
            "teacher": None,
        },
        "eventCount": len(events),
        "groupEventCounts": group_counts,
        "candidateDigest": candidate_digest,
        "events": tuples,
    }
    qa = {
        "qaReportId": "qa-dentistry-291-294-ec51c194-v1",
        "fixtureId": source["fixtureId"],
        "parsingJobId": job["jobId"],
        "sourceSha256": source["source"]["sha256"],
        "parserProfile": source["parserProfile"],
        "parserRulesVersion": source["parserRulesVersion"],
        "candidateDigest": candidate_digest,
        "eventCount": len(events),
        "groupEventCounts": group_counts,
        "status": status,
        "checks": checks,
        "warnings": warnings,
        "scheduleVersionCreated": False,
        "published": False,
    }
    evidence = {
        "fixtureId": source["fixtureId"],
        "sourceSha256": source["source"]["sha256"],
        "workbookGeometry": {"maxRow": geometry[0], "maxColumn": geometry[1], "mergedRangeCount": geometry[2], "nonEmptyCellCount": geometry[3]},
        "serviceWeekIntervals": week_intervals,
        "sourceCorrections": corrections,
        "lowerReferenceRows": lower_refs,
        "scheduleCoverage": {"expectedCells": expected_schedule_cells, "coveredCells": sorted(set(coverage)), "unmatched": unmatched_cells},
        "fragments": fragments_evidence,
        "crossDayExpectations": cross_day,
        "mixedCycle": {"rule": "S01/S08/S09", "events": cycle_evidence},
        "curatorHour": {"rule": "R17/S07", "events": curator_evidence},
        "deduplicatedSourceEvents": meta["deduplicated"],
        "normalization": {
            "disciplineCanonicalization": "R10/R36/R64/S05",
            "dateExpansion": "R01-R05/R13/R47/R48/R50/R51/R52/R86",
            "multiEventCells": "R34/R49",
            "intervalMerge": "R12/S08",
            "locations": "R18-R21/R35/R40/R42/R45/R57/R58/R62/S09",
            "deduplication": "R27",
            "explicitOverlapsPreserved": "R69/S04",
            "postprocessingBoundary": "P01: calendar-note/progress rendering is downstream and does not mutate this normalized draft",
        },
    }

    NORMALIZED_DIR.mkdir(parents=True, exist_ok=True)
    QA_DIR.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(normalized, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    QA_PATH.write_text(json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8")
    EVIDENCE_PATH.write_text(json.dumps(evidence, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "status": status,
        "eventCount": len(events),
        "groupEventCounts": group_counts,
        "candidateDigest": candidate_digest,
        "crossDay": [{"locator": item["originLocator"], "group": item["groupId"], "expected": item["expectedCount"], "occurrences": item["occurrenceCount"], "intervals": item["sourceIntervalCount"], "status": item["status"]} for item in cross_day],
        "unmatched": unmatched_cells,
        "corrections": corrections,
        "cycle": cycle_evidence,
        "curator": curator_evidence,
    }, ensure_ascii=False, indent=2))
    if status != "PASS":
        raise SystemExit(2)


if __name__ == "__main__":
    main()
