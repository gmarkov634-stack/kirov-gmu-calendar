#!/usr/bin/env python3
"""Mechanical XLSX source probe for KGMU Pediatrics course 1, groups 131-140.

This helper intentionally performs no semantic parsing. It discovers the current
official XLSX link from the KGMU pediatrics timetable page, then records checksum,
workbook geometry, merged ranges, non-empty cell values and hyperlink targets
for operator/ChatGPT review under canonical G/R rules.
"""
import hashlib
import html
import io
import json
import re
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "pediatrics-131-140-source-probe.json"
TIMETABLE_PAGE = "https://kirovgma.ru/raspisanie-pediatricheskiy-fakultet"
ALLOWED_PREFIX = "/sites/default/files/files/"
SOURCE_PATTERN = re.compile(r"1_ped[^\"'<>\s]*\.xlsx", re.IGNORECASE)


def request_bytes(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "kgmu-calendar-source-probe/1.0"})
    with urllib.request.urlopen(request, timeout=45) as response:
        return response.read()


def discover_source() -> dict:
    page_bytes = request_bytes(TIMETABLE_PAGE)
    page_text = html.unescape(page_bytes.decode("utf-8", errors="replace"))
    hrefs = re.findall(r"href\s*=\s*[\"']([^\"']+)[\"']", page_text, flags=re.IGNORECASE)
    matches = []
    for href in hrefs:
        decoded = urllib.parse.unquote(href)
        if SOURCE_PATTERN.search(decoded):
            absolute = urllib.parse.urljoin(TIMETABLE_PAGE, href)
            parsed = urllib.parse.urlparse(absolute)
            if parsed.netloc != "kirovgma.ru" or not parsed.path.startswith(ALLOWED_PREFIX):
                raise SystemExit(f"discovered source outside allowed KGMU path: {absolute}")
            matches.append(absolute)
    matches = sorted(set(matches))
    if len(matches) != 1:
        raise SystemExit(f"expected exactly one current XLSX for pediatrics 131-140, found {matches}")
    return {"program": "pediatrics", "course": 1, "groups": [str(value) for value in range(131, 141)], "url": matches[0]}


def workbook_dump(source: dict) -> dict:
    data = request_bytes(source["url"])
    if not data.startswith(b"PK"):
        raise SystemExit(f"source is not XLSX/ZIP: {source['url']}")
    workbook = load_workbook(io.BytesIO(data), data_only=False)
    sheets = []
    for sheet in workbook.worksheets:
        non_empty = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.value is None:
                    continue
                entry = {"coord": cell.coordinate, "value": str(cell.value).strip()}
                if cell.hyperlink is not None:
                    entry["hyperlink"] = {
                        "target": cell.hyperlink.target,
                        "location": cell.hyperlink.location,
                        "display": cell.hyperlink.display,
                    }
                non_empty.append(entry)
        sheets.append({
            "title": sheet.title,
            "maxRow": sheet.max_row,
            "maxColumn": sheet.max_column,
            "mergedRanges": [str(rng) for rng in sheet.merged_cells.ranges],
            "nonEmptyCellCount": len(non_empty),
            "nonEmptyCells": non_empty,
        })
    return {
        **source,
        "sha256": hashlib.sha256(data).hexdigest(),
        "byteLength": len(data),
        "sheetNames": workbook.sheetnames,
        "sheets": sheets,
    }


def main() -> None:
    payload = {
        "schema": "kgmu-mechanical-source-probe-v1",
        "semanticParsingPerformed": False,
        "timetablePage": TIMETABLE_PAGE,
        "source": workbook_dump(discover_source()),
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    source = payload["source"]
    print(json.dumps({
        "program": source["program"],
        "course": source["course"],
        "groups": source["groups"],
        "url": source["url"],
        "sha256": source["sha256"],
        "byteLength": source["byteLength"],
        "sheetNames": source["sheetNames"],
        "dimensions": [
            [sheet["title"], sheet["maxRow"], sheet["maxColumn"], len(sheet["mergedRanges"]), sheet["nonEmptyCellCount"]]
            for sheet in source["sheets"]
        ],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
