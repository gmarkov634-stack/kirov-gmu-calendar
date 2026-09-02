#!/usr/bin/env python3
"""Course-specific semantic source analyzer for KGMU Pediatrics course 4.

This helper does not publish anything and does not modify common parser/core code.
It verifies the immutable source fixture, downloads the official XLSX, and emits a
compact review artifact with the date axis, group cycle blocks and lower metadata
table needed to author explicit semantic decisions under parser profile C.
"""
import hashlib
import io
import json
import urllib.request
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
FIXTURE = ROOT / "fixtures/2026-2027-semester-1/pediatrics-431-436.source.json"
OUT = ROOT / "fixtures/2026-2027-semester-1/pediatrics-431-436.semantic-source.json"
MONTHS = {
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
    "январь": 1,
}


def request_bytes(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "kgmu-calendar-pediatrics-431-436-analyzer/1.0"})
    with urllib.request.urlopen(request, timeout=45) as response:
        return response.read()


def merged_range_for(sheet, row: int, column: int):
    for rng in sheet.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= column <= rng.max_col:
            return rng
    return None


def date_axis(sheet):
    current_month = None
    axis = {}
    for column in range(3, sheet.max_column + 1):
        month_value = sheet.cell(10, column).value
        if month_value is not None:
            key = str(month_value).strip().lower()
            if key in MONTHS:
                current_month = MONTHS[key]
        day_value = sheet.cell(11, column).value
        if current_month is None or day_value is None:
            continue
        try:
            day_number = int(day_value)
        except (TypeError, ValueError):
            continue
        year = 2026 if current_month >= 9 else 2027
        try:
            iso = date(year, current_month, day_number).isoformat()
        except ValueError:
            continue
        axis[column] = {
            "column": get_column_letter(column),
            "date": iso,
            "weekday": None if sheet.cell(12, column).value is None else str(sheet.cell(12, column).value).strip(),
        }
    return axis


def cycle_blocks(sheet, axis):
    blocks = []
    for row in range(13, 19):
        group = str(sheet.cell(row, 2).value).strip()
        seen_ranges = set()
        for column in range(3, sheet.max_column + 1):
            cell = sheet.cell(row, column)
            if isinstance(cell, MergedCell):
                continue
            value = cell.value
            if value is None or not str(value).strip():
                continue
            rng = merged_range_for(sheet, row, column)
            if rng is None:
                min_col = max_col = column
                locator = cell.coordinate
            else:
                locator = str(rng)
                if locator in seen_ranges:
                    continue
                seen_ranges.add(locator)
                min_col, max_col = rng.min_col, rng.max_col
            dates = [axis[col]["date"] for col in range(min_col, max_col + 1) if col in axis]
            blocks.append({
                "group": group,
                "locator": locator,
                "value": str(value).strip(),
                "startColumn": get_column_letter(min_col),
                "endColumn": get_column_letter(max_col),
                "dates": dates,
            })
    return blocks


def lower_table(sheet):
    rows = []
    for row in range(22, sheet.max_row + 1):
        values = []
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, column)
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            text = str(cell.value).strip()
            if not text:
                continue
            rng = merged_range_for(sheet, row, column)
            values.append({
                "coord": cell.coordinate,
                "range": str(rng) if rng is not None else cell.coordinate,
                "value": text,
            })
        if values:
            rows.append({"row": row, "cells": values})
    return rows


def main() -> None:
    fixture = json.loads(FIXTURE.read_text(encoding="utf-8"))
    data = request_bytes(fixture["source"]["url"])
    actual_sha = hashlib.sha256(data).hexdigest()
    if actual_sha != fixture["source"]["sha256"]:
        raise SystemExit(f"source sha mismatch: {actual_sha} != {fixture['source']['sha256']}")
    workbook = load_workbook(io.BytesIO(data), data_only=False)
    if workbook.sheetnames != fixture["workbookExpectations"]["sheetNames"]:
        raise SystemExit(f"sheet mismatch: {workbook.sheetnames}")
    sheet = workbook[workbook.sheetnames[0]]
    axis = date_axis(sheet)
    groups = [str(sheet.cell(row, 2).value).strip() for row in range(13, 19)]
    if groups != fixture["expectedGroupIds"]:
        raise SystemExit(f"group mismatch: {groups}")
    payload = {
        "schema": "kgmu-pediatrics-431-436-semantic-source-v1",
        "sourceSha256": actual_sha,
        "sheetName": sheet.title,
        "dateAxis": [axis[col] for col in sorted(axis)],
        "cycleBlocks": cycle_blocks(sheet, axis),
        "lowerTable": lower_table(sheet),
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "sourceSha256": actual_sha,
        "dateCount": len(payload["dateAxis"]),
        "cycleBlockCount": len(payload["cycleBlocks"]),
        "lowerTableRowCount": len(payload["lowerTable"]),
        "groups": groups,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
