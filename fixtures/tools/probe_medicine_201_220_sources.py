#!/usr/bin/env python3
"""Mechanical XLSX source probe for KGMU medicine groups 201-220.

This helper intentionally performs no semantic parsing. It discovers the current
official XLSX links from the KGMU medicine timetable page, then records
checksums, workbook geometry, merged ranges and non-empty cell values for
operator/ChatGPT review under canonical G/R rules.
"""
import hashlib
import html
import io
import json
import re
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "medicine-201-220-source-probe.json"
TIMETABLE_PAGE = "https://kirovgma.ru/lechebnyy-fakultet-raspisanie"
ALLOWED_PREFIX = "/sites/default/files/files/"
STREAM_PATTERNS = {
    "201-210": re.compile(r"2_lech\._1_potok[^\"'<>\s]*\.xlsx", re.IGNORECASE),
    "211-220": re.compile(r"2_lech\._2_potok[^\"'<>\s]*\.xlsx", re.IGNORECASE),
}


def request_bytes(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "kgmu-calendar-source-probe/1.0"})
    with urllib.request.urlopen(request, timeout=45) as response:
        return response.read()


def discover_sources() -> list[dict]:
    page_bytes = request_bytes(TIMETABLE_PAGE)
    page_text = html.unescape(page_bytes.decode("utf-8", errors="replace"))
    hrefs = re.findall(r"href\s*=\s*[\"']([^\"']+)[\"']", page_text, flags=re.IGNORECASE)
    result = []
    for stream, pattern in STREAM_PATTERNS.items():
        matches = []
        for href in hrefs:
            decoded = urllib.parse.unquote(href)
            if pattern.search(decoded):
                absolute = urllib.parse.urljoin(TIMETABLE_PAGE, href)
                parsed = urllib.parse.urlparse(absolute)
                if parsed.netloc != "kirovgma.ru" or not parsed.path.startswith(ALLOWED_PREFIX):
                    raise SystemExit(f"discovered source outside allowed KGMU path: {absolute}")
                matches.append(absolute)
        matches = sorted(set(matches))
        if len(matches) != 1:
            raise SystemExit(f"expected exactly one current XLSX for {stream}, found {matches}")
        result.append({"stream": stream, "url": matches[0]})
    return result


def workbook_dump(source: dict) -> dict:
    data = request_bytes(source["url"])
    if not data.startswith(b"PK"):
        raise SystemExit(f"source is not XLSX/ZIP: {source['url']}")
    workbook = load_workbook(io.BytesIO(data), data_only=False)
    sheets = []
    for sheet in workbook.worksheets:
        non_empty = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.value is None:
                    continue
                non_empty.append({"coord": cell.coordinate, "value": str(cell.value).strip()})
        sheets.append(
            {
                "title": sheet.title,
                "maxRow": sheet.max_row,
                "maxColumn": sheet.max_column,
                "mergedRanges": [str(rng) for rng in sheet.merged_cells.ranges],
                "nonEmptyCellCount": len(non_empty),
                "nonEmptyCells": non_empty,
            }
        )
    return {
        **source,
        "sha256": hashlib.sha256(data).hexdigest(),
        "byteLength": len(data),
        "sheetNames": workbook.sheetnames,
        "sheets": sheets,
    }


def main() -> None:
    discovered = discover_sources()
    payload = {
        "schema": "kgmu-mechanical-source-probe-v1",
        "semanticParsingPerformed": False,
        "timetablePage": TIMETABLE_PAGE,
        "sources": [workbook_dump(source) for source in discovered],
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    summary = [
        {
            "stream": source["stream"],
            "url": source["url"],
            "sha256": source["sha256"],
            "byteLength": source["byteLength"],
            "sheetNames": source["sheetNames"],
            "dimensions": [
                [sheet["title"], sheet["maxRow"], sheet["maxColumn"], len(sheet["mergedRanges"]), sheet["nonEmptyCellCount"]]
                for sheet in source["sheets"]
            ],
        }
        for source in payload["sources"]
    ]
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
