#!/usr/bin/env python3
"""Apply the confirmed course-specific PE interpretation for Pediatrics 531-537.

This is intentionally a postprocessing step over the course-specific explicit
manifest. It does not modify shared parser/core rules. The operator resolution
is pinned in pediatrics-531-537.operator-resolution.json.
"""
import hashlib
import io
import json
import re
import urllib.request
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
PERIOD = "2026-2027-semester-1"
FIXTURE_DIR = ROOT / "fixtures" / PERIOD
QA_DIR = ROOT / "qa" / PERIOD
SOURCE_PATH = FIXTURE_DIR / "pediatrics-531-537.source.json"
MANIFEST_PATH = FIXTURE_DIR / "pediatrics-531-537.decisions.json"
RESOLUTION_PATH = FIXTURE_DIR / "pediatrics-531-537.operator-resolution.json"
REVIEW_PATH = QA_DIR / "pediatrics-531-537.semantic-review.json"
GROUPS = ["531", "532", "533", "534", "535", "536", "537"]
DECISION_ID = "PED5-PE-WEEKDAY-RANGE-CONTRADICTION"
RAW_SCHEDULE = "Четверг с 04.09 по 18.12 14:30-16:00"
TIME_RE = re.compile(r"(?P<sh>\d{1,2})[:.](?P<sm>\d{2})\s*[-–—]\s*(?P<eh>\d{1,2})[:.](?P<em>\d{2})")


def read_json(path):
    return json.loads(path.read_text(encoding="utf-8"))


def norm(value):
    if value is None:
        return None
    return re.sub(r"\s+", " ", str(value)).strip()


def fetch(url):
    request = urllib.request.Request(url, headers={"User-Agent": "kgmu-calendar-pe-resolution/1.0"})
    with urllib.request.urlopen(request, timeout=45) as response:
        return response.read()


def parse_time(value):
    match = TIME_RE.search(norm(value) or "")
    if not match:
        raise SystemExit(f"PE time is not parseable: {value!r}")
    return (
        f"{int(match.group('sh')):02d}:{match.group('sm')}",
        f"{int(match.group('eh')):02d}:{match.group('em')}",
    )


def dates_for_weekday(start, end, weekday):
    current = start
    values = []
    while current <= end:
        if current.weekday() == weekday:
            values.append(current.isoformat())
        current += timedelta(days=1)
    return values


def mask(values, selected):
    indexes = {value: index for index, value in enumerate(values)}
    bits = 0
    for value in selected:
        if value not in indexes:
            raise SystemExit(f"resolved PE date outside source calendar: {value}")
        bits |= 1 << indexes[value]
    return format(bits, "x")


def join_location(base, address):
    base = norm(base)
    address = norm(address)
    if base and address:
        return f"{base}, {address}"
    return base or address or "Место не указано"


def main():
    source = read_json(SOURCE_PATH)
    manifest = read_json(MANIFEST_PATH)
    review = read_json(REVIEW_PATH)
    resolution = read_json(RESOLUTION_PATH)

    source_info = source["source"]
    if resolution["fixtureId"] != source["fixtureId"]:
        raise SystemExit("operator resolution fixture mismatch")
    if resolution["sourceSha256"] != source_info["sha256"] or manifest["sourceSha256"] != source_info["sha256"]:
        raise SystemExit("operator resolution is not pinned to the current source")
    if resolution["decisionId"] != DECISION_ID:
        raise SystemExit("unexpected operator decision id")
    if resolution["authority"] != "direct-user-confirmation":
        raise SystemExit("PE resolution must be an explicit operator confirmation")
    chosen = resolution["resolution"]
    if chosen["chosenInterpretation"] != "weekday-label" or chosen["weekdayLabel"] != "Четверг":
        raise SystemExit("this resolver only accepts the confirmed Thursday interpretation")
    if chosen["weekdayIndexMondayZero"] != 3 or chosen["appliesToGroups"] != GROUPS:
        raise SystemExit("unexpected PE weekday/group scope")

    data = fetch(source_info["url"])
    if hashlib.sha256(data).hexdigest() != source_info["sha256"]:
        raise SystemExit("official source SHA-256 changed")
    wb = load_workbook(io.BytesIO(data), data_only=False)
    ws = wb[wb.sheetnames[0]]
    discipline = norm(ws["C36"].value)
    raw_schedule = norm(ws["BT36"].value)
    if discipline != "Дисциплины по физической культуре и спорту":
        raise SystemExit(f"unexpected PE discipline: {discipline!r}")
    if raw_schedule != RAW_SCHEDULE or raw_schedule != resolution["rawSchedule"]:
        raise SystemExit(f"unexpected PE source schedule: {raw_schedule!r}")

    start = date.fromisoformat(chosen["rangeStart"])
    end = date.fromisoformat(chosen["rangeEnd"])
    dates = dates_for_weekday(start, end, chosen["weekdayIndexMondayZero"])
    if len(dates) != chosen["expectedOccurrenceCount"] or len(dates) != 15:
        raise SystemExit(f"unexpected resolved Thursday count: {len(dates)}")
    start_time, end_time = parse_time(raw_schedule)
    if f"{start_time}-{end_time}" != chosen["time"]:
        raise SystemExit("PE resolved time differs from pinned operator decision")

    if review.get("status") != "REVIEW_REQUIRED":
        raise SystemExit("base semantic review is expected to be REVIEW_REQUIRED before resolution")
    ambiguities = review.get("unresolvedAmbiguities", [])
    if len(ambiguities) != 1 or ambiguities[0].get("id") != DECISION_ID:
        raise SystemExit("expected the single PE ambiguity before applying resolution")
    if manifest.get("decisionCount") != 77 or len(manifest.get("decisions", [])) != 77:
        raise SystemExit("base manifest must contain 77 upper-grid decisions")

    if discipline in manifest["disciplineTable"]:
        discipline_index = manifest["disciplineTable"].index(discipline)
    else:
        discipline_index = len(manifest["disciplineTable"])
        manifest["disciplineTable"].append(discipline)

    location = join_location(ws["AN36"].value, ws["BL36"].value)
    if location in manifest["locationTable"]:
        location_index = manifest["locationTable"].index(location)
    else:
        location_index = len(manifest["locationTable"])
        manifest["locationTable"].append(location)

    assessment_label = norm(ws["S36"].value)
    if assessment_label:
        lowered = assessment_label.lower().replace("ё", "е")
        assessment_type = "exam" if "экзам" in lowered else "credit" if "зач" in lowered else "other"
        manifest.setdefault("assessmentMetadataByDisciplineIndex", {})[str(discipline_index)] = {
            "type": assessment_type,
            "label": assessment_label.lower(),
            "sourceRef": {"sourceId": source_info["sourceId"], "locator": f"{ws.title}!S36"},
        }

    pe_decision = [
        "BT36",
        format((1 << len(GROUPS)) - 1, "x"),
        mask(manifest["dateTable"], dates),
        start_time,
        end_time,
        discipline_index,
        0,
        location_index,
    ]
    manifest["decisions"].append(pe_decision)
    manifest["decisionCount"] = len(manifest["decisions"])
    manifest["logicalSourceCellCount"] = 78
    manifest["operatorResolutions"] = [{
        "decisionId": DECISION_ID,
        "resolutionFile": f"fixtures/{PERIOD}/pediatrics-531-537.operator-resolution.json",
        "chosenInterpretation": "weekday-label",
        "weekdayLabel": "Четверг",
        "dates": dates,
        "time": chosen["time"],
        "sourceLocator": resolution["sourceLocator"],
        "authority": resolution["authority"],
        "confirmedAt": resolution["confirmedAt"],
    }]

    review["status"] = "PASS"
    review["unresolvedAmbiguities"] = []
    review["resolvedAmbiguities"] = [{
        "id": DECISION_ID,
        "resolution": "weekday-label",
        "weekdayLabel": "Четверг",
        "dates": dates,
        "time": chosen["time"],
        "sourceRefs": [f"{ws.title}!C36", f"{ws.title}!BT36"],
        "authority": resolution["authority"],
        "confirmedAt": resolution["confirmedAt"],
        "resolutionFile": f"fixtures/{PERIOD}/pediatrics-531-537.operator-resolution.json",
    }]
    review["coverage"]["normalizedIndependentLowerScheduleCount"] = 1
    review["coverage"]["coveredEventBearingSourceBlockCount"] = 78
    review["coverage"]["resolvedIndependentLowerScheduleEventCount"] = len(dates) * len(GROUPS)
    review["normalizationSummary"] = {
        "draftScope": "all event-bearing source blocks after explicit operator resolution",
        "normalizedUpperGridEvents": review["coverage"]["upperGridNormalizedEventCount"],
        "normalizedIndependentLowerScheduleEvents": len(dates) * len(GROUPS),
        "groups": GROUPS,
        "excludedPendingReview": [],
        "sharedCoreChangeRequired": False,
        "publicationAllowed": False,
    }

    MANIFEST_PATH.write_text(json.dumps(manifest, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    REVIEW_PATH.write_text(json.dumps(review, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "decisionId": DECISION_ID,
        "resolution": "Thursday",
        "occurrenceCount": len(dates),
        "groups": len(GROUPS),
        "addedEvents": len(dates) * len(GROUPS),
        "manifestDecisionCount": manifest["decisionCount"],
        "reviewStatus": review["status"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
