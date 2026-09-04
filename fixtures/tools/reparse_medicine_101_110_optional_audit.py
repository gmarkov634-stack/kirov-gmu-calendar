#!/usr/bin/env python3
import hashlib
import io
import json
import re
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

from official_schedule_discovery import discover_schedule_link, fetch_url_bytes

ROOT = Path(__file__).resolve().parents[2]
SOURCE_META = ROOT / "fixtures/2026-2027-semester-1/medicine-101-110.source.json"
DECISIONS = ROOT / "fixtures/2026-2027-semester-1/medicine-101-110.decisions.json"
FACULTATIVES = ROOT / "fixtures/2026-2027-semester-1/medicine-101-110.facultatives.json"
DUMP_PATH = ROOT / "medicine-101-110-workbook-dump.json"
DISCOVERY_REPORT_PATH = ROOT / "medicine-101-110-source-discovery.json"
DISCOVERY_PAGE = "https://kirovgma.ru/lechebnyy-fakultet-raspisanie"
GROUP_LABEL = "101-110"
SEMESTER_LABEL = "первое полугодие"
SHEET_NAME = "1 леч.1"
EXPECTED_LOGICAL_CELLS = 145
OPTIONAL_RE = re.compile(r"факульт|электив|по выбор", re.IGNORECASE)
INTERVAL_RE = re.compile(r"(?<!\d)(\d{2}\.\d{2})-(\d{2}\.\d{2})(?!\d)")


def academic_date(value: str) -> date:
    day, month = map(int, value.split("."))
    year = 2027 if month == 1 else 2026
    return date(year, month, day)


def service_intervals(text: str):
    result = []
    for left, right in INTERVAL_RE.findall(text):
        start = academic_date(left)
        end = academic_date(right)
        if end < start:
            raise SystemExit(f"invalid service interval {left}-{right}")
        result.append((start, end))
    if len(result) != 19:
        raise SystemExit(f"expected 19 service-week intervals in A44, found {len(result)}")
    return result


def dates_for_weekday(intervals, weekday: int):
    result = []
    target = weekday - 1
    for start, end in intervals:
        current = start
        while current <= end:
            if current.weekday() == target:
                result.append(current.isoformat())
            current += timedelta(days=1)
    return sorted(set(result))


def inspect_sheet(workbook):
    if SHEET_NAME not in workbook.sheetnames:
        return None, [], []
    sheet = workbook[SHEET_NAME]
    full_cells = []
    full_sheet_optional = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value).strip()
            entry = {"coord": cell.coordinate, "value": text}
            full_cells.append(entry)
            if OPTIONAL_RE.search(text):
                full_sheet_optional.append(entry)
    return sheet, full_cells, full_sheet_optional


def main():
    meta = json.loads(SOURCE_META.read_text(encoding="utf-8"))
    manifest = json.loads(DECISIONS.read_text(encoding="utf-8"))
    facultatives = json.loads(FACULTATIVES.read_text(encoding="utf-8"))

    discovered = discover_schedule_link(
        page_url=DISCOVERY_PAGE,
        group_label=GROUP_LABEL,
        academic_year=meta["academicYear"],
        semester_label=SEMESTER_LABEL,
    )
    data = fetch_url_bytes(discovered.url, timeout=30)
    actual_sha = hashlib.sha256(data).hexdigest()
    workbook = load_workbook(io.BytesIO(data), data_only=False)
    sheet, full_cells, full_sheet_optional = inspect_sheet(workbook)

    current_workbook = {
        "sheetNames": workbook.sheetnames,
        "selectedSheet": SHEET_NAME if sheet is not None else None,
        "maxRow": sheet.max_row if sheet is not None else None,
        "maxColumn": sheet.max_column if sheet is not None else None,
        "mergedRangeCount": len(sheet.merged_cells.ranges) if sheet is not None else None,
        "nonEmptyCellCount": len(full_cells) if sheet is not None else None,
        "optionalTermMatchCount": len(full_sheet_optional) if sheet is not None else None,
    }
    fingerprint_matches = (
        actual_sha == meta["source"]["sha256"]
        and len(data) == meta["source"]["byteLength"]
    )
    discovery_report = {
        "status": "MATCH" if fingerprint_matches else "REVIEW_REQUIRED_SOURCE_CHANGED",
        "discoveryPage": DISCOVERY_PAGE,
        "groupLabel": GROUP_LABEL,
        "academicYear": meta["academicYear"],
        "semesterLabel": SEMESTER_LABEL,
        "discoveredLabel": discovered.label,
        "discoveredUrl": discovered.url,
        "frozenUrl": meta["source"]["url"],
        "discoveredSha256": actual_sha,
        "frozenSha256": meta["source"]["sha256"],
        "discoveredByteLength": len(data),
        "frozenByteLength": meta["source"]["byteLength"],
        "discoveredWorkbook": current_workbook,
        "frozenWorkbookExpectations": meta["workbookExpectations"],
    }
    DISCOVERY_REPORT_PATH.write_text(
        json.dumps(discovery_report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    dump = {
        "reviewStatus": discovery_report["status"],
        "sourceUrl": discovered.url,
        "sourceSha256": actual_sha,
        "byteLength": len(data),
        "sheetNames": workbook.sheetnames,
        "sheet": SHEET_NAME if sheet is not None else None,
        "maxRow": sheet.max_row if sheet is not None else None,
        "maxColumn": sheet.max_column if sheet is not None else None,
        "nonEmptyCells": full_cells,
        "mergedRanges": [str(rng) for rng in sheet.merged_cells.ranges] if sheet is not None else [],
        "optionalTermMatches": full_sheet_optional,
    }
    DUMP_PATH.write_text(json.dumps(dump, ensure_ascii=False, indent=2), encoding="utf-8")

    if not fingerprint_matches:
        raise SystemExit(
            "REVIEW_REQUIRED: official workbook fingerprint changed; structural dump preserved; "
            f"discoveredUrl={discovered.url} discoveredSha256={actual_sha} "
            f"frozenSha256={meta['source']['sha256']}"
        )

    if workbook.sheetnames != meta["workbookExpectations"]["sheetNames"]:
        raise SystemExit(f"sheet set changed: {workbook.sheetnames}")
    if sheet is None:
        raise SystemExit(f"expected worksheet is missing: {SHEET_NAME}")

    logical = []
    for row in sheet.iter_rows(min_row=9, max_row=42, min_col=2, max_col=11):
        for cell in row:
            if cell.value is not None:
                logical.append(cell.coordinate)
    if len(logical) != EXPECTED_LOGICAL_CELLS:
        raise SystemExit(f"logical source cell count changed: {len(logical)}")

    covered = set()
    for decision in manifest["decisions"]:
        locator = decision[0]
        match = re.match(r"([A-Z]+\d+)#s\d+$", locator)
        if not match:
            raise SystemExit(f"unexpected decision locator: {locator}")
        covered.add(match.group(1))
    missing = sorted(set(logical) - covered)
    extra = sorted(covered - set(logical))
    if missing or extra:
        raise SystemExit(f"main-table source coverage mismatch: missing={missing} extra={extra}")

    merged = {str(rng) for rng in sheet.merged_cells.ranges}
    if "A43:K43" not in merged:
        raise SystemExit("facultative scope merge A43:K43 is missing")
    if "A44:K44" not in merged:
        raise SystemExit("service-week merge A44:K44 is missing")

    a43 = str(sheet["A43"].value or "").strip()
    a44 = str(sheet["A44"].value or "").strip()
    if "ФАКУЛЬТАТИВЫ" not in a43.upper():
        raise SystemExit("A43 no longer contains the facultative block")
    if facultatives.get("schema") != "kgmu-medicine-facultatives-v1":
        raise SystemExit("unexpected facultative fixture schema")
    if facultatives.get("sourceSha256") != actual_sha:
        raise SystemExit("facultative fixture/source SHA mismatch")
    if facultatives.get("groupIds") != meta["expectedGroupIds"]:
        raise SystemExit("facultative group scope does not match groups 101-110")
    if facultatives.get("defaultSelected") is not False:
        raise SystemExit("facultatives must default to not selected")
    if facultatives.get("sourceLocator") != "A43" or facultatives.get("weekGridLocator") != "A44":
        raise SystemExit("facultative fixture must cite A43/A44")

    intervals = service_intervals(a44)
    items = facultatives.get("items") or []
    if len(items) != 5:
        raise SystemExit(f"expected five facultatives, found {len(items)}")

    expected_text = {
        "Актуальные вопросы биологии": ("понедельник", "16.50-18.20"),
        "Основы химии": ("вторник", "16.50-18.20"),
        "Физика": ("среда", "18.00-20.25"),
        "Математика": ("четверг", "17.20-19.45"),
        "Русский язык и культура речи": ("пятница", "16.50-18.20"),
    }
    counts = {}
    for item in items:
        discipline = item["discipline"]
        if discipline not in expected_text:
            raise SystemExit(f"unexpected facultative discipline: {discipline}")
        day_text, time_text = expected_text[discipline]
        if day_text not in a43.lower() or time_text not in a43:
            raise SystemExit(f"A43 no longer supports {discipline} {day_text} {time_text}")
        expected_dates = dates_for_weekday(intervals, item["weekday"])
        if item["dates"] != expected_dates:
            raise SystemExit(f"R90 date expansion mismatch for {discipline}: expected={expected_dates} actual={item['dates']}")
        counts[item["facultativeId"]] = len(item["dates"]) * len(facultatives["groupIds"])

    result = {
        "sourceUrl": discovered.url,
        "sourceSha256": actual_sha,
        "logicalSourceCellCount": len(logical),
        "coveredSourceCellCount": len(covered),
        "fullSheetOptionalTermMatches": full_sheet_optional,
        "facultativeSourceCell": "A43:K43",
        "serviceWeekSourceCell": "A44:K44",
        "facultativeCount": len(items),
        "facultativeEventCount": sum(counts.values()),
        "facultativeCountsById": counts,
        "decision": "PASS_CONFIRMED_FACULTATIVES_R90",
        "note": "A43 facultatives cover groups 101-110 by merged geometry and expand over every academic service week in A44 under operator-confirmed R90 semantics; default selection is none.",
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
