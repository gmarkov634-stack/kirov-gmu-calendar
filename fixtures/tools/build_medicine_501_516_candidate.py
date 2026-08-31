#!/usr/bin/env python3
"""Build the explicit semantic candidate for KGMU medicine course 5, groups 501-516.

The workbook remains the source of truth. The only source-specific operator decision
implemented here is the user's 2026-08-31 instruction to model the two physical-
education streams as mutually exclusive personalization options for every group,
without inferring a group-to-stream mapping.
"""
from __future__ import annotations

import hashlib
import io
import json
import re
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
PERIOD = "2026-2027-semester-1"
OUT = ROOT / "fixtures" / PERIOD
QA = ROOT / "qa" / PERIOD
SOURCE_URL = "https://kirovgma.ru/sites/default/files/files/2026/08/24/1078/5_lech-24-08-2026-13.xlsx"
SOURCE_PAGE = "https://kirovgma.ru/lechebnyy-fakultet-raspisanie"
SOURCE_SHA256 = "43ecb37de9db7ba69153c8514f62de0b058e51c2032e0ee320b117378a740c62"
SOURCE_BYTES = 31755
SHEET = "2026+-2027 осень 5 курс  Леч"
GROUPS = [str(value) for value in range(501, 517)]
RULES_VERSION = "kgmu-2026-08-30-v4"
MONTHS = {"Сентябрь": 9, "Октябрь": 10, "Ноябрь": 11, "Декабрь": 12, "Январь": 1}
PE_NAME = "Дисциплины по физической культуре и спорту"
PE_SELECTION_GROUP = "medicine-5-physical-education-stream-2026-s1"

UPPER_TO_REFERENCE_ROW = {
    "Психиатрия, мед. психология": 33,
    "Госпитальная терапия": 34,
    "Эндокринология": 35,
    "Педиатрия": 36,
    "Детские инфекции": 37,
    "Инфекционные болезни": 38,
    "Медицина катастроф": 39,
    "Акушерство и гинекология": 40,
    "Госпитальная хирургия": 41,
    "Мед. реабилитация, СМ": 42,
    "КПА": 43,
    "К П/Ф": 44,
}

ASSESSMENT_MAP = {
    "Экзамен": ("exam", "экзамен"),
    "Зачёт": ("credit", "зачет"),
    "Зачет": ("credit", "зачет"),
    "Зачёт с оценкой": ("graded_credit", "зачет с оценкой"),
    "Зачет с оценкой": ("graded_credit", "зачет с оценкой"),
}


def fetch() -> bytes:
    request = urllib.request.Request(SOURCE_URL, headers={"User-Agent": "kgmu-calendar-candidate-builder/1.0"})
    with urllib.request.urlopen(request, timeout=45) as response:
        data = response.read()
    actual = hashlib.sha256(data).hexdigest()
    if actual != SOURCE_SHA256:
        raise SystemExit(f"source sha mismatch: {actual}")
    if len(data) != SOURCE_BYTES:
        raise SystemExit(f"source byte length mismatch: {len(data)}")
    return data


def text(value) -> str | None:
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def parse_time(value: str | None) -> tuple[str, str] | None:
    if not value:
        return None
    match = re.search(r"(\d{1,2})[.:](\d{2})\s*[-–—]\s*(\d{1,2})[.:](\d{2})", value)
    if not match:
        return None
    return (f"{int(match.group(1)):02d}:{match.group(2)}", f"{int(match.group(3)):02d}:{match.group(4)}")


def mask_for_indexes(indexes: list[int]) -> str:
    value = 0
    for index in indexes:
        value |= 1 << index
    if value == 0:
        raise ValueError("empty mask")
    return format(value, "x")


def location_for(ws, row: int) -> str:
    base = text(ws[f"AN{row}"].value)
    address = text(ws[f"BL{row}"].value)
    parts = [part for part in (base, address) if part]
    if not parts:
        raise SystemExit(f"reference row {row} has no location")
    return ", ".join(parts)


def reference_discipline(ws, row: int) -> str:
    value = text(ws[f"C{row}"].value)
    if not value:
        raise SystemExit(f"reference row {row} has no discipline")
    return value


def assessment_for(ws, row: int, source_id: str) -> dict | None:
    raw = text(ws[f"S{row}"].value)
    if not raw:
        return None
    if raw not in ASSESSMENT_MAP:
        raise SystemExit(f"unrecognized assessment {raw!r} in S{row}")
    kind, label = ASSESSMENT_MAP[raw]
    return {"type": kind, "label": label, "sourceRef": {"sourceId": source_id, "locator": f"{SHEET}!S{row}"}}


def main() -> None:
    data = fetch()
    workbook = load_workbook(io.BytesIO(data), data_only=False)
    if workbook.sheetnames != [SHEET]:
        raise SystemExit(f"unexpected sheets: {workbook.sheetnames}")
    ws = workbook[SHEET]
    if (ws.max_row, ws.max_column) != (52, 125):
        raise SystemExit(f"unexpected dimensions: {(ws.max_row, ws.max_column)}")
    merged = list(ws.merged_cells.ranges)
    if len(merged) != 331:
        raise SystemExit(f"unexpected merged range count: {len(merged)}")
    non_empty = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
    if non_empty != 605:
        raise SystemExit(f"unexpected non-empty count: {non_empty}")

    month_by_col: dict[int, int] = {}
    for rng in merged:
        if rng.min_row == 10 and rng.max_row == 10:
            value = text(ws.cell(10, rng.min_col).value)
            if value in MONTHS:
                for col in range(rng.min_col, rng.max_col + 1):
                    month_by_col[col] = MONTHS[value]

    date_by_col: dict[int, str] = {}
    for col in range(3, ws.max_column + 1):
        month = month_by_col.get(col)
        raw_day = ws.cell(11, col).value
        if month is None or raw_day is None:
            continue
        day = int(raw_day)
        year = 2027 if month == 1 else 2026
        date_by_col[col] = date(year, month, day).isoformat()
    if not date_by_col:
        raise SystemExit("calendar grid dates not found")
    date_table = list(dict.fromkeys(date_by_col[col] for col in sorted(date_by_col)))
    date_index = {value: index for index, value in enumerate(date_table)}

    source_id = "medicine-course-5-2026-2027-s1"
    source = {
        "fixtureId": "medicine-501-516-2026-2027-semester-1",
        "universityId": "kirov-gmu",
        "programId": "medicine",
        "course": 5,
        "academicYear": "2026-2027",
        "academicPeriodId": PERIOD,
        "parserProfile": "cyclic",
        "parserRulesVersion": RULES_VERSION,
        "source": {
            "sourceId": source_id,
            "url": SOURCE_URL,
            "sha256": SOURCE_SHA256,
            "byteLength": SOURCE_BYTES,
            "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
        "expectedGroupIds": GROUPS,
        "workbookExpectations": {
            "sheetNames": [SHEET],
            "maxRow": 52,
            "maxColumn": 125,
            "mergedRangeCount": 331,
            "nonEmptyCellCount": 605,
        },
        "discovery": {
            "pageUrl": SOURCE_PAGE,
            "strategy": "html-xlsx-links",
            "verifiedAt": "2026-08-31",
        },
        "storagePolicy": {
            "repositoryStoresBinarySource": False,
            "productionSourceIsServerFetched": True,
            "immutableSourceStoredInObjectStorage": True,
        },
    }

    reference_rows = sorted(set(UPPER_TO_REFERENCE_ROW.values()))
    disciplines: list[str] = [reference_discipline(ws, row) for row in reference_rows]
    # Keep two distinct indexes with the exact same source discipline so each stream
    # can carry independent selection metadata without changing the calendar title.
    pe_index_1 = len(disciplines)
    disciplines.append(reference_discipline(ws, 45))
    pe_index_2 = len(disciplines)
    disciplines.append(reference_discipline(ws, 45))
    if disciplines[pe_index_1] != PE_NAME or disciplines[pe_index_2] != PE_NAME:
        raise SystemExit("physical-education discipline text changed in source")

    discipline_index_by_row = {row: index for index, row in enumerate(reference_rows)}
    locations = [location_for(ws, row) for row in reference_rows + [45]]
    location_table = list(dict.fromkeys(locations))
    location_index_by_row = {row: location_table.index(location_for(ws, row)) for row in reference_rows + [45]}

    assessment_metadata: dict[str, dict] = {}
    for row in reference_rows:
        item = assessment_for(ws, row, source_id)
        if item:
            assessment_metadata[str(discipline_index_by_row[row])] = item
    pe_assessment = assessment_for(ws, 45, source_id)
    if pe_assessment:
        assessment_metadata[str(pe_index_1)] = pe_assessment
        assessment_metadata[str(pe_index_2)] = pe_assessment

    group_row_by_id: dict[str, int] = {}
    for row in range(13, 29):
        group = text(ws.cell(row, 2).value)
        if group:
            group_row_by_id[group] = row
    if list(group_row_by_id) != GROUPS:
        raise SystemExit(f"unexpected group rows: {group_row_by_id}")

    block_ranges = [rng for rng in merged if 13 <= rng.min_row <= 28 and rng.min_row == rng.max_row and rng.min_col >= 3]
    decisions: list[list] = []
    service_blocks: list[dict] = []
    upper_blocks = 0
    starred_blocks = 0
    source_date_coverage = 0

    for group_index, group in enumerate(GROUPS):
        row = group_row_by_id[group]
        ranges = sorted((rng for rng in block_ranges if rng.min_row == row), key=lambda rng: rng.min_col)
        if not ranges:
            raise SystemExit(f"group {group} has no cycle blocks")
        for rng in ranges:
            raw = text(ws.cell(row, rng.min_col).value)
            if not raw:
                continue
            upper_blocks += 1
            dates = [date_by_col[col] for col in range(rng.min_col, rng.max_col + 1) if col in date_by_col]
            if not dates:
                raise SystemExit(f"block {rng} has no calendar dates")
            source_date_coverage += len(dates)
            locator = f"{get_column_letter(rng.min_col)}{row}"
            if raw.casefold() == "экзамен":
                service_blocks.append({"groupId": group, "locator": locator, "range": str(rng), "dates": dates, "rule": "C14"})
                continue

            starred = raw.endswith("*")
            key = raw[:-1].strip() if starred else raw
            if key not in UPPER_TO_REFERENCE_ROW:
                raise SystemExit(f"unmapped cycle label {raw!r} at {locator}")
            ref_row = UPPER_TO_REFERENCE_ROW[key]
            discipline_index = discipline_index_by_row[ref_row]
            group_mask = mask_for_indexes([group_index])
            first_time = parse_time(text(ws[f"BT{ref_row}"].value))
            second_time = parse_time(text(ws[f"BX{ref_row}"].value))
            if first_time is None:
                # Some source rows contain a single explicit shift in BX only.
                first_time = second_time
            if first_time is None:
                raise SystemExit(f"reference row {ref_row} has no usable time")
            location_index = location_index_by_row[ref_row]

            if starred:
                starred_blocks += 1
                if second_time is None:
                    raise SystemExit(f"starred block {locator} requires second-shift time in row {ref_row}")
                decisions.append([
                    f"{locator}#first", group_mask, mask_for_indexes([date_index[dates[0]]]),
                    second_time[0], second_time[1], discipline_index, 0, location_index,
                ])
                if len(dates) > 1:
                    decisions.append([
                        f"{locator}#rest", group_mask,
                        mask_for_indexes([date_index[value] for value in dates[1:]]),
                        first_time[0], first_time[1], discipline_index, 0, location_index,
                    ])
            else:
                decisions.append([
                    locator, group_mask, mask_for_indexes([date_index[value] for value in dates]),
                    first_time[0], first_time[1], discipline_index, 0, location_index,
                ])

    pe_1_raw = text(ws["BT45"].value)
    pe_2_raw = text(ws["BX45"].value)
    pe_1_time = parse_time(pe_1_raw)
    pe_2_time = parse_time(pe_2_raw)
    if pe_1_time != ("13:30", "15:00") or pe_2_time != ("15:10", "16:40"):
        raise SystemExit(f"unexpected PE stream times: {pe_1_time} / {pe_2_time}")
    if not pe_1_raw or "1 поток" not in pe_1_raw or "понедельник" not in pe_1_raw or "07.09-21.12" not in pe_1_raw:
        raise SystemExit("BT45 PE stream 1 rule changed")
    if not pe_2_raw or "2 поток" not in pe_2_raw or "понедельник" not in pe_2_raw or "07.09-21.12" not in pe_2_raw:
        raise SystemExit("BX45 PE stream 2 rule changed")

    pe_dates: list[str] = []
    cursor = date(2026, 9, 7)
    end = date(2026, 12, 21)
    while cursor <= end:
        if cursor.weekday() == 0:
            value = cursor.isoformat()
            if value not in date_index:
                raise SystemExit(f"PE date {value} is absent from source calendar grid")
            pe_dates.append(value)
        cursor += timedelta(days=1)
    if len(pe_dates) != 16:
        raise SystemExit(f"expected 16 PE Mondays, got {len(pe_dates)}")

    all_groups_mask = mask_for_indexes(list(range(len(GROUPS))))
    pe_date_mask = mask_for_indexes([date_index[value] for value in pe_dates])
    pe_location = location_index_by_row[45]
    decisions.append(["BT45#stream-1", all_groups_mask, pe_date_mask, pe_1_time[0], pe_1_time[1], pe_index_1, 0, pe_location])
    decisions.append(["BX45#stream-2", all_groups_mask, pe_date_mask, pe_2_time[0], pe_2_time[1], pe_index_2, 0, pe_location])

    selection_metadata = {
        str(pe_index_1): {
            "selectionGroupId": PE_SELECTION_GROUP,
            "selectionOptionId": "stream-1",
            "selectionGroupLabel": "Поток физкультуры",
            "selectionOptionLabel": "1 поток",
        },
        str(pe_index_2): {
            "selectionGroupId": PE_SELECTION_GROUP,
            "selectionOptionId": "stream-2",
            "selectionGroupLabel": "Поток физкультуры",
            "selectionOptionLabel": "2 поток",
        },
    }

    logical_source_cells = upper_blocks + 2
    manifest = {
        "schema": "kgmu-explicit-semantic-decisions-v3",
        "fixtureId": source["fixtureId"],
        "sourceSha256": SOURCE_SHA256,
        "parserRulesVersion": RULES_VERSION,
        "sheetName": SHEET,
        "semanticDecisionMode": "operator-authored-explicit",
        "logicalSourceCellCount": logical_source_cells,
        "decisionCount": len(decisions),
        "dateTable": date_table,
        "disciplineTable": disciplines,
        "locationTable": location_table,
        "assessmentMetadataByDisciplineIndex": assessment_metadata,
        "selectionMetadataByDisciplineIndex": selection_metadata,
        "groupTable": GROUPS,
        "lessonTypeTable": ["practice"],
        "tupleFields": ["locator", "groupMaskHex", "dateMaskHex", "startTime", "endTime", "disciplineIndex", "lessonTypeIndex", "locationIndex"],
        "decisions": decisions,
    }

    review = {
        "fixtureId": source["fixtureId"],
        "status": "resolved-by-user-source-specific-decision",
        "parserProfile": "cyclic",
        "sourceSha256": SOURCE_SHA256,
        "rules": ["G01", "G02", "G04", "G06", "G10", "G11", "G12", "G13", "G14", "G15", "G16", "G17", "G18", "G19", "G20", "G21", "C01", "C02", "C12", "C13", "C14"],
        "sourceSpecificDecisions": [{
            "decisionId": "M501-PE-STREAM-PERSONALIZATION",
            "sourceLocators": [f"{SHEET}!C45", f"{SHEET}!BT45", f"{SHEET}!BX45"],
            "decision": "Model physical education as a user-selectable stream for every group 501-516. Preserve both source schedules as selection alternatives; do not infer group-to-stream membership.",
            "selectionGroupId": PE_SELECTION_GROUP,
            "options": [
                {"selectionOptionId": "stream-1", "label": "1 поток", "schedule": "понедельник 07.09-21.12 13:30-15:00"},
                {"selectionOptionId": "stream-2", "label": "2 поток", "schedule": "понедельник 07.09-21.12 15:10-16:40"},
            ],
            "confirmedBy": "user",
            "confirmedAt": "2026-08-31",
        }],
        "coverage": {
            "upperCycleBlockCount": upper_blocks,
            "serviceExamBlockCount": len(service_blocks),
            "independentPeScheduleCount": 2,
            "logicalSourceCellCount": logical_source_cells,
            "coveredSourceCellCount": logical_source_cells,
            "sourceCalendarDateOccurrencesInUpperBlocks": source_date_coverage,
            "starredCycleBlockCount": starred_blocks,
        },
        "serviceBlocks": service_blocks,
        "unresolvedAmbiguities": [],
    }

    OUT.mkdir(parents=True, exist_ok=True)
    QA.mkdir(parents=True, exist_ok=True)
    (OUT / "medicine-501-516.source.json").write_text(json.dumps(source, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (OUT / "medicine-501-516.decisions.json").write_text(json.dumps(manifest, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    (QA / "medicine-501-516.semantic-review.json").write_text(json.dumps(review, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "sourceSha256": SOURCE_SHA256,
        "groups": GROUPS,
        "dateTableCount": len(date_table),
        "upperCycleBlockCount": upper_blocks,
        "serviceExamBlockCount": len(service_blocks),
        "starredCycleBlockCount": starred_blocks,
        "peDates": pe_dates,
        "logicalSourceCellCount": logical_source_cells,
        "decisionCount": len(decisions),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
