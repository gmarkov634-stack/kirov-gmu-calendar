#!/usr/bin/env python3
"""Mechanical XLSX source probe for KGMU Pediatrics course 4, groups 431-436."""
import hashlib
import html
import io
import json
import re
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "fixtures/2026-2027-semester-1/pediatrics-431-436.source-probe.json"
TIMETABLE_PAGE = "https://kirovgma.ru/raspisanie-pediatricheskiy-fakultet"
ALLOWED_PREFIX = "/sites/default/files/files/"
SOURCE_PATTERN = re.compile(r"4_ped[^\"'<>\s]*\.xlsx", re.IGNORECASE)
CYCLE_ROWS = range(13, 19)


def request_bytes(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "kgmu-calendar-source-probe/1.0"})
    with urllib.request.urlopen(request, timeout=45) as response:
        return response.read()


def discover_source() -> dict:
    page_bytes = request_bytes(TIMETABLE_PAGE)
    page_text = html.unescape(page_bytes.decode("utf-8", errors="replace"))
    hrefs = re.findall(r"href\s*=\s*[\"']([^\"']+)[\"']", page_text, flags=re.IGNORECASE)
    matches = []
    for href in hrefs:
        decoded = urllib.parse.unquote(href)
        if SOURCE_PATTERN.search(decoded):
            absolute = urllib.parse.urljoin(TIMETABLE_PAGE, href)
            parsed = urllib.parse.urlparse(absolute)
            if parsed.netloc != "kirovgma.ru" or not parsed.path.startswith(ALLOWED_PREFIX):
                raise SystemExit(f"discovered source outside allowed KGMU path: {absolute}")
            matches.append(absolute)
    matches = sorted(set(matches))
    if len(matches) != 1:
        raise SystemExit(f"expected exactly one current XLSX for pediatrics 431-436, found {matches}")
    return {"program": "pediatrics", "course": 4, "groups": [str(value) for value in range(431, 437)], "url": matches[0]}


def json_scalar(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def color_dump(color) -> dict | None:
    if color is None:
        return None
    result = {"type": json_scalar(color.type)}
    for key in ("rgb", "indexed", "theme", "tint"):
        value = getattr(color, key, None)
        if value is not None:
            result[key] = json_scalar(value)
    return result


def border_side_dump(side) -> dict:
    return {"style": json_scalar(side.style), "color": color_dump(side.color)}


def visual_dump(cell) -> dict:
    return {
        "styleId": json_scalar(getattr(cell, "style_id", None)),
        "fill": {
            "type": json_scalar(cell.fill.fill_type),
            "fgColor": color_dump(cell.fill.fgColor),
            "bgColor": color_dump(cell.fill.bgColor),
        },
        "font": {
            "bold": json_scalar(cell.font.bold),
            "italic": json_scalar(cell.font.italic),
            "underline": json_scalar(cell.font.underline),
            "color": color_dump(cell.font.color),
        },
        "border": {
            "left": border_side_dump(cell.border.left),
            "right": border_side_dump(cell.border.right),
            "top": border_side_dump(cell.border.top),
            "bottom": border_side_dump(cell.border.bottom),
        },
        "alignment": {
            "horizontal": json_scalar(cell.alignment.horizontal),
            "vertical": json_scalar(cell.alignment.vertical),
            "textRotation": json_scalar(cell.alignment.textRotation),
            "wrapText": json_scalar(cell.alignment.wrapText),
        },
        "numberFormat": json_scalar(cell.number_format),
    }


def cycle_visual_evidence(sheet) -> list[dict]:
    evidence = []
    for merged in sorted(sheet.merged_cells.ranges, key=lambda item: (item.min_row, item.min_col, item.max_row, item.max_col)):
        if merged.min_row != merged.max_row or merged.min_row not in CYCLE_ROWS:
            continue
        anchor = sheet.cell(merged.min_row, merged.min_col)
        value = None if anchor.value is None else str(anchor.value).strip()
        if not value:
            continue
        visual_cells = []
        for column in range(merged.min_col, merged.max_col + 1):
            cell = sheet.cell(merged.min_row, column)
            visual_cells.append({"coord": cell.coordinate, **visual_dump(cell)})
        evidence.append({
            "range": str(merged),
            "anchor": anchor.coordinate,
            "value": value,
            "visualCells": visual_cells,
        })
    return evidence


def workbook_dump(source: dict) -> dict:
    data = request_bytes(source["url"])
    if not data.startswith(b"PK"):
        raise SystemExit(f"source is not XLSX/ZIP: {source['url']}")
    workbook = load_workbook(io.BytesIO(data), data_only=False)
    sheets = []
    for sheet in workbook.worksheets:
        non_empty = []
        comments = []
        formulas = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.comment is not None:
                    comments.append({"coord": cell.coordinate, "text": cell.comment.text, "author": cell.comment.author})
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({"coord": cell.coordinate, "formula": cell.value})
                if cell.value is None:
                    continue
                entry = {"coord": cell.coordinate, "value": str(cell.value).strip()}
                if cell.hyperlink is not None:
                    entry["hyperlink"] = {"target": cell.hyperlink.target, "location": cell.hyperlink.location, "display": cell.hyperlink.display}
                non_empty.append(entry)
        validations = []
        if sheet.data_validations is not None:
            for validation in sheet.data_validations.dataValidation:
                validations.append({
                    "sqref": str(validation.sqref),
                    "type": json_scalar(validation.type),
                    "formula1": json_scalar(validation.formula1),
                    "formula2": json_scalar(validation.formula2),
                })
        hidden_rows = [index for index, dimension in sheet.row_dimensions.items() if dimension.hidden]
        hidden_columns = [key for key, dimension in sheet.column_dimensions.items() if dimension.hidden]
        sheets.append({
            "title": sheet.title,
            "state": sheet.sheet_state,
            "maxRow": sheet.max_row,
            "maxColumn": sheet.max_column,
            "mergedRanges": [str(rng) for rng in sheet.merged_cells.ranges],
            "nonEmptyCellCount": len(non_empty),
            "nonEmptyCells": non_empty,
            "structuralEvidence": {
                "hiddenRows": hidden_rows,
                "hiddenColumns": hidden_columns,
                "comments": comments,
                "formulas": formulas,
                "dataValidations": validations,
                "imageCount": len(sheet._images),
                "chartCount": len(sheet._charts),
                "cycleMergedBlockVisuals": cycle_visual_evidence(sheet),
                "timeCells": {
                    coord: {
                        "value": None if sheet[coord].value is None else str(sheet[coord].value).strip(),
                        **visual_dump(sheet[coord]),
                    }
                    for coord in ("CE35", "CE36")
                },
            },
        })
    defined_names = []
    for name, definition in workbook.defined_names.items():
        defined_names.append({"name": name, "attrText": json_scalar(definition.attr_text), "hidden": json_scalar(definition.hidden)})
    return {
        **source,
        "sha256": hashlib.sha256(data).hexdigest(),
        "byteLength": len(data),
        "sheetNames": workbook.sheetnames,
        "definedNames": defined_names,
        "sheets": sheets,
    }


def main() -> None:
    payload = {"schema": "kgmu-mechanical-source-probe-v2", "semanticParsingPerformed": False, "timetablePage": TIMETABLE_PAGE, "source": workbook_dump(discover_source())}
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    source = payload["source"]
    print(json.dumps({"program": source["program"], "course": source["course"], "groups": source["groups"], "url": source["url"], "sha256": source["sha256"], "byteLength": source["byteLength"], "sheetNames": source["sheetNames"], "dimensions": [[s["title"], s["maxRow"], s["maxColumn"], len(s["mergedRanges"]), s["nonEmptyCellCount"]] for s in source["sheets"]], "structuralEvidence": [[s["title"], len(s["structuralEvidence"]["hiddenRows"]), len(s["structuralEvidence"]["hiddenColumns"]), len(s["structuralEvidence"]["comments"]), len(s["structuralEvidence"]["formulas"]), len(s["structuralEvidence"]["dataValidations"]), s["structuralEvidence"]["imageCount"], s["structuralEvidence"]["chartCount"]] for s in source["sheets"]]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
