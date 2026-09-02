#!/usr/bin/env python3
"""Build the deterministic normalized draft for KGMU Dentistry course 5.

Course-specific only. The script validates the pinned official XLSX, derives
calendar dates from the workbook, expands the merged rotation blocks for
591-594, splits the source-explicit Medicine of Disasters / Physical Training
combined block by lower-table times, and adds the source-explicit Friday PE
elective recurrence. It never opens the production DB or publishes anything.
"""
from __future__ import annotations

import argparse
import hashlib
import io
import json
import re
import urllib.request
from collections import Counter, defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

ROOT = Path(__file__).resolve().parents[2]
PERIOD = "2026-2027-semester-1"
FIXTURE = ROOT / "fixtures" / PERIOD / "dentistry-591-594.source.json"
JOB = ROOT / "fixtures" / PERIOD / "dentistry-591-594.parsing-job.json"
OUT = ROOT / "fixtures" / PERIOD / "normalized" / "dentistry-591-594.normalized.compact.json"
SHEET = "2026-2027 осень 5 курс Стом"
EXPECTED_SHA = "0c8b13b7e4dc409eaec551f8d4720d77dee88d76e8e7e89e4efcfe2aeed42109"
EXPECTED_RULES = "kgmu-2026-08-27-v3"
EXPECTED_COURSE_RULES = "dentistry-591-594-v1"
GROUP_ROW = {"591": 15, "592": 16, "593": 17, "594": 18}
MONTH_BY_COL = [
    ("C", "AB", 9, 2026),
    ("AC", "BC", 10, 2026),
    ("BD", "CA", 11, 2026),
    ("CB", "DA", 12, 2026),
    ("DB", "DT", 1, 2027),
]
ACADEMIC_LAST_COL = column_index_from_string("CV")
SERVICE_FIRST_COL = column_index_from_string("CW")
SERVICE_LABELS = {"экзамены", "практика", "каникулы"}

DENTAL_LOCATION = "Консультативно-диагностическое отделение клиники Кировского ГМУ, ул. Никитская, 161"
MEDCAT_LOCATION = "Кировский ГМУ, ул. Красноармейская, 35"
PE_LOCATION = "Кировский ГМУ, учебный корпус № 3 Физкультурно-оздоровительный комплекс, ул. Владимирская, 112"

META = {
    "Детская стоматология": {"row": 22, "time": ("08:00", "11:55"), "location": DENTAL_LOCATION, "assessment": "exam"},
    "Челюстно-лицевая хирургия": {"row": 23, "time": ("08:00", "11:55"), "location": DENTAL_LOCATION, "assessment": None},
    "Заболевания слизистой оболочки полости рта (модуль)": {"row": 24, "time": ("08:00", "11:55"), "location": DENTAL_LOCATION, "assessment": "credit"},
    "Ортодонтия и детское протезирование": {"row": 25, "time": ("08:00", "11:55"), "location": DENTAL_LOCATION, "assessment": "exam"},
    "Комплексное зубопротезирование и имплантология": {"row": 26, "time": ("08:00", "11:55"), "location": DENTAL_LOCATION, "assessment": None},
    "Пародонтология": {"row": 27, "time": ("08:00", "11:55"), "location": DENTAL_LOCATION, "assessment": "credit"},
    "Костнопластические материалы и технологии": {"row": 28, "time": None, "location": DENTAL_LOCATION, "assessment": "credit"},
    "Медицина катастроф": {"row": 29, "time": ("08:30", "11:35"), "location": MEDCAT_LOCATION, "assessment": "credit"},
    "Физическая подготовка": {"row": 30, "time": ("13:00", "14:30"), "location": PE_LOCATION, "assessment": "credit"},
    "Элективные дисциплины по физической культуре и спорту": {"row": 31, "time": ("14:30", "16:00"), "location": PE_LOCATION, "assessment": None},
}

ALIASES = {
    "детская стоматология": "Детская стоматология",
    "челюстно-лицевая хирургия": "Челюстно-лицевая хирургия",
    "ортодонтия и детское протезирование": "Ортодонтия и детское протезирование",
    "пародонтология": "Пародонтология",
    "костнопластические материалы и технологии": "Костнопластические материалы и технологии",
    "кзп": "Комплексное зубопротезирование и имплантология",
}

WEEKDAY_RU = {0: "пн", 1: "вт", 2: "ср", 3: "чт", 4: "пт", 5: "сб", 6: "вс"}


def compact(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\\n", " ")).strip()


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def fetch(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "kgmu-calendar-dentistry-591-594/1.0"})
    with urllib.request.urlopen(request, timeout=45) as response:
        return response.read()


def canonical(raw: str):
    text = compact(raw)
    key = text.casefold()
    if key in ALIASES:
        return ALIASES[key]
    if "медицина катастроф" in key and "физическая подготовка" in key:
        return "__COMBINED__"
    if "комплексное зубопротез" in key and "имплантолог" in key:
        return "Комплексное зубопротезирование и имплантология"
    if "заболевания слизистой" in key and "оболочки" in key:
        return "Заболевания слизистой оболочки полости рта (модуль)"
    return None


def parse_time(value: str):
    match = re.search(r"(\d{1,2})[.:](\d{2})\s*[-–—]\s*(\d{1,2})[.:](\d{2})", value)
    if not match:
        raise SystemExit(f"cannot parse time from {value!r}")
    return f"{int(match.group(1)):02d}:{match.group(2)}", f"{int(match.group(3)):02d}:{match.group(4)}"


def col_month(col: int):
    for first, last, month, year in MONTH_BY_COL:
        if column_index_from_string(first) <= col <= column_index_from_string(last):
            return year, month
    raise SystemExit(f"date column {col} outside configured month bands")


def date_for_col(ws, col: int):
    raw_day = ws.cell(13, col).value
    if raw_day in (None, ""):
        raise SystemExit(f"missing day in row 13, column {col}")
    year, month = col_month(col)
    result = date(year, month, int(raw_day))
    source_weekday = compact(ws.cell(14, col).value).casefold()
    if source_weekday != WEEKDAY_RU[result.weekday()]:
        raise SystemExit(f"weekday mismatch for {result}: source={source_weekday!r}")
    return result


def anchor_value(ws, rng):
    return ws.cell(rng.min_row, rng.min_col).value


def source_range_for_cell(ws, row: int, col: int):
    matches = [rng for rng in ws.merged_cells.ranges if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col]
    if len(matches) != 1:
        raise SystemExit(f"expected one merged range at row={row} col={col}, got {[str(r) for r in matches]}")
    return matches[0]


def metadata_checks(ws):
    expected = {
        22: ("Детская стоматология", "ул.Никитская, 161", "8.00-11.55"),
        23: ("Челюстно-лицевая хирургия", "ул.Никитская, 161", "8.00-11.55"),
        24: ("Заболевания слизистой оболочки полости рта", "ул.Никитская, 161", "8.00-11.55"),
        25: ("Ортодонтия и детское протезирование", "ул.Никитская, 161", "8.00-11.55"),
        26: ("Комплексное зубопротезирование и имплантология", "ул.Никитская, 161", "8.00-11.55"),
        27: ("Пародонтология", "ул.Никитская, 161", "8.00-11.55"),
        28: ("Костнопластические материалы и технологии", "ул.Никитская, 161", "гр. 591, 593"),
        29: ("Медицина катастроф", "ул. Красноармейская, 35", "8:30-11:35"),
        30: ("Физическая подготовка", "ул. Владимирская, 112", "13:00-14:30"),
        31: ("Элективные дисциплины по физической культуре и спорту", "ул. Владимирская, 112", "04.09"),
    }
    for row, (discipline, address, time_token) in expected.items():
        if discipline.casefold() not in compact(ws.cell(row, 3).value).casefold():
            raise SystemExit(f"lower-reference discipline mismatch at C{row}")
        if address.casefold() not in compact(ws.cell(row, 75).value).casefold():  # BW
            raise SystemExit(f"lower-reference address mismatch at BW{row}")
        line = " ".join(compact(ws.cell(row, c).value) for c in range(83, 91))  # CE:CL merged variants
        if time_token.casefold() not in line.casefold():
            raise SystemExit(f"lower-reference time mismatch at row {row}: {line!r}")
    # Special second time for costoplastic materials.
    second = " ".join(compact(ws.cell(28, c).value) for c in range(87, 91))  # CI:CL
    if "12.30-16.25" not in second or "592, 594" not in second:
        raise SystemExit("missing source-explicit 592/594 second shift in row 28")


def event_tuple(group: str, day: date, discipline: str, locator: str):
    meta = META[discipline]
    if discipline == "Костнопластические материалы и технологии":
        start, end = ("08:00", "11:55") if group in {"591", "593"} else ("12:30", "16:25")
    else:
        start, end = meta["time"]
    signature = "|".join([
        group, day.isoformat(), start, end, discipline, "practice", meta["location"], meta["assessment"] or "", locator
    ])
    event_id = "kgmu-" + hashlib.sha256(signature.encode("utf-8")).hexdigest()[:24]
    return [event_id, group, day.isoformat(), start, end, discipline, "practice", meta["location"], meta["assessment"], locator]


def build_events(ws):
    events = []
    classified_ranges = []
    for group, row in GROUP_ROW.items():
        seen_ranges = set()
        covered = set()
        for col in range(3, ACADEMIC_LAST_COL + 1):
            rng = source_range_for_cell(ws, row, col)
            key = str(rng)
            covered.add(col)
            if key in seen_ranges:
                continue
            seen_ranges.add(key)
            raw = compact(anchor_value(ws, rng))
            discipline = canonical(raw)
            if discipline is None:
                raise SystemExit(f"unclassified academic merged range {key}: {raw!r}")
            if rng.min_row != row or rng.max_row != row:
                raise SystemExit(f"academic range unexpectedly spans groups: {key}")
            classified_ranges.append({"groupId": group, "range": key, "raw": raw, "classification": discipline})
            for event_col in range(rng.min_col, rng.max_col + 1):
                day = date_for_col(ws, event_col)
                if discipline == "__COMBINED__":
                    events.append(event_tuple(group, day, "Медицина катастроф", key + "#medcat"))
                    events.append(event_tuple(group, day, "Физическая подготовка", key + "#physical"))
                else:
                    events.append(event_tuple(group, day, discipline, key))
        if covered != set(range(3, ACADEMIC_LAST_COL + 1)):
            raise SystemExit(f"incomplete academic date coverage for group {group}")

        # Service periods must be source-labelled and must not generate events.
        for col in range(SERVICE_FIRST_COL, column_index_from_string("DT") + 1):
            rng = source_range_for_cell(ws, row, col)
            raw = compact(anchor_value(ws, rng)).casefold()
            if raw not in SERVICE_LABELS:
                raise SystemExit(f"unclassified service range {rng}: {raw!r}")

    # Lower reference row 31 is an explicit all-group Friday recurrence.
    recurrence = compact(ws["CE31"].value)
    if "пятница" not in recurrence.casefold() or "04.09" not in recurrence or "18.12" not in recurrence:
        raise SystemExit(f"unexpected elective recurrence source: {recurrence!r}")
    start_time, end_time = parse_time(recurrence)
    if (start_time, end_time) != META["Элективные дисциплины по физической культуре и спорту"]["time"]:
        raise SystemExit("elective recurrence time differs from configured source interpretation")
    current = date(2026, 9, 4)
    while current <= date(2026, 12, 18):
        if current.weekday() == 4:
            for group in GROUP_ROW:
                events.append(event_tuple(group, current, "Элективные дисциплины по физической культуре и спорту", "CE31:CL31#friday"))
        current += timedelta(days=1)

    events.sort(key=lambda e: (e[1], e[2], e[3], e[5], e[9]))
    return events, classified_ranges


def overlap_pairs(events):
    def minutes(value):
        hour, minute = map(int, value.split(":"))
        return hour * 60 + minute
    by_day = defaultdict(list)
    for event in events:
        by_day[(event[1], event[2])].append(event)
    overlaps = []
    for (group, day), items in by_day.items():
        for index, first in enumerate(items):
            for second in items[index + 1:]:
                if minutes(first[3]) < minutes(second[4]) and minutes(second[3]) < minutes(first[4]):
                    overlaps.append({"groupId": group, "date": day, "a": first[5], "b": second[5], "sourceBacked": True})
    return overlaps


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--write", action="store_true")
    args = parser.parse_args()

    fixture = read_json(FIXTURE)
    job = read_json(JOB)
    if fixture["parserRulesVersion"] != EXPECTED_RULES or fixture["courseParserRulesVersion"] != EXPECTED_COURSE_RULES:
        raise SystemExit("parser rule version mismatch")
    if job["parserRulesVersion"] != EXPECTED_RULES or job["expectedGroupIds"] != fixture["expectedGroupIds"]:
        raise SystemExit("ParsingJob is not bound to the source fixture")
    if job["sourceObjectKey"] != fixture["source"]["objectKey"]:
        raise SystemExit("ParsingJob sourceObjectKey mismatch")

    data = fetch(fixture["source"]["url"])
    actual_sha = hashlib.sha256(data).hexdigest()
    if actual_sha != EXPECTED_SHA or actual_sha != fixture["source"]["sha256"]:
        raise SystemExit(f"official source SHA changed: {actual_sha}")
    if len(data) != fixture["source"]["byteLength"]:
        raise SystemExit("official source byte length changed")

    wb = load_workbook(io.BytesIO(data), data_only=False)
    if wb.sheetnames != [SHEET]:
        raise SystemExit(f"unexpected sheets {wb.sheetnames}")
    ws = wb[SHEET]
    if ws.max_row != 36 or ws.max_column != 124 or len(ws.merged_cells.ranges) != 127:
        raise SystemExit("workbook geometry changed")
    non_empty = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
    if non_empty != 407:
        raise SystemExit(f"non-empty cell count changed: {non_empty}")
    if compact(ws["B15"].value) != "591" or compact(ws["B18"].value) != "594":
        raise SystemExit("group rows changed")
    metadata_checks(ws)

    events, classified = build_events(ws)
    counts = dict(sorted(Counter(event[1] for event in events).items()))
    duplicate_signatures = len(events) - len({tuple(event[1:]) for event in events})
    if counts != {"591": 123, "592": 123, "593": 123, "594": 123} or len(events) != 492:
        raise SystemExit(f"unexpected event counts: {counts}, total={len(events)}")
    if duplicate_signatures:
        raise SystemExit(f"duplicate normalized signatures: {duplicate_signatures}")
    digest = "sha256:" + hashlib.sha256(json.dumps(events, ensure_ascii=False, separators=(",", ":")).encode("utf-8")).hexdigest()
    overlaps = overlap_pairs(events)
    if len(overlaps) != 3:
        raise SystemExit(f"unexpected source-backed overlap count: {len(overlaps)}")

    payload = {
        "schema": "kgmu-normalized-compact-v1",
        "fixtureId": fixture["fixtureId"],
        "parsingJobId": job["jobId"],
        "sourceArtifactId": fixture["source"]["sourceArtifactId"],
        "sourceSha256": actual_sha,
        "parserProfile": fixture["parserProfile"],
        "parserRulesVersion": EXPECTED_RULES,
        "courseParserRulesVersion": EXPECTED_COURSE_RULES,
        "encoding": "normalized-event-tuples-with-assessment-v1",
        "tupleFields": ["eventId", "groupId", "date", "startTime", "endTime", "discipline", "lessonType", "location", "assessmentType", "sourceLocator"],
        "constants": {"universityId": "kirov-gmu", "academicPeriodId": PERIOD, "timeSemantics": "floating", "sourceId": "dentistry", "teacher": None},
        "eventCount": len(events),
        "groupEventCounts": counts,
        "candidateDigest": digest,
        "duplicateSignatures": duplicate_signatures,
        "sourceBackedOverlapCount": len(overlaps),
        "events": events,
    }
    if args.write:
        OUT.parent.mkdir(parents=True, exist_ok=True)
        OUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")

    print(json.dumps({
        "status": "PASS",
        "sourceSha256": actual_sha,
        "eventCount": len(events),
        "groupEventCounts": counts,
        "candidateDigest": digest,
        "classifiedAcademicRanges": len(classified),
        "sourceBackedOverlapCount": len(overlaps),
        "duplicateSignatures": duplicate_signatures,
        "output": str(OUT.relative_to(ROOT)) if args.write else None,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
