#!/usr/bin/env python3
"""Mechanical XLSX source probe for KGMU Dentistry course 4.

This helper intentionally performs no semantic parsing. It discovers the current
official XLSX link from the KGMU dentistry timetable page, then records checksum,
workbook geometry, merged ranges, non-empty cell values, discovered group scope
and hyperlink targets for operator/ChatGPT review under canonical G/R/C/S rules.
"""
import hashlib
import html
import io
import json
import re
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "dentistry-494-source-probe.json"
XLSX_OUT = ROOT / "dentistry-494-source.xlsx"
TIMETABLE_PAGE = "https://kirovgma.ru/raspisanie-stomatologicheskiy-fakultet"
ALLOWED_PREFIX = "/sites/default/files/files/"
SOURCE_PATTERN = re.compile(r"4_stomat[^\"'<>\s]*\.xlsx", re.IGNORECASE)
GROUP_PATTERN = re.compile(r"^\d{3}$")


def request_bytes(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "kgmu-calendar-source-probe/1.0"})
    with urllib.request.urlopen(request, timeout=45) as response:
        return response.read()


def discover_source() -> dict:
    page_bytes = request_bytes(TIMETABLE_PAGE)
    page_text = html.unescape(page_bytes.decode("utf-8", errors="replace"))
    hrefs = re.findall(r"href\s*=\s*[\"']([^\"']+)[\"']", page_text, flags=re.IGNORECASE)
    matches = []
    for href in hrefs:
        decoded = urllib.parse.unquote(href)
        if SOURCE_PATTERN.search(decoded):
            absolute = urllib.parse.urljoin(TIMETABLE_PAGE, href)
            parsed = urllib.parse.urlparse(absolute)
            if parsed.netloc != "kirovgma.ru" or not parsed.path.startswith(ALLOWED_PREFIX):
                raise SystemExit(f"discovered source outside allowed KGMU path: {absolute}")
            matches.append(absolute)
    matches = sorted(set(matches))
    if len(matches) != 1:
        raise SystemExit(f"expected exactly one current XLSX for dentistry course 4, found {matches}")
    return {"program": "dentistry", "course": 4, "url": matches[0]}


def workbook_dump(source: dict) -> dict:
    data = request_bytes(source["url"])
    if not data.startswith(b"PK"):
        raise SystemExit(f"source is not XLSX/ZIP: {source['url']}")
    XLSX_OUT.write_bytes(data)
    workbook = load_workbook(io.BytesIO(data), data_only=False)
    sheets = []
    discovered_groups = []
    for sheet in workbook.worksheets:
        non_empty = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.value is None:
                    continue
                value = str(cell.value).strip()
                entry = {"coord": cell.coordinate, "value": value}
                if cell.column == 2 and GROUP_PATTERN.fullmatch(value):
                    discovered_groups.append(value)
                if cell.hyperlink is not None:
                    entry["hyperlink"] = {
                        "target": cell.hyperlink.target,
                        "location": cell.hyperlink.location,
                        "display": cell.hyperlink.display,
                    }
                non_empty.append(entry)
        sheets.append({
            "title": sheet.title,
            "maxRow": sheet.max_row,
            "maxColumn": sheet.max_column,
            "mergedRanges": [str(rng) for rng in sheet.merged_cells.ranges],
            "nonEmptyCellCount": len(non_empty),
            "nonEmptyCells": non_empty,
        })
    groups = sorted(set(discovered_groups))
    if not groups:
        raise SystemExit("no three-digit group ids discovered in workbook column B")
    return {
        **source,
        "groups": groups,
        "sha256": hashlib.sha256(data).hexdigest(),
        "byteLength": len(data),
        "sheetNames": workbook.sheetnames,
        "sheets": sheets,
    }


def main() -> None:
    payload = {
        "schema": "kgmu-mechanical-source-probe-v1",
        "semanticParsingPerformed": False,
        "timetablePage": TIMETABLE_PAGE,
        "source": workbook_dump(discover_source()),
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    source = payload["source"]
    print(json.dumps({
        "program": source["program"],
        "course": source["course"],
        "groups": source["groups"],
        "url": source["url"],
        "sha256": source["sha256"],
        "byteLength": source["byteLength"],
        "sheetNames": source["sheetNames"],
        "dimensions": [
            [sheet["title"], sheet["maxRow"], sheet["maxColumn"], len(sheet["mergedRanges"]), sheet["nonEmptyCellCount"]]
            for sheet in source["sheets"]
        ],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
