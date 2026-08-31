#!/usr/bin/env python3
"""Compact structural inventory for the pinned KGMU medicine 5th-year XLSX.

This is an operator aid only: it records source geometry and source text without
performing semantic normalization.
"""
import hashlib
import io
import json
import urllib.request
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "medicine-501-516-source-inventory.json"
SOURCE_URL = "https://kirovgma.ru/sites/default/files/files/2026/08/24/1078/5_lech-24-08-2026-13.xlsx"
EXPECTED_SHA256 = "43ecb37de9db7ba69153c8514f62de0b058e51c2032e0ee320b117378a740c62"
EXPECTED_SHEET = "2026+-2027 осень 5 курс  Леч"
MONTHS = {"Сентябрь": 9, "Октябрь": 10, "Ноябрь": 11, "Декабрь": 12, "Январь": 1}


def fetch() -> bytes:
    req = urllib.request.Request(SOURCE_URL, headers={"User-Agent": "kgmu-calendar-source-inventory/1.0"})
    with urllib.request.urlopen(req, timeout=45) as response:
        data = response.read()
    actual = hashlib.sha256(data).hexdigest()
    if actual != EXPECTED_SHA256:
        raise SystemExit(f"source sha mismatch: {actual}")
    return data


def top_left_value(ws, rng):
    value = ws.cell(rng.min_row, rng.min_col).value
    return None if value is None else str(value).strip()


def value_at(ws, coord):
    value = ws[coord].value
    return None if value is None else str(value).strip()


def main():
    wb = load_workbook(io.BytesIO(fetch()), data_only=False)
    if wb.sheetnames != [EXPECTED_SHEET]:
        raise SystemExit(f"unexpected sheets: {wb.sheetnames}")
    ws = wb[EXPECTED_SHEET]
    ranges = list(ws.merged_cells.ranges)

    month_by_col = {}
    for rng in ranges:
        if rng.min_row == 10 and rng.max_row == 10:
            value = top_left_value(ws, rng)
            if value in MONTHS:
                for col in range(rng.min_col, rng.max_col + 1):
                    month_by_col[col] = MONTHS[value]
    date_by_col = {}
    for col in range(3, ws.max_column + 1):
        raw_day = ws.cell(11, col).value
        month = month_by_col.get(col)
        if raw_day is None or month is None:
            continue
        try:
            day = int(raw_day)
        except (TypeError, ValueError):
            continue
        year = 2027 if month == 1 else 2026
        try:
            date_by_col[col] = date(year, month, day).isoformat()
        except ValueError:
            pass

    group_rows = []
    value_usage = defaultdict(lambda: {"groups": set(), "locators": [], "dateCount": 0})
    for row in range(13, 29):
        group = str(ws.cell(row, 2).value or "").strip()
        if not group.isdigit():
            continue
        blocks = []
        for rng in ranges:
            if rng.min_row != row or rng.max_row != row or rng.min_col < 3:
                continue
            value = top_left_value(ws, rng)
            if value is None:
                continue
            covered_dates = [date_by_col[c] for c in range(rng.min_col, rng.max_col + 1) if c in date_by_col]
            locator = f"{get_column_letter(rng.min_col)}{row}"
            block = {
                "locator": locator,
                "range": str(rng),
                "value": value,
                "startDate": covered_dates[0] if covered_dates else None,
                "endDate": covered_dates[-1] if covered_dates else None,
                "dateCount": len(covered_dates),
                "dates": covered_dates,
            }
            blocks.append(block)
            usage = value_usage[value]
            usage["groups"].add(group)
            usage["locators"].append(locator)
            usage["dateCount"] += len(covered_dates)
        blocks.sort(key=lambda item: ws[item["locator"]].column)
        group_rows.append({"group": group, "row": row, "blocks": blocks})

    reference_rows = []
    reference_summary = []
    for row in range(31, 46):
        entries = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value is None:
                continue
            owning = None
            for rng in ranges:
                if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                    owning = rng
                    break
            if owning and (owning.min_row != row or owning.min_col != col):
                continue
            entry = {"coord": cell.coordinate, "value": str(cell.value).strip()}
            if owning:
                entry["range"] = str(owning)
            entries.append(entry)
        if entries:
            reference_rows.append({"row": row, "entries": entries})
        if row >= 33:
            reference_summary.append({
                "row": row,
                "number": value_at(ws, f"B{row}"),
                "discipline": value_at(ws, f"C{row}"),
                "assessment": value_at(ws, f"S{row}"),
                "department": value_at(ws, f"Y{row}"),
                "base": value_at(ws, f"AN{row}"),
                "address": value_at(ws, f"BL{row}"),
                "firstShift": value_at(ws, f"BT{row}"),
                "secondShift": value_at(ws, f"BX{row}"),
            })

    unique_block_values = []
    for value, usage in sorted(value_usage.items()):
        unique_block_values.append({
            "value": value,
            "groups": sorted(usage["groups"], key=int),
            "locators": usage["locators"],
            "dateCountAcrossBlocks": usage["dateCount"],
        })

    payload = {
        "schema": "kgmu-source-structural-inventory-v1",
        "semanticParsingPerformed": False,
        "sourceUrl": SOURCE_URL,
        "sourceSha256": EXPECTED_SHA256,
        "sheetName": EXPECTED_SHEET,
        "calendarColumns": [{"column": get_column_letter(c), "date": d} for c, d in sorted(date_by_col.items())],
        "uniqueBlockValues": unique_block_values,
        "groupRows": group_rows,
        "referenceSummary": reference_summary,
        "referenceRows": reference_rows,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "groups": [item["group"] for item in group_rows],
        "blockCounts": {item["group"]: len(item["blocks"]) for item in group_rows},
        "uniqueBlockValues": unique_block_values,
        "referenceSummary": reference_summary,
        "calendarColumnCount": len(date_by_col),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
