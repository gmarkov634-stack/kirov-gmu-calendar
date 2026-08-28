#!/usr/bin/env python3
import collections
import datetime as dt
import hashlib
import itertools
import json
import re
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple

ROOT = Path(__file__).resolve().parents[2]
SOURCE_META = ROOT / "fixtures/2026-2027-semester-1/dentistry-191-194.source.json"
OUT_DIR = ROOT / "fixtures/2026-2027-semester-1/normalized"
QA_DIR = ROOT / "qa/2026-2027-semester-1"

UNIVERSITY_ID = "kirov-gmu"
PERIOD_ID = "2026-2027-semester-1"
SOURCE_ID = "dentistry"
SHEET_NAME = "1 стомат"
GROUP_ID = "191"
PARSER_RULES_VERSION = "kgmu-2026-08-27-v3"

DATE_PAT = r"\d{1,2}\.(?:12|11|10|0?9|0?1)(?!\d)"
TIME_TOKEN = r"\d{1,2}\.\d{2}"

DISC_PATTERNS = [
    ("Элективные дисциплины по физической культуре и спорту", r"ЭЛЕКТИВНЫЕ ДИСЦИПЛИНЫ\s*\(МОДУЛИ\)\s*ПО ФИЗИЧЕСКОЙ КУЛЬТУРЕ И СПОРТУ"),
    ("Основы российской государственности", r"Основы российской государственности"),
    ("Гистология, эмбриология, цитология-гистология полости рта", r"Гистология,\s*эмбриология,\s*цитология[-–—]гистология полости рта"),
    ("Анатомия человека-анатомия головы и шеи", r"Анатомия человека[-–—]анатомия головы и шеи"),
    ("Медицинская и биологическая физика", r"Медицинская и биологическая физика"),
    ("Общая и биоорганическая химия", r"Общая и биоорганическая химия"),
    ("Русский язык и культура речи", r"Русский язык и культура речи(?:\s*\(факультатив\))?"),
    ("Медицинская информатика", r"Медицинская информатика"),
    ("Физика, математика", r"Физика,\s*математика"),
    ("Иностранный язык", r"Иностранный язык"),
    ("История медицины", r"История медицины"),
    ("История России", r"История России"),
    ("Латинский язык", r"Латинский язык"),
    ("Биология", r"Биология"),
    ("Библиотечный час", r"Библиотечный час"),
    ("Час куратора", r"Час куратора"),
]
DISC_ALT = "|".join(f"(?P<d{i}>{pat})" for i, (_, pat) in enumerate(DISC_PATTERNS))
TIME_PREFIX = rf"(?P<times>{TIME_TOKEN}\s*-\s*{TIME_TOKEN}(?:\s*,\s*{TIME_TOKEN}\s*-\s*{TIME_TOKEN})?|{TIME_TOKEN}\s*-\s*{TIME_TOKEN}\s*-\s*{TIME_TOKEN}\s*-\s*{TIME_TOKEN})"
EVENT_RE = re.compile(rf"{TIME_PREFIX}\s+(?P<lecture>Лекция\s+)?(?:{DISC_ALT})", re.I)
CONTROL_RE = re.compile(
    rf"(?P<date>{DATE_PAT}).{{0,45}}?(?P<control>зач[её]т(?:\s+с\s+оценкой)?|экзамен).{{0,25}}?(?P<t1>{TIME_TOKEN})\s*-\s*(?P<t2>{TIME_TOKEN})",
    re.I,
)
CONTROL_RE2 = re.compile(
    rf"(?P<control>зач[её]т(?:\s+с\s+оценкой)?|экзамен).{{0,25}}?(?P<date>{DATE_PAT}).{{0,15}}?(?P<t1>{TIME_TOKEN})\s*-\s*(?P<t2>{TIME_TOKEN})",
    re.I,
)

ADDR_BY_CORPUS = {
    "1": "ул. Владимирская, 137",
    "2": "ул. Пролетарская, 38",
    "3": "ул. Владимирская, 112",
}
DEFAULT_LOCATIONS = {
    "Основы российской государственности": "1 корпус, ул. Владимирская, 137",
    "История медицины": "1 корпус, ул. Владимирская, 137",
    "Латинский язык": "1 корпус, ул. Владимирская, 137",
    "Физика, математика": "3 корпус, ул. Владимирская, 112",
    "Общая и биоорганическая химия": "1 корпус, ул. Владимирская, 137",
    "Анатомия человека-анатомия головы и шеи": "2 корпус, ул. Пролетарская, 38",
    "История России": "1 корпус, ул. Владимирская, 137",
    "Иностранный язык": "1 корпус, ул. Владимирская, 137",
    "Медицинская информатика": "3 корпус, ул. Владимирская, 112",
    "Медицинская и биологическая физика": "3 корпус, ул. Владимирская, 112",
    "Биология": "3 корпус, ул. Владимирская, 112",
    "Гистология, эмбриология, цитология-гистология полости рта": "1 корпус, ул. Владимирская, 137",
    "Русский язык и культура речи": "ул. Красноармейская, 35",
    "Элективные дисциплины по физической культуре и спорту": "ФОК, 3 корпус, ул. Владимирская, 112",
}

DAY_ROWS = [
    (10, 16, 0),
    (17, 22, 1),
    (23, 27, 2),
    (29, 34, 3),
    (36, 42, 4),
    (43, 49, 5),
]

CELL_LOCATION_SCOPE = {
    "B43": "1 корпус, аудитория 306, ул. Владимирская, 137",
}


def ddmm(value):
    day, month = map(int, value.split("."))
    return dt.date(2026 if month >= 9 else 2027, month, day)


def week_intervals(values):
    return [(ddmm(a), ddmm(b)) for a, b in values]


WEEK1 = week_intervals([
    ("01.09", "05.09"),
    ("14.09", "19.09"),
    ("28.09", "03.10"),
    ("12.10", "17.10"),
    ("26.10", "31.10"),
    ("09.11", "14.11"),
    ("23.11", "28.11"),
    ("07.12", "12.12"),
    ("21.12", "26.12"),
    ("11.01", "16.01"),
])
WEEK2 = week_intervals([
    ("07.09", "12.09"),
    ("21.09", "26.09"),
    ("05.10", "10.10"),
    ("19.10", "24.10"),
    ("02.11", "07.11"),
    ("16.11", "21.11"),
    ("30.11", "05.12"),
    ("14.12", "19.12"),
    ("28.12", "30.12"),
])
ALL_WEEKS = WEEK1 + WEEK2


def in_grid(value):
    return any(left <= value <= right for left, right in ALL_WEEKS)


def week_no(value):
    if any(left <= value <= right for left, right in WEEK1):
        return 1
    if any(left <= value <= right for left, right in WEEK2):
        return 2
    return None


def dates_in_range(start, end, weekday, parity=None):
    out = []
    current = start
    while current <= end:
        if current.weekday() == weekday and in_grid(current):
            if parity is None or week_no(current) == parity:
                out.append(current)
        current += dt.timedelta(days=1)
    return out


def weekday_for_row(row):
    for first, last, weekday in DAY_ROWS:
        if first <= row <= last:
            return weekday
    raise ValueError(f"row outside timetable: {row}")


def parse_time(value):
    hour, minute = map(int, value.split("."))
    return f"{hour:02d}:{minute:02d}"


def time_bounds(value):
    tokens = re.findall(TIME_TOKEN, value)
    return parse_time(tokens[0]), parse_time(tokens[-1])


def location_from_code(corpus, auditorium):
    corpus = str(int(corpus))
    auditorium = str(int(auditorium))
    return f"{corpus} корпус, аудитория {auditorium}, {ADDR_BY_CORPUS[corpus]}"


def explicit_location(text, discipline=None):
    full = list(re.finditer(
        r"([123])\s*корпус\s*,?\s*ауд(?:итория|\.)?\s*(\d{3})",
        text,
        re.I,
    ))
    if full:
        return location_from_code(*full[-1].groups())
    compact = list(re.finditer(r"(?<![\d.])([123])\s*-\s*(\d{3})(?!\d)", text))
    if compact:
        return location_from_code(*compact[-1].groups())
    if re.search(r"\bФОК\b", text, re.I):
        return "ФОК, 3 корпус, ул. Владимирская, 112"
    return DEFAULT_LOCATIONS.get(discipline)


def find_segments(text):
    matches = list(EVENT_RE.finditer(text))
    result = []
    for index, match in enumerate(matches):
        discipline = None
        for i, (name, _) in enumerate(DISC_PATTERNS):
            if match.group(f"d{i}") is not None:
                discipline = name
                break
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        result.append((match, discipline, text[match.end():end].strip()))
    return result


def event(value_date, start, end, discipline, lesson_type, location, coord, segment, extra=""):
    locator = f"{SHEET_NAME}!{coord}#s{segment}"
    key = f"{GROUP_ID}|{value_date.isoformat()}|{start}|{end}|{discipline}|{lesson_type}|{locator}|{extra}"
    return {
        "eventId": "kgmu-" + hashlib.sha256(key.encode()).hexdigest()[:24],
        "universityId": UNIVERSITY_ID,
        "groupId": GROUP_ID,
        "academicPeriodId": PERIOD_ID,
        "date": value_date.isoformat(),
        "startTime": start,
        "endTime": end,
        "timeSemantics": "floating",
        "discipline": discipline,
        "lessonType": lesson_type,
        "teacher": None,
        "location": location,
        "sourceRef": {"sourceId": SOURCE_ID, "locator": locator},
    }


def parse_segment(coord, row, segment, match, discipline, tail):
    weekday = weekday_for_row(row)
    start, end = time_bounds(match.group("times"))
    lesson = (
        "lecture"
        if match.group("lecture")
        else ("other" if discipline in ("Библиотечный час", "Час куратора") else "practice")
    )
    base_location = (
        explicit_location(tail, discipline)
        or CELL_LOCATION_SCOPE.get(coord)
        or DEFAULT_LOCATIONS.get(discipline)
    )
    work = tail
    events = []
    notes = []

    if discipline == "Час куратора" and not re.search(DATE_PAT, work):
        for value_date in dates_in_range(dt.date(2026, 9, 1), dt.date(2027, 1, 16), weekday)[:2]:
            events.append(event(value_date, start, end, discipline, "other", base_location, coord, segment, "R17"))
        return events, notes

    location_overrides = {}
    for paren in list(re.finditer(r"\(([^()]*)\)", work)):
        content = paren.group(1)
        loc = re.search(r"(?<!\d)([123])\s*-\s*(\d{3})(?!\d)", content)
        if loc and not re.search(r"зач|экзам", content, re.I):
            values = [ddmm(x) for x in re.findall(DATE_PAT, content[:loc.start()])]
            if values:
                location = location_from_code(*loc.groups())
                for value_date in values:
                    location_overrides[value_date] = location
                work = work.replace(paren.group(0), " ")

    extras = []
    for paren in list(re.finditer(r"\(([^()]*)\)", work)):
        content = paren.group(1)
        control = CONTROL_RE.search(content) or CONTROL_RE2.search(content)
        if control:
            value_date = ddmm(control.group("date"))
            c_start = parse_time(control.group("t1"))
            c_end = parse_time(control.group("t2"))
            c_type = "exam" if "экзам" in control.group("control").lower() else "credit"
            extras.append((value_date, c_start, c_end, c_type, explicit_location(content, discipline) or base_location, "control"))
            work = work.replace(paren.group(0), " ")
            m = re.search(r"(\d+)\s+(занятие|занятия|лекци[яи])\s+в\s+(пн|вт|ср|чт|пт|сб|вс)", content, re.I)
            if m:
                notes.append((int(m.group(1)), m.group(2).lower(), m.group(3).lower()))
            continue

        override = re.search(
            rf"(?P<date>{DATE_PAT})\s*-\s*(?P<t1>{TIME_TOKEN})\s*-\s*(?P<t2>{TIME_TOKEN})",
            content,
        )
        if override:
            extras.append((
                ddmm(override.group("date")),
                parse_time(override.group("t1")),
                parse_time(override.group("t2")),
                lesson,
                base_location,
                "override",
            ))
            work = work.replace(paren.group(0), " ")
            continue

        m = re.search(r"(\d+)\s+(занятие|занятия|лекци[яи])\s+в\s+(пн|вт|ср|чт|пт|сб|вс)", content, re.I)
        if m:
            notes.append((int(m.group(1)), m.group(2).lower(), m.group(3).lower()))

    for m in re.finditer(r"(\d+)\s+(занятие|занятия|лекци[яи])\s+в\s+(пн|вт|ср|чт|пт|сб|вс)", work, re.I):
        notes.append((int(m.group(1)), m.group(2).lower(), m.group(3).lower()))

    override_pattern = re.compile(
        rf"(?P<date>{DATE_PAT})\s*-\s*(?P<t1>{TIME_TOKEN})\s*-\s*(?P<t2>{TIME_TOKEN})"
    )
    overrides = []
    for override in list(override_pattern.finditer(work)):
        overrides.append((
            ddmm(override.group("date")),
            parse_time(override.group("t1")),
            parse_time(override.group("t2")),
        ))
    work = override_pattern.sub(lambda item: item.group("date"), work)

    parity = None
    parity_match = re.search(r"([12])\s*недел", work, re.I)
    if parity_match:
        parity = int(parity_match.group(1))

    date_records = []
    bounded = re.search(rf"\bс\s*({DATE_PAT})\s*по\s*({DATE_PAT})", work, re.I)
    if bounded:
        date_records.extend(dates_in_range(ddmm(bounded.group(1)), ddmm(bounded.group(2)), weekday, parity))
    else:
        until = re.search(rf"\bпо\s*({DATE_PAT})", work, re.I)
        if parity and until:
            date_records.extend(dates_in_range(dt.date(2026, 9, 1), ddmm(until.group(1)), weekday, parity))
            date_records.extend(ddmm(x) for x in re.findall(DATE_PAT, work[until.end():]))
        else:
            temp = re.sub(r"(?<!\d)[123]\s*-\s*\d{3}(?!\d)", " ", work)
            temp = re.split(r"\b[123]\s*корпус\b", temp, maxsplit=1, flags=re.I)[0]
            range_pattern = re.compile(rf"({DATE_PAT})\s*-\s*({DATE_PAT})")
            ranges = list(range_pattern.finditer(temp))
            if ranges:
                for item in ranges:
                    date_records.extend(dates_in_range(ddmm(item.group(1)), ddmm(item.group(2)), weekday, parity))
                temp = range_pattern.sub(" ", temp)
            date_records.extend(ddmm(x) for x in re.findall(DATE_PAT, temp))

    override_map = {value_date: (o_start, o_end) for value_date, o_start, o_end in overrides}
    paren_override_map = {
        value_date: (o_start, o_end, o_type, o_location)
        for value_date, o_start, o_end, o_type, o_location, kind in extras
        if kind == "override"
    }

    for value_date in sorted(set(date_records)):
        cur_start, cur_end = start, end
        cur_type = lesson
        cur_location = location_overrides.get(value_date, base_location)
        if value_date in override_map:
            cur_start, cur_end = override_map[value_date]
        if value_date in paren_override_map:
            cur_start, cur_end, cur_type, cur_location = paren_override_map[value_date]
        events.append(event(value_date, cur_start, cur_end, discipline, cur_type, cur_location, coord, segment))

    for value_date, e_start, e_end, e_type, e_location, kind in extras:
        if kind == "override":
            continue
        if kind == "control":
            events = [
                item
                for item in events
                if not (item["date"] == value_date.isoformat() and item["discipline"] == discipline)
            ]
        events.append(event(value_date, e_start, e_end, discipline, e_type, e_location, coord, segment, kind))

    if not events:
        raise ValueError(f"{coord} segment {segment}: no events from {tail!r}")
    return events, notes


DAY_CODE = {"пн": 0, "вт": 1, "ср": 2, "чт": 3, "пт": 4, "сб": 5, "вс": 6}


def source_coord(item):
    return re.search(r"!([A-Z]+\d+)#", item["sourceRef"]["locator"]).group(1)


def overlaps(left, right):
    def minutes(value):
        hour, minute = map(int, value.split(":"))
        return hour * 60 + minute
    return minutes(left["startTime"]) < minutes(right["endTime"]) and minutes(right["startTime"]) < minutes(left["endTime"])


def main():
    meta = json.loads(SOURCE_META.read_text(encoding="utf-8"))
    if meta["parserRulesVersion"] != PARSER_RULES_VERSION:
        raise SystemExit("parser rules version mismatch")

    data = urllib.request.urlopen(meta["source"]["url"], timeout=30).read()
    actual_sha = hashlib.sha256(data).hexdigest()
    if actual_sha != meta["source"]["sha256"]:
        raise SystemExit(f"SHA mismatch: {actual_sha}")
    if len(data) != meta["source"]["byteLength"]:
        raise SystemExit(f"byteLength mismatch: {len(data)}")

    temp = ROOT / ".tmp-kgmu-dentistry-191-194.xlsx"
    temp.write_bytes(data)
    try:
        workbook = load_workbook(temp, data_only=False)
    finally:
        temp.unlink(missing_ok=True)

    exp = meta["workbookExpectations"]
    if workbook.sheetnames != exp["sheetNames"]:
        raise SystemExit(f"sheet names changed: {workbook.sheetnames}")
    sheet = workbook[SHEET_NAME]
    if sheet.max_row != exp["maxRow"] or sheet.max_column != exp["maxColumn"]:
        raise SystemExit(f"workbook geometry changed: {sheet.max_row}x{sheet.max_column}")
    if str(sheet.calculate_dimension()) != exp["dimension"]:
        raise SystemExit(f"worksheet dimension changed: {sheet.calculate_dimension()}")
    if len(sheet.merged_cells.ranges) != exp["mergedRangeCount"]:
        raise SystemExit(f"merged range count changed: {len(sheet.merged_cells.ranges)}")

    anchors = {}
    for row in range(10, 50):
        coord = f"B{row}"
        anchor = coord
        for merged in sheet.merged_cells.ranges:
            if coord in merged:
                anchor = sheet.cell(merged.min_row, merged.min_col).coordinate
                break
        anchors[anchor] = row

    logical = []
    ignored = []
    for coord in sorted(anchors, key=lambda c: coordinate_to_tuple(c)):
        cell = sheet[coord]
        if cell.value is None:
            continue
        raw = str(cell.value).strip()
        row, _ = coordinate_to_tuple(coord)
        if row == 49 or raw.upper().startswith("ФАКУЛЬТАТИВЫ:"):
            ignored.append({"sourceLocator": f"{SHEET_NAME}!{coord}", "rule": "R39", "text": raw})
            continue
        logical.append(cell)

    normalized = []
    parse_notes = []
    unclassified = []

    for cell in logical:
        row, _ = coordinate_to_tuple(cell.coordinate)
        segments = find_segments(str(cell.value))
        if not segments:
            unclassified.append({"sourceLocator": f"{SHEET_NAME}!{cell.coordinate}", "text": str(cell.value)})
            continue
        for index, (match, discipline, tail) in enumerate(segments, 1):
            try:
                events, notes = parse_segment(cell.coordinate, row, index, match, discipline, tail)
                normalized.extend(events)
                for count, kind, day in notes:
                    parse_notes.append({
                        "sourceLocator": f"{SHEET_NAME}!{cell.coordinate}#s{index}",
                        "discipline": discipline,
                        "lessonType": "lecture" if "лекц" in kind else "practice",
                        "expectedCount": count,
                        "weekday": day,
                    })
            except Exception as exc:
                unclassified.append({
                    "sourceLocator": f"{SHEET_NAME}!{cell.coordinate}#s{index}",
                    "text": str(cell.value),
                    "error": str(exc),
                })

    normalized.sort(key=lambda item: (
        item["date"], item["startTime"], item["endTime"], item["discipline"],
        item["lessonType"], item["sourceRef"]["locator"]
    ))

    deduped = []
    seen = set()
    duplicates = []
    for item in normalized:
        sig = (
            item["groupId"], item["date"], item["startTime"], item["endTime"],
            item["discipline"], item["lessonType"], item["location"]
        )
        if sig in seen:
            duplicates.append(item)
            continue
        seen.add(sig)
        deduped.append(item)
    normalized = deduped

    note_checks = []
    for note in parse_notes:
        weekday = DAY_CODE[note["weekday"]]
        origin_coord = re.search(r"!([A-Z]+\d+)#", note["sourceLocator"]).group(1)
        matches = [
            item for item in normalized
            if item["discipline"] == note["discipline"]
            and item["lessonType"] == note["lessonType"]
            and dt.date.fromisoformat(item["date"]).weekday() == weekday
            and source_coord(item) != origin_coord
        ]
        note_checks.append({
            **note,
            "matchedCount": len(matches),
            "matchedEventIds": [item["eventId"] for item in matches],
            "status": "pass" if len(matches) == note["expectedCount"] else "review_required",
        })

    by_date = collections.defaultdict(list)
    for item in normalized:
        by_date[item["date"]].append(item)
    overlap_rows = []
    for value_date, items in sorted(by_date.items()):
        for left, right in itertools.combinations(items, 2):
            if overlaps(left, right):
                overlap_rows.append({
                    "date": value_date,
                    "left": {
                        "startTime": left["startTime"],
                        "endTime": left["endTime"],
                        "discipline": left["discipline"],
                        "sourceLocator": left["sourceRef"]["locator"],
                    },
                    "right": {
                        "startTime": right["startTime"],
                        "endTime": right["endTime"],
                        "discipline": right["discipline"],
                        "sourceLocator": right["sourceRef"]["locator"],
                    },
                })

    warnings = []
    if unclassified:
        warnings.append(f"{len(unclassified)} source fragment(s) require review")
    failed_notes = [row for row in note_checks if row["status"] != "pass"]
    if failed_notes:
        warnings.append(f"{len(failed_notes)} additional-event completeness note(s) require review")
    if duplicates:
        warnings.append(f"{len(duplicates)} duplicate normalized event(s) removed")
    if overlap_rows:
        warnings.append(f"{len(overlap_rows)} explicit time overlap(s) preserved under G16/R69")

    canonical = json.dumps(normalized, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    candidate_digest = hashlib.sha256(canonical.encode()).hexdigest()

    parsing_result = {
        "jobId": "parsing-job-dentistry-191-candidate-1",
        "universityId": UNIVERSITY_ID,
        "academicPeriodId": PERIOD_ID,
        "parserRulesVersion": PARSER_RULES_VERSION,
        "events": normalized,
        "warnings": warnings,
    }
    evidence = {
        "fixtureId": meta["fixtureId"],
        "sourceSha256": actual_sha,
        "parserRulesVersion": PARSER_RULES_VERSION,
        "groupId": GROUP_ID,
        "logicalSourceCellCount": len(logical),
        "ignoredSourceCells": ignored,
        "eventCount": len(normalized),
        "coveredSourceCellCount": len({source_coord(item) for item in normalized}),
        "unclassifiedSourceFragments": unclassified,
        "additionalEventChecks": note_checks,
        "duplicateEventsRemoved": len(duplicates),
        "explicitOverlapWarningCount": len(overlap_rows),
        "explicitOverlapWarnings": overlap_rows,
        "candidateDigest": "sha256:" + candidate_digest,
        "decision": "review_required" if (unclassified or failed_notes) else "candidate_ready_for_qa",
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    QA_DIR.mkdir(parents=True, exist_ok=True)
    result_path = OUT_DIR / "dentistry-191.parsing-result.candidate.json"
    evidence_path = QA_DIR / "dentistry-191.candidate.evidence.json"
    result_path.write_text(json.dumps(parsing_result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    evidence_path.write_text(json.dumps(evidence, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(json.dumps({
        "decision": evidence["decision"],
        "eventCount": evidence["eventCount"],
        "logicalSourceCellCount": evidence["logicalSourceCellCount"],
        "coveredSourceCellCount": evidence["coveredSourceCellCount"],
        "unclassified": len(unclassified),
        "additionalChecks": len(note_checks),
        "additionalChecksFailed": len(failed_notes),
        "duplicatesRemoved": len(duplicates),
        "overlaps": len(overlap_rows),
        "candidateDigest": candidate_digest,
        "result": str(result_path.relative_to(ROOT)),
        "evidence": str(evidence_path.relative_to(ROOT)),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
