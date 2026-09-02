#!/usr/bin/env python3
"""Inspect source-local visual evidence for C20-ambiguous dentistry 391-394 cycles.

This is a mechanical evidence probe only. It does not infer semantics from styles.
It records whether cells inside the affected merged ranges carry source-visible
formatting differences that could justify a later source-specific interpretation.
"""
from __future__ import annotations

import hashlib
import io
import json
import re
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.colors import Color

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "qa/2026-2027-semester-1/dentistry-391-394.c20-style-probe.json"
SOURCE_URL = "https://kirovgma.ru/sites/default/files/files/2026/08/24/1097/3_stomat-24-08-2026-14.xlsx"
SOURCE_SHA256 = "82fcb873776634553f9dcc5bf3da581654d59f4ef10db5ad6a779aa6d53f950d"
TARGET_LABELS = {
    "Философия": "philosophy",
    "Акушерство": "obstetrics",
    "ИОК врача-стоматолога": "iok",
}
ALIASES = {"ИОК врача- стоматолога": "ИОК врача-стоматолога"}
EXPECTED_GROUPS = {"391", "392", "393", "394"}


def request_bytes(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "kgmu-calendar-c20-style-probe/1.0"})
    with urllib.request.urlopen(req, timeout=45) as response:
        return response.read()


def compact(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def color_value(color: Color | None):
    if color is None:
        return None
    return {
        "type": color.type,
        "rgb": color.rgb if color.type == "rgb" else None,
        "indexed": color.indexed if color.type == "indexed" else None,
        "theme": color.theme if color.type == "theme" else None,
        "tint": color.tint,
        "auto": color.auto,
    }


def side_value(side):
    return {"style": side.style, "color": color_value(side.color)}


def semantic_style(cell):
    fill = cell.fill
    font = cell.font
    alignment = cell.alignment
    return {
        "styleId": cell.style_id,
        "cellClass": "MergedCell" if isinstance(cell, MergedCell) else "Cell",
        "fill": {
            "fillType": fill.fill_type,
            "fgColor": color_value(fill.fgColor),
            "bgColor": color_value(fill.bgColor),
        },
        "font": {
            "name": font.name,
            "size": font.sz,
            "bold": font.b,
            "italic": font.i,
            "underline": font.u,
            "strike": font.strike,
            "color": color_value(font.color),
        },
        "alignment": {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
            "textRotation": alignment.textRotation,
            "wrapText": alignment.wrapText,
            "shrinkToFit": alignment.shrinkToFit,
            "indent": alignment.indent,
        },
        "numberFormat": cell.number_format,
        "protection": {"locked": cell.protection.locked, "hidden": cell.protection.hidden},
    }


def border_style(cell):
    border = cell.border
    return {
        "left": side_value(border.left),
        "right": side_value(border.right),
        "top": side_value(border.top),
        "bottom": side_value(border.bottom),
        "diagonal": side_value(border.diagonal),
    }


def semantic_key(style: dict) -> str:
    # styleId is intentionally excluded because merged-edge borders can alter it.
    reduced = {key: value for key, value in style.items() if key not in {"styleId", "cellClass"}}
    return json.dumps(reduced, ensure_ascii=False, sort_keys=True, default=str)


def conditional_formatting(ws):
    records = []
    for cf in ws.conditional_formatting:
        rules = []
        for rule in cf.rules:
            dxf = rule.dxf
            rules.append({
                "type": rule.type,
                "operator": rule.operator,
                "formula": list(rule.formula or []),
                "dxf": None if dxf is None else {
                    "fill": None if dxf.fill is None else {
                        "fillType": dxf.fill.fill_type,
                        "fgColor": color_value(dxf.fill.fgColor),
                        "bgColor": color_value(dxf.fill.bgColor),
                    },
                    "font": None if dxf.font is None else {
                        "bold": dxf.font.b,
                        "italic": dxf.font.i,
                        "color": color_value(dxf.font.color),
                    },
                },
            })
        records.append({"sqref": str(cf.sqref), "rules": rules})
    return records


def main() -> None:
    data = request_bytes(SOURCE_URL)
    sha = hashlib.sha256(data).hexdigest()
    if sha != SOURCE_SHA256:
        raise SystemExit(f"official source changed: {sha}")
    wb = load_workbook(io.BytesIO(data), data_only=False)
    if len(wb.worksheets) != 1:
        raise SystemExit("expected one worksheet")
    ws = wb.worksheets[0]

    group_rows = {}
    for row in range(1, ws.max_row + 1):
        value = compact(ws.cell(row=row, column=2).value)
        if value in EXPECTED_GROUPS:
            group_rows[row] = value
    if set(group_rows.values()) != EXPECTED_GROUPS:
        raise SystemExit(f"group rows mismatch: {group_rows}")

    date_labels = {column: compact(ws.cell(row=11, column=column).value) for column in range(1, ws.max_column + 1)}
    target_ranges = []
    for merged in ws.merged_cells.ranges:
        if merged.min_row != merged.max_row or merged.min_row not in group_rows:
            continue
        top_left = ws.cell(row=merged.min_row, column=merged.min_col)
        raw_label = compact(top_left.value)
        normalized = ALIASES.get(raw_label, raw_label)
        kind = TARGET_LABELS.get(normalized)
        if kind is None:
            continue

        cells = []
        semantic_keys = set()
        for column in range(merged.min_col, merged.max_col + 1):
            cell = ws.cell(row=merged.min_row, column=column)
            style = semantic_style(cell)
            semantic_keys.add(semantic_key(style))
            cells.append({
                "coordinate": cell.coordinate,
                "dateLabelRow11": date_labels.get(column),
                "value": compact(cell.value) or None,
                "semanticStyle": style,
                "border": border_style(cell),
            })
        target_ranges.append({
            "range": str(merged),
            "groupId": group_rows[merged.min_row],
            "kind": kind,
            "rawLabel": raw_label,
            "normalizedLabel": normalized,
            "columnCount": merged.max_col - merged.min_col + 1,
            "distinctSemanticStyleCount": len(semantic_keys),
            "hasWithinRangeSemanticStyleVariation": len(semantic_keys) > 1,
            "cells": cells,
        })

    target_ranges.sort(key=lambda item: (item["groupId"], item["range"]))
    if len(target_ranges) != 12:
        raise SystemExit(f"expected 12 C20 target ranges, found {len(target_ranges)}")

    payload = {
        "schema": "kgmu-c20-source-style-probe-v1",
        "semanticInferencePerformed": False,
        "sourceUrl": SOURCE_URL,
        "sourceSha256": sha,
        "sheetName": ws.title,
        "targetRangeCount": len(target_ranges),
        "conditionalFormatting": conditional_formatting(ws),
        "targetRanges": target_ranges,
        "summary": {
            "rangesWithWithinRangeSemanticStyleVariation": sum(
                item["hasWithinRangeSemanticStyleVariation"] for item in target_ranges
            ),
            "allTargetRangesSemanticallyUniformByCellStyle": all(
                not item["hasWithinRangeSemanticStyleVariation"] for item in target_ranges
            ),
        },
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
    print(json.dumps(payload["summary"], ensure_ascii=False, indent=2))
    print(f"conditional formatting regions: {len(payload['conditionalFormatting'])}")


if __name__ == "__main__":
    main()
