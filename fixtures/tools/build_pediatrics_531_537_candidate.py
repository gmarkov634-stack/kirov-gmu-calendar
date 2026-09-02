#!/usr/bin/env python3
"""Build the deterministic normalized draft for KGMU Pediatrics course 5.

The upper cyclic grid is normalized under the existing G+C profile. The lower
physical-education schedule is deliberately left unresolved because its weekday
label conflicts with its date range; G04/G21 require manual confirmation.
"""
import hashlib
import io
import json
import re
import urllib.request
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
PERIOD = "2026-2027-semester-1"
FIXTURE_DIR = ROOT / "fixtures" / PERIOD
QA_DIR = ROOT / "qa" / PERIOD
SOURCE_PATH = FIXTURE_DIR / "pediatrics-531-537.source.json"
JOB_PATH = FIXTURE_DIR / "pediatrics-531-537.parsing-job.json"
MANIFEST_PATH = FIXTURE_DIR / "pediatrics-531-537.decisions.json"
REVIEW_PATH = QA_DIR / "pediatrics-531-537.semantic-review.json"

RAW_TO_REFERENCE_ROW = {
    "Акушерство и гинекология": 32,
    "Госпитальная хирургия": 27,
    "Госпитальная терапия": 33,
    "Дерматовенерология": 25,
    "Детская хирургия": 35,
    "Инфекционные болезни": 28,
    "Медицина катастроф": 31,
    "Онкология": 29,
    "Психиатрия, мед. психология": 26,
    "Травматология, ортопедия": 30,
    "Факульт. педиатрия, эндокринология": 34,
}
MONTHS = {"Сентябрь": 9, "Октябрь": 10, "Ноябрь": 11, "Декабрь": 12, "Январь": 1}
TIME_RE = re.compile(r"(?P<sh>\d{1,2})[:.](?P<sm>\d{2})\s*[-–—]\s*(?P<eh>\d{1,2})[:.](?P<em>\d{2})")


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def norm(value):
    if value is None:
        return None
    return re.sub(r"\s+", " ", str(value)).strip()


def fetch(url):
    req = urllib.request.Request(url, headers={"User-Agent": "kgmu-calendar-candidate-builder/1.0"})
    with urllib.request.urlopen(req, timeout=45) as response:
        return response.read()


def parse_time(value):
    raw = norm(value)
    if not raw:
        return None
    match = TIME_RE.search(raw)
    if not match:
        return None
    return (
        f"{int(match.group('sh')):02d}:{match.group('sm')}",
        f"{int(match.group('eh')):02d}:{match.group('em')}",
    )


def group_mask(index):
    return format(1 << index, "x")


def date_mask(date_table, dates):
    lookup = {value: index for index, value in enumerate(date_table)}
    mask = 0
    for value in dates:
        if value not in lookup:
            raise SystemExit(f"date outside source calendar: {value}")
        mask |= 1 << lookup[value]
    if mask == 0:
        raise SystemExit("empty date mask")
    return format(mask, "x")


def join_location(base, address):
    base = norm(base)
    address = norm(address)
    if base and address:
        return f"{base}, {address}"
    return base or address or "Место не указано"


def assessment(value, source_id, sheet, row):
    label = norm(value)
    if not label:
        return None
    lowered = label.lower().replace("ё", "е")
    if "экзам" in lowered:
        kind = "exam"
    elif "зач" in lowered:
        kind = "credit"
    else:
        kind = "other"
    return {
        "type": kind,
        "label": label.lower(),
        "sourceRef": {"sourceId": source_id, "locator": f"{sheet}!S{row}"},
    }


def dates_for_weekday(start, end, weekday):
    current = start
    result = []
    while current <= end:
        if current.weekday() == weekday:
            result.append(current.isoformat())
        current += timedelta(days=1)
    return result


def main():
    source = read_json(SOURCE_PATH)
    job = read_json(JOB_PATH)
    source_info = source["source"]
    if source["parserProfile"] != "cyclic":
        raise SystemExit("pediatrics 531-537 must use the cyclic profile")
    if source["parserRulesVersion"] != "kgmu-2026-08-27-v3":
        raise SystemExit("unexpected parser rules version")
    if job["sourceId"] != source_info["sourceId"] or job["sourceObjectKey"] != source_info["objectKey"]:
        raise SystemExit("ParsingJob is not bound to the pinned SourceArtifact")
    if job["expectedGroupIds"] != source["expectedGroupIds"]:
        raise SystemExit("ParsingJob expected groups differ from source fixture")

    data = fetch(source_info["url"])
    if hashlib.sha256(data).hexdigest() != source_info["sha256"]:
        raise SystemExit("official source SHA-256 changed")
    if len(data) != source_info["byteLength"]:
        raise SystemExit("official source byte length changed")

    wb = load_workbook(io.BytesIO(data), data_only=False)
    if wb.sheetnames != source["workbookExpectations"]["sheetNames"]:
        raise SystemExit(f"unexpected sheet names: {wb.sheetnames}")
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    merged = list(ws.merged_cells.ranges)
    expected = source["workbookExpectations"]
    non_empty = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
    actual_geometry = (ws.max_row, ws.max_column, len(merged), non_empty)
    expected_geometry = (
        expected["maxRow"], expected["maxColumn"], expected["mergedRangeCount"], expected["nonEmptyCellCount"]
    )
    if actual_geometry != expected_geometry:
        raise SystemExit(f"workbook geometry changed: {actual_geometry}")

    month_by_col = {}
    for rng in merged:
        if rng.min_row == 10 and rng.max_row == 10:
            value = norm(ws.cell(10, rng.min_col).value)
            if value in MONTHS:
                for col in range(rng.min_col, rng.max_col + 1):
                    month_by_col[col] = MONTHS[value]

    date_by_col = {}
    for col in range(3, ws.max_column + 1):
        raw_day = ws.cell(11, col).value
        month = month_by_col.get(col)
        if raw_day is None or month is None:
            continue
        year = 2027 if month == 1 else 2026
        date_by_col[col] = date(year, month, int(raw_day)).isoformat()
    date_table = [date_by_col[col] for col in sorted(date_by_col)]
    if len(date_table) != 122 or len(set(date_table)) != 122:
        raise SystemExit(f"unexpected source calendar: {len(date_table)} columns")

    references = {}
    for row in range(25, 36):
        discipline = norm(ws[f"C{row}"].value)
        if not discipline:
            continue
        first = parse_time(ws[f"BT{row}"].value)
        second = parse_time(ws[f"BX{row}"].value)
        if first is None and second is None:
            raise SystemExit(f"reference row {row} has no usable cycle time")
        references[row] = {
            "discipline": discipline,
            "assessment": assessment(ws[f"S{row}"].value, source_info["sourceId"], sheet_name, row),
            "location": join_location(ws[f"AN{row}"].value, ws[f"BL{row}"].value),
            "time": first or second,
            "timeSource": f"BT{row}" if first else f"BX{row}",
        }

    discipline_table = []
    location_table = []
    assessment_by_index = {}
    discipline_index_by_row = {}
    location_index_by_row = {}
    for row in sorted(references):
        ref = references[row]
        discipline_index_by_row[row] = len(discipline_table)
        discipline_table.append(ref["discipline"])
        if ref["location"] not in location_table:
            location_table.append(ref["location"])
        location_index_by_row[row] = location_table.index(ref["location"])
        if ref["assessment"]:
            assessment_by_index[str(discipline_index_by_row[row])] = ref["assessment"]

    group_table = source["expectedGroupIds"]
    decisions = []
    block_records = []
    upper_date_coverage = 0
    seen_group_blocks = set()
    for row in range(13, 20):
        group = norm(ws[f"B{row}"].value)
        if group not in group_table:
            raise SystemExit(f"unexpected group in B{row}: {group!r}")
        group_index = group_table.index(group)
        row_ranges = sorted(
            [rng for rng in merged if rng.min_row == row and rng.max_row == row and rng.min_col >= 3],
            key=lambda rng: rng.min_col,
        )
        for rng in row_ranges:
            raw = norm(ws.cell(row, rng.min_col).value)
            if not raw:
                continue
            if raw not in RAW_TO_REFERENCE_ROW:
                raise SystemExit(f"unmapped upper block {get_column_letter(rng.min_col)}{row}: {raw!r}")
            ref_row = RAW_TO_REFERENCE_ROW[raw]
            ref = references[ref_row]
            dates = [date_by_col[col] for col in range(rng.min_col, rng.max_col + 1) if col in date_by_col]
            if not dates:
                raise SystemExit(f"upper block without source dates: {rng}")
            locator = f"{get_column_letter(rng.min_col)}{row}"
            seen_group_blocks.add(locator)
            start_time, end_time = ref["time"]
            decisions.append([
                locator,
                group_mask(group_index),
                date_mask(date_table, dates),
                start_time,
                end_time,
                discipline_index_by_row[ref_row],
                0,
                location_index_by_row[ref_row],
            ])
            upper_date_coverage += len(dates)
            block_records.append({
                "locator": locator,
                "range": str(rng),
                "groupId": group,
                "rawValue": raw,
                "normalizedDiscipline": ref["discipline"],
                "referenceRow": ref_row,
                "timeSource": ref["timeSource"],
                "startTime": start_time,
                "endTime": end_time,
                "startDate": dates[0],
                "endDate": dates[-1],
                "dateCount": len(dates),
            })

    if len(seen_group_blocks) != 77:
        raise SystemExit(f"expected 77 event-bearing upper blocks, got {len(seen_group_blocks)}")
    if upper_date_coverage != 805:
        raise SystemExit(f"expected 805 normalized upper-grid events, got {upper_date_coverage}")

    shared = []
    for rng in merged:
        if rng.min_row <= 19 and rng.max_row >= 13 and rng.min_row != rng.max_row and rng.min_col >= 3:
            value = norm(ws.cell(rng.min_row, rng.min_col).value)
            if value:
                shared.append((str(rng), value, rng.min_col, rng.max_col))
    if len(shared) != 1 or shared[0][0] != "DN13:DT19" or shared[0][1].lower() != "экзамены":
        raise SystemExit(f"unexpected shared group blocks: {shared}")
    exam_dates = [date_by_col[col] for col in range(shared[0][2], shared[0][3] + 1) if col in date_by_col]

    pe_discipline = norm(ws["C36"].value)
    pe_raw = norm(ws["BT36"].value)
    if pe_discipline != "Дисциплины по физической культуре и спорту":
        raise SystemExit(f"unexpected PE discipline: {pe_discipline!r}")
    if pe_raw != "Четверг с 04.09 по 18.12 14:30-16:00":
        raise SystemExit(f"unexpected PE schedule: {pe_raw!r}")
    pe_start = date(2026, 9, 4)
    pe_end = date(2026, 12, 18)
    thursdays = dates_for_weekday(pe_start, pe_end, 3)
    fridays = dates_for_weekday(pe_start, pe_end, 4)
    if len(thursdays) != 15 or len(fridays) != 16:
        raise SystemExit("unexpected PE alternative date counts")

    manifest = {
        "schema": "kgmu-explicit-semantic-decisions-v3",
        "fixtureId": source["fixtureId"],
        "sourceSha256": source_info["sha256"],
        "parserRulesVersion": source["parserRulesVersion"],
        "sheetName": sheet_name,
        "semanticDecisionMode": "operator-authored-explicit",
        "logicalSourceCellCount": 78,
        "decisionCount": len(decisions),
        "dateTable": date_table,
        "disciplineTable": discipline_table,
        "locationTable": location_table,
        "assessmentMetadataByDisciplineIndex": assessment_by_index,
        "groupTable": group_table,
        "lessonTypeTable": ["practice"],
        "tupleFields": [
            "locator", "groupMaskHex", "dateMaskHex", "startTime", "endTime",
            "disciplineIndex", "lessonTypeIndex", "locationIndex"
        ],
        "decisions": decisions,
    }

    review = {
        "reviewId": "review-pediatrics-531-537-2026-09-02-v1",
        "parsingJobId": job["jobId"],
        "fixtureId": source["fixtureId"],
        "sourceArtifactId": source_info["sourceArtifactId"],
        "sourceSha256": source_info["sha256"],
        "parserProfile": source["parserProfile"],
        "parserRulesVersion": source["parserRulesVersion"],
        "status": "REVIEW_REQUIRED",
        "rulesApplied": ["G01", "G02", "G04", "G06", "G11", "G12", "G13", "G14", "G19", "G20", "G21", "C01", "C07", "C08", "C12", "C14"],
        "coverage": {
            "upperCycleBlockCount": 77,
            "normalizedUpperCycleBlockCount": 77,
            "upperGridNormalizedEventCount": upper_date_coverage,
            "serviceExamBlockCount": 1,
            "serviceExamDateCount": len(exam_dates),
            "independentLowerScheduleCount": 1,
            "normalizedIndependentLowerScheduleCount": 0,
            "eventBearingLogicalSourceBlockCount": 78,
            "coveredEventBearingSourceBlockCount": 77,
        },
        "serviceBlocks": [{
            "locator": "DN13",
            "range": "DN13:DT19",
            "value": "экзамены",
            "dates": exam_dates,
            "classification": "service-exam-period-no-event",
            "rules": ["C07", "C14"],
        }],
        "upperCycleBlocks": block_records,
        "unresolvedAmbiguities": [{
            "id": "PED5-PE-WEEKDAY-RANGE-CONTRADICTION",
            "severity": "blocking",
            "sourceRefs": [f"{sheet_name}!C36", f"{sheet_name}!BT36"],
            "discipline": pe_discipline,
            "rawSchedule": pe_raw,
            "facts": {
                "weekdayLabel": "Четверг",
                "rangeStart": pe_start.isoformat(),
                "rangeStartWeekday": "Friday",
                "rangeEnd": pe_end.isoformat(),
                "rangeEndWeekday": "Friday",
                "time": "14:30-16:00",
                "thursdayInterpretation": {"dateCount": len(thursdays), "dates": thursdays},
                "fridayRangeEndpointInterpretation": {"dateCount": len(fridays), "dates": fridays},
            },
            "rules": ["G04", "G06", "G21", "C12"],
            "reason": "The source gives weekday 'Thursday' but both explicit range endpoints 04.09.2026 and 18.12.2026 are Fridays. Existing general/cyclic rules do not authorize choosing one interpretation for this source.",
            "requiredConfirmation": "Confirm whether PE is scheduled on Thursdays within 04.09–18.12 or on Fridays following the explicit range endpoints; do not publish before confirmation.",
        }],
        "normalizationSummary": {
            "draftScope": "all unambiguous upper cyclic blocks only",
            "normalizedUpperGridEvents": upper_date_coverage,
            "groups": group_table,
            "excludedPendingReview": ["independent PE schedule C36/BT36"],
            "sharedCoreChangeRequired": False,
            "publicationAllowed": False,
        },
    }

    QA_DIR.mkdir(parents=True, exist_ok=True)
    MANIFEST_PATH.write_text(json.dumps(manifest, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    REVIEW_PATH.write_text(json.dumps(review, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "sourceArtifactId": source_info["sourceArtifactId"],
        "parsingJobId": job["jobId"],
        "upperCycleBlocks": len(decisions),
        "normalizedUpperGridEvents": upper_date_coverage,
        "serviceExamBlocks": 1,
        "unresolvedAmbiguities": len(review["unresolvedAmbiguities"]),
        "reviewStatus": review["status"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
