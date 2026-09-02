#!/usr/bin/env python3
"""Mechanical XLSX source probe for KGMU Dentistry course 5, groups 591-594."""
import hashlib
import html
import io
import json
import re
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "fixtures/2026-2027-semester-1/dentistry-591-594.source-probe.json"
TIMETABLE_PAGE = "https://kirovgma.ru/raspisanie-stomatologicheskiy-fakultet"
ALLOWED_PREFIX = "/sites/default/files/files/"
SOURCE_PATTERN = re.compile(r"5_stomat[^\"'<>\s]*\.xlsx", re.IGNORECASE)
EXPECTED_GROUPS = ["591", "592", "593", "594"]


def request_bytes(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "kgmu-calendar-source-probe/1.0"})
    with urllib.request.urlopen(req, timeout=45) as response:
        return response.read()


def discover_source() -> dict:
    page = html.unescape(request_bytes(TIMETABLE_PAGE).decode("utf-8", errors="replace"))
    hrefs = re.findall(r"href\s*=\s*[\"']([^\"']+)[\"']", page, flags=re.IGNORECASE)
    matches = []
    for href in hrefs:
        decoded = urllib.parse.unquote(href)
        if not SOURCE_PATTERN.search(decoded):
            continue
        absolute = urllib.parse.urljoin(TIMETABLE_PAGE, href)
        parsed = urllib.parse.urlparse(absolute)
        if parsed.netloc != "kirovgma.ru" or not parsed.path.startswith(ALLOWED_PREFIX):
            raise SystemExit(f"source outside allowed KGMU path: {absolute}")
        matches.append(absolute)
    matches = sorted(set(matches))
    if len(matches) != 1:
        raise SystemExit(f"expected exactly one current XLSX for dentistry 591-594, found {matches}")
    return {"program": "dentistry", "course": 5, "groups": EXPECTED_GROUPS, "url": matches[0]}


def scalar(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def workbook_dump(source: dict) -> dict:
    data = request_bytes(source["url"])
    if not data.startswith(b"PK"):
        raise SystemExit(f"source is not XLSX/ZIP: {source['url']}")
    wb = load_workbook(io.BytesIO(data), data_only=False)
    sheets = []
    for ws in wb.worksheets:
        non_empty = []
        comments = []
        formulas = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.comment is not None:
                    comments.append({"coord": cell.coordinate, "text": cell.comment.text, "author": cell.comment.author})
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({"coord": cell.coordinate, "formula": cell.value})
                if cell.value is not None:
                    non_empty.append({"coord": cell.coordinate, "value": str(cell.value).strip()})
        validations = []
        if ws.data_validations is not None:
            for item in ws.data_validations.dataValidation:
                validations.append({"sqref": str(item.sqref), "type": scalar(item.type), "formula1": scalar(item.formula1), "formula2": scalar(item.formula2)})
        sheets.append({
            "title": ws.title,
            "state": ws.sheet_state,
            "maxRow": ws.max_row,
            "maxColumn": ws.max_column,
            "mergedRanges": [str(rng) for rng in ws.merged_cells.ranges],
            "nonEmptyCellCount": len(non_empty),
            "nonEmptyCells": non_empty,
            "structuralEvidence": {
                "hiddenRows": [idx for idx, dim in ws.row_dimensions.items() if dim.hidden],
                "hiddenColumns": [key for key, dim in ws.column_dimensions.items() if dim.hidden],
                "comments": comments,
                "formulas": formulas,
                "dataValidations": validations,
                "imageCount": len(ws._images),
                "chartCount": len(ws._charts),
            },
        })
    return {
        **source,
        "sha256": hashlib.sha256(data).hexdigest(),
        "byteLength": len(data),
        "sheetNames": wb.sheetnames,
        "definedNames": [{"name": name, "attrText": scalar(defn.attr_text), "hidden": scalar(defn.hidden)} for name, defn in wb.defined_names.items()],
        "sheets": sheets,
    }


def main() -> None:
    payload = {
        "schema": "kgmu-mechanical-source-probe-v2",
        "semanticParsingPerformed": False,
        "timetablePage": TIMETABLE_PAGE,
        "source": workbook_dump(discover_source()),
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    src = payload["source"]
    print(json.dumps({
        "url": src["url"],
        "sha256": src["sha256"],
        "byteLength": src["byteLength"],
        "sheetNames": src["sheetNames"],
        "dimensions": [[s["title"], s["maxRow"], s["maxColumn"], len(s["mergedRanges"]), s["nonEmptyCellCount"]] for s in src["sheets"]],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
