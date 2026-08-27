#!/usr/bin/env python3
import collections
import datetime as dt
import hashlib
import itertools
import json
import re
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple

ROOT = Path(__file__).resolve().parents[2]
SOURCE_META = ROOT / 'fixtures/2026-2027-semester-1/medicine-101-110.source.json'
OUT_DIR = ROOT / 'fixtures/2026-2027-semester-1/normalized'
QA_DIR = ROOT / 'qa/2026-2027-semester-1'
PERIOD_ID = '2026-2027-semester-1'
UNIVERSITY_ID = 'kirov-gmu'
SOURCE_ID = 'medicine'
SHEET_NAME = '1 леч.1'
EXPECTED_GROUP_COUNTS = {'101':336,'102':336,'103':335,'104':335,'105':336,'106':336,'107':338,'108':338,'109':361,'110':361}
EXPECTED_TOTAL = 3412
EXPECTED_SOURCE_CELLS = 145
EXPECTED_OVERLAPS = 8

DATE_PAT = r'\d{1,2}\.(?:12|11|10|0?9|0?1)(?!\d)'
TIME_TOKEN = r'\d{1,2}\.\d{2}'

DISC_PATTERNS = [
    ('Элективные дисциплины по физической культуре и спорту', r'ЭЛЕКТИВНЫЕ ДИСЦИПЛИНЫ \(МОДУЛИ\) ПО ФИЗИЧЕСКОЙ КУЛЬТУРЕ И СПОРТУ'),
    ('Основы российской государственности', r'Основы российской государственности|ОСНОВЫ РОССИЙСКОЙ ГОСУДАРСТВЕННОСТИ'),
    ('Общая и биоорганическая химия', r'Общая и биоорганическая химия|ОБЩАЯ И БИООРГАНИЧЕСКАЯ ХИМИЯ'),
    ('Безопасность жизнедеятельности', r'Безопасность жизнедеятельности|БЕЗОПАСНОСТЬ ЖИЗНЕДЕЯТЕЛЬНОСТИ'),
    ('Психология и педагогика', r'Психология и педагогика|ПСИХОЛОГИЯ И ПЕДАГОГИКА'),
    ('Физика, математика', r'Физика, математика|ФИЗИКА, МАТЕМАТИКА'),
    ('Иностранный язык', r'Иностранный язык'),
    ('История медицины', r'История медицины|ИСТОРИЯ МЕДИЦИНЫ'),
    ('История России', r'История России|ИСТОРИЯ РОССИИ'),
    ('Латинский язык', r'Латинский язык|Латинский\b'),
    ('Правоведение', r'Правоведение|ПРАВОВЕДЕНИЕ'),
    ('Экономика', r'Экономика|ЭКОНОМИКА'),
    ('Анатомия', r'Анатомия|АНАТОМИЯ'),
    ('Биология', r'Биология|БИОЛОГИЯ'),
    ('УПО. Общий уход', r'УПО\.?\s*(?:ЛЕКЦИЯ\s*)?ОБЩИЙ УХОД|УПО\.Общий уход|УПО\. Общий уход'),
    ('Библиотечный час', r'Библиотечный час'),
    ('Час куратора', r'Час куратора'),
]
DISC_ALT = '|'.join(f'(?P<d{i}>{pat})' for i, (_, pat) in enumerate(DISC_PATTERNS))
TIME_PREFIX = rf'(?P<times>{TIME_TOKEN}\s*-\s*{TIME_TOKEN}(?:\s*,\s*{TIME_TOKEN}\s*-\s*{TIME_TOKEN})?|{TIME_TOKEN}\s*-\s*{TIME_TOKEN}\s*-\s*{TIME_TOKEN}\s*-\s*{TIME_TOKEN})'
EVENT_RE = re.compile(rf'{TIME_PREFIX}\s+(?P<lecture>ЛЕКЦИЯ\s+)?(?:{DISC_ALT})', re.I)
CONTROL_RE = re.compile(rf'(?P<date>{DATE_PAT}).{{0,45}}?(?P<control>зач[её]т(?:\s+с\s+оценкой)?|экзамен).{{0,25}}?(?P<t1>{TIME_TOKEN})\s*-\s*(?P<t2>{TIME_TOKEN})', re.I)
CONTROL_RE2 = re.compile(rf'(?P<control>зач[её]т(?:\s+с\s+оценкой)?|экзамен).{{0,25}}?(?P<date>{DATE_PAT}).{{0,15}}?(?P<t1>{TIME_TOKEN})\s*-\s*(?P<t2>{TIME_TOKEN})', re.I)

ADDR_BY_CORPUS = {'1':'ул. Владимирская, 137','2':'ул. Пролетарская, 38','3':'ул. Владимирская, 112'}
DEFAULT_LOCATIONS = {
    'История России':'1 корпус, ул. Владимирская, 137',
    'Экономика':'1 корпус, ул. Владимирская, 137',
    'Общая и биоорганическая химия':'1 корпус, ул. Владимирская, 137',
    'Физика, математика':'3 корпус, ул. Владимирская, 112',
    'УПО. Общий уход':'КОГБУЗ «Кировский областной клинический онкологический диспансер», пр-т Строителей, 23',
    'Элективные дисциплины по физической культуре и спорту':'ФОК, ул. Владимирская, 112',
    'Безопасность жизнедеятельности':'ул. Красноармейская, 35',
    'Психология и педагогика':'1 корпус, ул. Владимирская, 137',
    'Иностранный язык':'1 корпус, ул. Владимирская, 137',
    'Латинский язык':'1 корпус, ул. Владимирская, 137',
    'Биология':'3 корпус, ул. Владимирская, 112',
    'Анатомия':'2 корпус, ул. Пролетарская, 38',
    'Правоведение':'1 корпус, ул. Владимирская, 137',
    'История медицины':'1 корпус, ул. Владимирская, 137',
    'Основы российской государственности':'1 корпус, ул. Владимирская, 137',
}
DAY_ROWS = [(9,15,0),(16,21,1),(22,27,2),(28,33,3),(34,37,4),(38,42,5)]


def ddmm(value):
    day, month = map(int, value.split('.'))
    return dt.date(2026 if month >= 9 else 2027, month, day)


def week_intervals(values):
    return [(ddmm(a), ddmm(b)) for a,b in values]


WEEK1 = week_intervals([('01.09','05.09'),('14.09','19.09'),('28.09','03.10'),('12.10','17.10'),('26.10','31.10'),('09.11','14.11'),('23.11','28.11'),('07.12','12.12'),('21.12','26.12'),('11.01','16.01')])
WEEK2 = week_intervals([('07.09','12.09'),('21.09','26.09'),('05.10','10.10'),('19.10','24.10'),('02.11','07.11'),('16.11','21.11'),('30.11','05.12'),('14.12','19.12'),('28.12','30.12')])
ALL_WEEKS = WEEK1 + WEEK2


def in_grid(value):
    return any(start <= value <= end for start,end in ALL_WEEKS)


def week_no(value):
    if any(start <= value <= end for start,end in WEEK1): return 1
    if any(start <= value <= end for start,end in WEEK2): return 2
    return None


def dates_in_range(start, end, weekday, parity=None):
    result=[]
    current=start
    while current <= end:
        if current.weekday() == weekday and in_grid(current) and (parity is None or week_no(current) == parity):
            result.append(current)
        current += dt.timedelta(days=1)
    return result


def weekday_for_row(row):
    for first,last,weekday in DAY_ROWS:
        if first <= row <= last: return weekday
    raise ValueError(f'row outside timetable: {row}')


def parse_time(value):
    hour, minute = map(int, value.split('.'))
    return f'{hour:02d}:{minute:02d}'


def time_bounds(value):
    tokens = re.findall(TIME_TOKEN, value)
    return parse_time(tokens[0]), parse_time(tokens[-1])


def location_from_code(corpus, auditorium):
    corpus=str(int(corpus)); auditorium=str(int(auditorium))
    return f'{corpus} корпус, аудитория {auditorium}, {ADDR_BY_CORPUS[corpus]}'


def explicit_location(text, discipline=None):
    full=list(re.finditer(r'([123])\s*корпус\s*,?\s*аудитория\s*(\d{3})', text, re.I))
    if full:
        return location_from_code(*full[-1].groups())
    compact=list(re.finditer(r'(?<![\d.])([123])\s*-\s*(\d{3})(?!\d)', text))
    if compact:
        return location_from_code(*compact[-1].groups())
    if re.search(r'\bФОК\b', text, re.I):
        return 'ФОК, ул. Владимирская, 112'
    return DEFAULT_LOCATIONS.get(discipline)


def find_segments(text):
    matches=list(EVENT_RE.finditer(text))
    result=[]
    for index, match in enumerate(matches):
        discipline=None
        for i,(name,_) in enumerate(DISC_PATTERNS):
            if match.group(f'd{i}') is not None:
                discipline=name; break
        end=matches[index+1].start() if index+1 < len(matches) else len(text)
        result.append((match, discipline, text[match.end():end].strip()))
    return result


def event(group, value_date, start, end, discipline, lesson_type, location, coord, segment, extra=''):
    locator=f'{SHEET_NAME}!{coord}#s{segment}'
    key=f'{group}|{value_date.isoformat()}|{start}|{end}|{discipline}|{lesson_type}|{locator}|{extra}'
    return {
        'eventId':'kgmu-'+hashlib.sha256(key.encode()).hexdigest()[:24],
        'universityId':UNIVERSITY_ID,
        'groupId':group,
        'academicPeriodId':PERIOD_ID,
        'date':value_date.isoformat(),
        'startTime':start,
        'endTime':end,
        'timeSemantics':'floating',
        'discipline':discipline,
        'lessonType':lesson_type,
        'teacher':None,
        'location':location,
        'sourceRef':{'sourceId':SOURCE_ID,'locator':locator},
    }


def parse_segment(coord, row, groups, segment, match, discipline, tail):
    weekday=weekday_for_row(row)
    start,end=time_bounds(match.group('times'))
    lesson='lecture' if match.group('lecture') else ('other' if discipline in ('Библиотечный час','Час куратора') else 'practice')
    base_location=explicit_location(tail, discipline)
    work=tail
    events=[]

    if discipline == 'Час куратора' and re.search(r'\(2\s*недел', work, re.I) and re.search(r'\(1\s*недел', work, re.I):
        second=re.search(rf'({TIME_TOKEN})\s*-\s*({TIME_TOKEN})\s*\(1\s*недел', work, re.I)
        if not second: raise ValueError(f'{coord}: cannot parse R89')
        week1_start,week1_end=parse_time(second.group(1)),parse_time(second.group(2))
        for left,right in WEEK2:
            for value_date in dates_in_range(left,right,weekday,2):
                for group in groups: events.append(event(group,value_date,start,end,discipline,'other',None,coord,segment,'R89-w2'))
        for left,right in WEEK1:
            for value_date in dates_in_range(left,right,weekday,1):
                for group in groups: events.append(event(group,value_date,week1_start,week1_end,discipline,'other',None,coord,segment,'R89-w1'))
        return events

    if discipline == 'Час куратора' and not re.search(DATE_PAT, work):
        for value_date in dates_in_range(dt.date(2026,9,1),dt.date(2027,1,16),weekday)[:2]:
            for group in groups: events.append(event(group,value_date,start,end,discipline,'other',None,coord,segment,'R17'))
        return events

    location_overrides={}
    for paren in list(re.finditer(r'\(([^()]*)\)', work)):
        content=paren.group(1)
        loc=re.search(r'(?<!\d)([123])\s*-\s*(\d{3})(?!\d)', content)
        if loc and not re.search(r'зач|экзам', content, re.I):
            values=[ddmm(x) for x in re.findall(DATE_PAT, content[:loc.start()])]
            if values:
                location=location_from_code(*loc.groups())
                for value_date in values: location_overrides[value_date]=location
                work=work.replace(paren.group(0),' ')

    extras=[]
    for paren in list(re.finditer(r'\(([^()]*)\)', work)):
        content=paren.group(1)
        control=CONTROL_RE.search(content) or CONTROL_RE2.search(content)
        if control:
            value_date=ddmm(control.group('date'))
            control_start,control_end=parse_time(control.group('t1')),parse_time(control.group('t2'))
            control_type='exam' if 'экзам' in control.group('control').lower() else 'credit'
            extras.append((value_date,control_start,control_end,control_type,explicit_location(content,discipline) or base_location,'control'))
            work=work.replace(paren.group(0),' '); continue
        override=re.search(rf'(?P<date>{DATE_PAT})\s*-\s*(?P<t1>{TIME_TOKEN})\s*-\s*(?P<t2>{TIME_TOKEN})', content)
        if override:
            extras.append((ddmm(override.group('date')),parse_time(override.group('t1')),parse_time(override.group('t2')),lesson,base_location,'override'))
            work=work.replace(paren.group(0),' '); continue
        extra=re.search(rf'(?P<date>{DATE_PAT}).{{0,12}}?(?P<t1>{TIME_TOKEN})\s*-\s*(?P<t2>{TIME_TOKEN})', content)
        if extra:
            extras.append((ddmm(extra.group('date')),parse_time(extra.group('t1')),parse_time(extra.group('t2')),lesson,base_location,'extra'))
            work=work.replace(paren.group(0),' ')

    override_pattern=re.compile(rf'(?P<date>{DATE_PAT})\s*-\s*(?P<t1>{TIME_TOKEN})\s*-\s*(?P<t2>{TIME_TOKEN})')
    overrides=[]
    for override in list(override_pattern.finditer(work)):
        overrides.append((ddmm(override.group('date')),parse_time(override.group('t1')),parse_time(override.group('t2'))))
    work=override_pattern.sub(lambda item:item.group('date'), work)

    parity=None
    parity_match=re.search(r'([12])\s*недел', work, re.I)
    if parity_match: parity=int(parity_match.group(1))
    date_records=[]
    bounded=re.search(rf'с\s*({DATE_PAT})\s*по\s*({DATE_PAT})', work, re.I)
    if bounded:
        date_records.extend(dates_in_range(ddmm(bounded.group(1)),ddmm(bounded.group(2)),weekday,parity))
    else:
        until=re.search(rf'\bпо\s*({DATE_PAT})', work, re.I)
        if parity and until:
            date_records.extend(dates_in_range(dt.date(2026,9,1),ddmm(until.group(1)),weekday,parity))
            date_records.extend(ddmm(x) for x in re.findall(DATE_PAT, work[until.end():]))
        else:
            temp=re.sub(r'(?<!\d)[123]\s*-\s*\d{3}(?!\d)',' ',work)
            temp=re.split(r'\b[123]\s*корпус\b',temp,maxsplit=1,flags=re.I)[0]
            range_pattern=re.compile(rf'({DATE_PAT})\s*-\s*({DATE_PAT})')
            ranges=list(range_pattern.finditer(temp))
            if ranges:
                for item in ranges:
                    date_records.extend(dates_in_range(ddmm(item.group(1)),ddmm(item.group(2)),weekday,parity))
                temp=range_pattern.sub(' ',temp)
            date_records.extend(ddmm(x) for x in re.findall(DATE_PAT,temp))

    override_map={value_date:(override_start,override_end) for value_date,override_start,override_end in overrides}
    paren_override_map={value_date:(override_start,override_end,override_type,override_location) for value_date,override_start,override_end,override_type,override_location,kind in extras if kind == 'override'}
    for value_date in sorted(set(date_records)):
        current_start,current_end=start,end
        current_type=lesson
        current_location=location_overrides.get(value_date,base_location)
        if value_date in override_map: current_start,current_end=override_map[value_date]
        if value_date in paren_override_map: current_start,current_end,current_type,current_location=paren_override_map[value_date]
        for group in groups: events.append(event(group,value_date,current_start,current_end,discipline,current_type,current_location,coord,segment))

    for value_date,extra_start,extra_end,extra_type,extra_location,kind in extras:
        if kind == 'override': continue
        if kind == 'control':
            events=[item for item in events if not (item['date'] == value_date.isoformat() and item['discipline'] == discipline)]
        for group in groups: events.append(event(group,value_date,extra_start,extra_end,discipline,extra_type,extra_location,coord,segment,kind))
    if not events: raise ValueError(f'{coord} segment {segment}: no events from {tail!r}')
    return events


def overlaps(left,right):
    def minutes(value):
        hour,minute=map(int,value.split(':')); return hour*60+minute
    return minutes(left['startTime']) < minutes(right['endTime']) and minutes(right['startTime']) < minutes(left['endTime'])


def source_coord(item):
    return re.search(r'!([A-Z]+\d+)#',item['sourceRef']['locator']).group(1)


def main():
    meta=json.loads(SOURCE_META.read_text(encoding='utf-8'))
    data=urllib.request.urlopen(meta['source']['url'],timeout=30).read()
    actual=hashlib.sha256(data).hexdigest()
    if actual != meta['source']['sha256']: raise SystemExit(f'SHA mismatch: {actual}')
    if len(data) != meta['source']['byteLength']: raise SystemExit(f'byteLength mismatch: {len(data)}')
    xlsx=ROOT/'.tmp-kgmu-101-110.xlsx'; xlsx.write_bytes(data)
    try:
        workbook=load_workbook(xlsx,data_only=False)
    finally:
        xlsx.unlink(missing_ok=True)
    if workbook.sheetnames != meta['workbookExpectations']['sheetNames']: raise SystemExit('sheet names changed')
    sheet=workbook[SHEET_NAME]
    if sheet.max_row != meta['workbookExpectations']['maxRow'] or sheet.max_column != meta['workbookExpectations']['maxColumn']: raise SystemExit('workbook geometry changed')
    if len(sheet.merged_cells.ranges) != meta['workbookExpectations']['mergedRangeCount']: raise SystemExit('merged range count changed')
    nonempty=sum(1 for row in sheet.iter_rows() for cell in row if cell.value is not None)
    if nonempty != meta['workbookExpectations']['nonEmptyCellCount']: raise SystemExit('non-empty cell count changed')

    groups_by_column={column:str(99+column) for column in range(2,12)}
    logical=[]
    for row in sheet.iter_rows(min_row=9,max_row=42,min_col=2,max_col=11):
        for cell in row:
            if cell.value is not None: logical.append(cell)
    if len(logical) != EXPECTED_SOURCE_CELLS: raise SystemExit(f'logical cell count changed: {len(logical)}')

    normalized=[]
    for cell in logical:
        row,column=coordinate_to_tuple(cell.coordinate)
        covered=[column]
        for merged in sheet.merged_cells.ranges:
            if cell.coordinate in merged:
                covered=list(range(merged.min_col,merged.max_col+1)); break
        groups=[groups_by_column[value] for value in covered if value in groups_by_column]
        segments=find_segments(str(cell.value))
        if not segments: raise SystemExit(f'unclassified source cell: {cell.coordinate}: {cell.value}')
        for index,(match,discipline,tail) in enumerate(segments,1):
            normalized.extend(parse_segment(cell.coordinate,row,groups,index,match,discipline,tail))

    normalized.sort(key=lambda item:(int(item['groupId']),item['date'],item['startTime'],item['endTime'],item['discipline'],item['lessonType'],item['sourceRef']['locator']))
    counts=dict(sorted(collections.Counter(item['groupId'] for item in normalized).items(),key=lambda pair:int(pair[0])))
    if len(normalized) != EXPECTED_TOTAL: raise SystemExit(f'event count changed: {len(normalized)}')
    if counts != EXPECTED_GROUP_COUNTS: raise SystemExit(f'group counts changed: {counts}')
    if len({item['eventId'] for item in normalized}) != len(normalized): raise SystemExit('duplicate eventId')
    signatures=[(item['groupId'],item['date'],item['startTime'],item['endTime'],item['discipline'],item['lessonType'],item['location']) for item in normalized]
    if len(set(signatures)) != len(signatures): raise SystemExit('duplicate normalized events')
    source_cells=sorted({source_coord(item) for item in normalized})
    if len(source_cells) != EXPECTED_SOURCE_CELLS: raise SystemExit(f'source coverage changed: {len(source_cells)}')
    if any(not in_grid(dt.date.fromisoformat(item['date'])) for item in normalized): raise SystemExit('event outside service-week grid')
    if any(item['timeSemantics'] != 'floating' for item in normalized): raise SystemExit('non-floating event')

    by_group_date=collections.defaultdict(list)
    for item in normalized: by_group_date[(item['groupId'],item['date'])].append(item)
    overlap_rows=[]
    for (group,value_date),items in sorted(by_group_date.items()):
        for left,right in itertools.combinations(items,2):
            if overlaps(left,right):
                overlap_rows.append({'groupId':group,'date':value_date,'left':{'startTime':left['startTime'],'endTime':left['endTime'],'discipline':left['discipline'],'sourceLocator':left['sourceRef']['locator']},'right':{'startTime':right['startTime'],'endTime':right['endTime'],'discipline':right['discipline'],'sourceLocator':right['sourceRef']['locator']}})
    if len(overlap_rows) != EXPECTED_OVERLAPS: raise SystemExit(f'overlap count changed: {len(overlap_rows)}')

    specs=[
      {'note':'C9','group':'102','disc':'Биология','coord':'C34','dates':['2027-01-15']},
      {'note':'K9','group':'110','disc':'Биология','coord':'K27','dates':['2026-12-02','2026-12-16']},
      {'note':'B10','group':'101','disc':'Биология','coord':'B34','dates':['2027-01-15']},
      {'note':'B10','group':'102','disc':'Биология','coord':'C34','dates':['2027-01-15']},
      {'note':'F17','group':'105','disc':'Биология','coord':'F27','dates':['2026-12-09']},
      {'note':'G18','group':'106','disc':'Биология','coord':'G27','dates':['2026-11-25']},
      {'note':'E22','group':'104','disc':'Биология','coord':'E26','dates':['2026-10-21']},
    ]
    for group in map(str,range(101,111)):
        specs.append({'note':'B24','group':group,'disc':'Психология и педагогика','coord':'B25','dates':['2026-12-16','2026-12-23']})
    specs += [
      {'note':'H26','group':'107','disc':'История России','coord':'H9','dates':['2026-12-07','2026-12-14','2026-12-21']},
      {'note':'J29','group':'109','disc':'Биология','coord':'J9','dates':['2026-10-19','2026-10-26']},
      {'note':'G34','group':'106','disc':'Физика, математика','coord':'G28','dates':['2026-12-03']},
      {'note':'I34','group':'108','disc':'Физика, математика','coord':'I9','dates':['2026-11-30']},
      {'note':'D35','group':'103','disc':'Биология','coord':'D26','dates':['2026-10-28']},
      {'note':'H35','group':'107','disc':'Биология','coord':'H29','dates':['2027-01-14']},
      {'note':'I35','group':'108','disc':'Биология','coord':'I35','dates':['2026-11-30'],'startTime':'15:30','endTime':'17:55'},
      {'note':'H36','group':'107','disc':'Общая и биоорганическая химия','coord':'H9','dates':['2026-11-30']},
      {'note':'D40','group':'103','disc':'Физика, математика','coord':'D26','dates':['2026-12-23']},
    ]
    r83=[]
    for spec in specs:
        matches=[item for item in normalized if item['groupId']==spec['group'] and item['discipline']==spec['disc'] and source_coord(item)==spec['coord'] and item['date'] in spec['dates']]
        if 'startTime' in spec: matches=[item for item in matches if item['startTime']==spec['startTime'] and item['endTime']==spec['endTime']]
        ok=sorted(item['date'] for item in matches)==sorted(spec['dates'])
        if not ok: raise SystemExit(f'R83 check failed: {spec}')
        r83.append({**spec,'status':'pass','matchedEventIds':[item['eventId'] for item in matches]})

    canonical=json.dumps(normalized,ensure_ascii=False,separators=(',',':'),sort_keys=True)
    candidate_digest=hashlib.sha256(canonical.encode()).hexdigest()
    tuples=[[item['eventId'],item['groupId'],item['date'],item['startTime'],item['endTime'],item['discipline'],item['lessonType'],item['location'],item['sourceRef']['locator']] for item in normalized]
    fixture={
      'fixtureId':meta['fixtureId'],
      'encoding':'normalized-event-tuples-v1',
      'tupleFields':['eventId','groupId','date','startTime','endTime','discipline','lessonType','location','sourceLocator'],
      'constants':{'universityId':UNIVERSITY_ID,'academicPeriodId':PERIOD_ID,'timeSemantics':'floating','sourceId':SOURCE_ID,'teacher':None},
      'eventCount':len(normalized),
      'candidateDigest':'sha256:'+candidate_digest,
      'events':tuples,
    }
    evidence={
      'fixtureId':meta['fixtureId'],'sourceSha256':actual,'parserRulesVersion':meta['parserRulesVersion'],
      'eventCount':len(normalized),'groupEventCounts':counts,'logicalSourceCellCount':len(logical),'coveredSourceCellCount':len(source_cells),
      'unresolvedAmbiguities':0,'duplicateEvents':0,'explicitOverlapWarningCount':len(overlap_rows),'explicitOverlapWarnings':overlap_rows,
      'r83Checks':r83,'candidateDigest':'sha256:'+candidate_digest,
    }
    qa_report={
      'qaReportId':'qa-kgmu-2026-2027-s1-medicine-101-110-v1',
      'parsingJobId':'parsing-job-101-110-1','candidateDigest':'sha256:'+candidate_digest,'decision':'pass',
      'checks':[
        {'code':'all-group-content-accounted-for','status':'pass','message':f'{len(source_cells)}/{len(logical)} logical timetable cells accounted for'},
        {'code':'dates-within-academic-period','status':'pass','message':'All generated dates are inside the official service-week grid through 2027-01-16'},
        {'code':'source-groups-match-expected-groups','status':'pass','message':'Groups 101-110 only'},
        {'code':'duplicate-events-resolved','status':'pass','message':'No duplicate normalized event signatures'},
        {'code':'explicit-overlaps-preserved','status':'warning','message':f'{len(overlap_rows)} source-explicit overlaps preserved under G16/R69'},
        {'code':'unresolved-ambiguities-zero-before-pass','status':'pass','message':'0 unresolved ambiguities'},
        {'code':'r83-additional-event-counts','status':'pass','message':f'{len(r83)} R83 completeness assertions passed'},
      ],
      'createdAt':'2026-08-27T04:11:00Z',
    }
    OUT_DIR.mkdir(parents=True,exist_ok=True); QA_DIR.mkdir(parents=True,exist_ok=True)
    (OUT_DIR/'medicine-101-110.normalized.compact.json').write_text(json.dumps(fixture,ensure_ascii=False,separators=(',',':'))+'\n',encoding='utf-8')
    (QA_DIR/'medicine-101-110.evidence.json').write_text(json.dumps(evidence,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
    (QA_DIR/'medicine-101-110.qa-report.json').write_text(json.dumps(qa_report,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
    print(json.dumps({'eventCount':len(normalized),'groupEventCounts':counts,'sourceCells':len(source_cells),'overlaps':len(overlap_rows),'candidateDigest':candidate_digest},ensure_ascii=False))


if __name__ == '__main__':
    main()
