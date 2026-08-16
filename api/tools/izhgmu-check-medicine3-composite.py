#!/usr/bin/env python3
import argparse
import hashlib
import json
import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

CORE_DISCIPLINES = {
    'Фармакология': 'Фармакология',
    'Стоматология': 'Стоматология',
    'Патофизиология': 'Патофизиология',
    'Общая хирургия': 'Общая хирургия',
    'ОЗЗ': 'Общественное здоровье и здравоохранение',
    'Патанатомия': 'Патологическая анатомия',
    'Пр.вн.бол': 'Пропедевтика внутренних болезней',
}

WEEKDAYS = {
    'Понедельник': 0,
    'Вторник': 1,
    'Среда': 2,
    'Четверг': 3,
    'Пятница': 4,
    'Суббота': 5,
}

MONTHS = {
    'Февраль': 2,
    'Март': 3,
    'Апрель': 4,
    'Май': 5,
}

REVIEWED_EARLY_LECTURE = {'start': '08:30', 'end': '10:05'}
REVIEWED_DENTISTRY_LECTURE = {'start': '11:20', 'end': '12:50'}
REVIEWED_PHYSICAL_EDUCATION = {'start': '16:35', 'end': '18:00'}


def norm(value):
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def normalize_clock(value):
    match = re.fullmatch(r'(\d{1,2})[.:](\d{2})', norm(value))
    if not match:
        return None
    return f'{int(match.group(1)):02d}:{match.group(2)}'


def parse_clock_range(value):
    match = re.fullmatch(
        r'\s*(\d{1,2}[.:]\d{2})\s*[-–]\s*(\d{1,2}[.:]\d{2})\s*',
        norm(value),
    )
    if not match:
        return None
    return {
        'start': normalize_clock(match.group(1)),
        'end': normalize_clock(match.group(2)),
    }


def parse_cycle_time_variants(value):
    match = re.fullmatch(
        r'\s*(\d{1,2}[.:]\d{2}\s*[-–]\s*\d{1,2}[.:]\d{2})\s*'
        r'\((\d{1,2}[.:]\d{2}\s*[-–]\s*\d{1,2}[.:]\d{2})\)\s*',
        norm(value),
    )
    if not match:
        raise AssertionError(f'IZH-M3 unsupported cycle time variants: {value!r}')
    primary = parse_clock_range(match.group(1))
    after_lecture = parse_clock_range(match.group(2))
    if not primary or not after_lecture or after_lecture['start'] != '10:15':
        raise AssertionError(f'IZH-M3 unexpected cycle time semantics: {value!r}')
    return {'primary': primary, 'afterLecture': after_lecture}


def sha256(data):
    return hashlib.sha256(data).hexdigest()


def find_source(report, source_kind):
    matches = [
        item for item in report.get('files', [])
        if item.get('status') == 'downloaded'
        and item.get('faculty') == 'medicine'
        and int(item.get('course') or 0) == 3
        and item.get('term') == 'spring'
        and item.get('language') == 'ru'
        and item.get('sourceKind') == source_kind
    ]
    if len(matches) != 1:
        raise AssertionError(f'IZH-M3 expected one {source_kind} source, got {len(matches)}')
    return matches[0]


def month_columns(sheet):
    anchors = []
    for col in range(1, sheet.max_column + 1):
        value = norm(sheet.cell(6, col).value)
        if value in MONTHS:
            anchors.append((col, MONTHS[value]))
    if [month for _, month in anchors] != [2, 3, 4, 5]:
        raise AssertionError(f'IZH-M3 lecture month grid changed: {anchors}')
    result = {}
    for index, (start_col, month) in enumerate(anchors):
        end_col = anchors[index + 1][0] - 1 if index + 1 < len(anchors) else 22
        for col in range(start_col, end_col + 1):
            result[col] = month
    return result


def stream_groups(label):
    text = norm(label)
    if 'англ' in text.lower():
        return None
    if 'все 3 потока' in text:
        return list(range(301, 327))
    match = re.search(r'([123])\s*поток\s*\((\d{3})-(\d{3})\s*групп', text, re.I)
    if not match:
        return None
    expected = {
        '1': (301, 310),
        '2': (311, 318),
        '3': (319, 326),
    }[match.group(1)]
    actual = (int(match.group(2)), int(match.group(3)))
    if actual != expected:
        raise AssertionError(f'IZH-M3 stream range changed: {text!r}')
    return list(range(actual[0], actual[1] + 1))


def parse_lecture_workbook(filename):
    workbook = load_workbook(filename, data_only=True, read_only=False)
    if len(workbook.sheetnames) != 1:
        raise AssertionError(f'IZH-M3 lecture sheet count changed: {workbook.sheetnames}')
    sheet = workbook[workbook.sheetnames[0]]
    if sheet.title != 'Лекции все':
        raise AssertionError(f'IZH-M3 lecture sheet changed: {sheet.title}')

    months = month_columns(sheet)
    current_day = None
    core = []
    invalid_weekday = []
    physical_education = []
    ignored_english_pe = []

    for row in range(7, sheet.max_row + 1):
        day_value = norm(sheet.cell(row, 1).value)
        if day_value:
            current_day = day_value
        if current_day not in WEEKDAYS:
            continue

        discipline_raw = norm(sheet.cell(row, 3).value)
        time_range = parse_clock_range(sheet.cell(row, 2).value)
        if not discipline_raw or not time_range:
            continue

        occurrences = []
        for col, month in months.items():
            raw_day = sheet.cell(row, col).value
            if not isinstance(raw_day, (int, float)) or not float(raw_day).is_integer():
                continue
            current_date = date(2026, month, int(raw_day))
            occurrence = {
                'date': current_date.isoformat(),
                'reference': f'{sheet.title}!{sheet.cell(row, col).coordinate}',
                'weekday': current_day,
                'row': row,
                'time': time_range,
                'disciplineRaw': discipline_raw,
                'location': norm(sheet.cell(row, 4).value) or None,
            }
            if current_date.weekday() != WEEKDAYS[current_day]:
                invalid_weekday.append(occurrence)
            else:
                occurrences.append(occurrence)

        canonical_discipline = CORE_DISCIPLINES.get(discipline_raw)
        if canonical_discipline:
            for occurrence in occurrences:
                core.append({**occurrence, 'discipline': canonical_discipline})
            continue

        if 'физическая культура' in discipline_raw.lower() or 'физическая культураи спорт' in discipline_raw.lower():
            groups = stream_groups(discipline_raw)
            if groups is None:
                ignored_english_pe.extend(occurrences)
                continue
            if time_range != REVIEWED_PHYSICAL_EDUCATION:
                raise AssertionError(f'IZH-M3 PE slot changed at row {row}: {time_range}')
            for occurrence in occurrences:
                physical_education.append({
                    **occurrence,
                    'discipline': 'Физическая культура и спорт',
                    'groups': groups,
                })

    return {
        'sheet': sheet.title,
        'core': core,
        'invalidWeekday': invalid_weekday,
        'physicalEducation': physical_education,
        'ignoredEnglishPhysicalEducation': ignored_english_pe,
    }


def blocker_key(item):
    return (
        item.get('kind'),
        item.get('discipline'),
        item.get('date'),
        tuple(item.get('references') or []),
    )


def dedupe(items):
    result = []
    seen = set()
    for item in items:
        key = blocker_key(item)
        if key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def build_group_pair(pair, lecture):
    groups = [int(group) for group in pair['groups']]
    core_by_key = {}
    for occurrence in lecture['core']:
        core_by_key.setdefault((occurrence['discipline'], occurrence['date']), []).append(occurrence)

    invalid_by_key = {}
    for occurrence in lecture['invalidWeekday']:
        discipline = CORE_DISCIPLINES.get(occurrence['disciplineRaw'])
        if discipline:
            invalid_by_key.setdefault((discipline, occurrence['date']), []).append(occurrence)

    blockers = []
    practice_events = []
    lecture_events = []

    for series in pair['series']:
        variants = parse_cycle_time_variants(series['timeRaw'])
        for current_date in series['dates']:
            key = (series['discipline'], current_date)
            invalid = invalid_by_key.get(key, [])
            if invalid:
                blockers.append({
                    'kind': 'medicine3_lecture_weekday_mismatch',
                    'discipline': series['discipline'],
                    'date': current_date,
                    'references': [item['reference'] for item in invalid],
                })

            matches = core_by_key.get(key, [])
            if len(matches) > 1:
                blockers.append({
                    'kind': 'medicine3_lecture_duplicate_exact_date',
                    'discipline': series['discipline'],
                    'date': current_date,
                    'references': [item['reference'] for item in matches],
                })
            lecture_occurrence = matches[0] if len(matches) == 1 else None

            shifted = False
            practice_time = variants['primary']
            if lecture_occurrence:
                if lecture_occurrence['time'] == REVIEWED_EARLY_LECTURE:
                    practice_time = variants['afterLecture']
                    shifted = True
                elif lecture_occurrence['time'] == REVIEWED_DENTISTRY_LECTURE:
                    practice_time = variants['primary']
                else:
                    blockers.append({
                        'kind': 'medicine3_core_lecture_slot_unreviewed',
                        'discipline': series['discipline'],
                        'date': current_date,
                        'references': [lecture_occurrence['reference']],
                    })

            practice_events.append({
                'sourceRole': 'cycle_practice',
                'date': current_date,
                'discipline': series['discipline'],
                'startTime': practice_time['start'],
                'endTime': practice_time['end'],
                'location': series.get('location'),
                'assessment': series.get('assessment'),
                'shiftedAfterMorningLecture': shifted,
                'sourceRanges': list(series.get('sourceRanges') or []),
                'timeReference': series.get('refs', {}).get('time'),
                'ruleIds': ['IZH-C3-01', 'IZH-C3-02', 'IZH-C3-07', 'IZH-C3-13'],
            })

            if lecture_occurrence:
                lecture_events.append({
                    'sourceRole': 'cycle_lecture',
                    'date': current_date,
                    'discipline': series['discipline'],
                    'startTime': lecture_occurrence['time']['start'],
                    'endTime': lecture_occurrence['time']['end'],
                    'location': lecture_occurrence.get('location'),
                    'reference': lecture_occurrence['reference'],
                    'ruleIds': ['IZH-C3-14'],
                })

    physical_education_events = []
    for occurrence in lecture['physicalEducation']:
        if not any(group in occurrence['groups'] for group in groups):
            continue
        physical_education_events.append({
            'sourceRole': 'physical_education',
            'date': occurrence['date'],
            'discipline': occurrence['discipline'],
            'startTime': occurrence['time']['start'],
            'endTime': occurrence['time']['end'],
            'location': occurrence.get('location'),
            'reference': occurrence['reference'],
            'ruleIds': ['IZH-C3-15'],
        })

    blockers = dedupe(blockers)
    expected_practice = 92
    if len(practice_events) != expected_practice:
        blockers.append({
            'kind': 'medicine3_practice_coverage_changed',
            'discipline': None,
            'date': None,
            'references': [],
        })

    if len(physical_education_events) != 30:
        blockers.append({
            'kind': 'medicine3_physical_education_coverage_changed',
            'discipline': 'Физическая культура и спорт',
            'date': None,
            'references': [item['reference'] for item in physical_education_events],
        })

    blockers = dedupe(blockers)
    return {
        'groupSpan': pair['groupSpan'],
        'groups': groups,
        'sourceResolved': len(blockers) == 0,
        'blockers': blockers,
        'stats': {
            'practiceEvents': len(practice_events),
            'shiftedPracticeEvents': sum(1 for item in practice_events if item['shiftedAfterMorningLecture']),
            'coreLectureEvents': len(lecture_events),
            'physicalEducationEvents': len(physical_education_events),
            'totalEvents': len(practice_events) + len(lecture_events) + len(physical_education_events),
        },
        'practiceEvents': practice_events,
        'lectureEvents': lecture_events,
        'physicalEducationEvents': physical_education_events,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--input-dir', default='/tmp/izhgmu-current')
    parser.add_argument('--output')
    args = parser.parse_args()

    root = Path(args.input_dir)
    report = json.loads((root / 'download-report.json').read_text(encoding='utf-8'))
    class_source = find_source(report, 'class')
    lecture_source = find_source(report, 'lecture')

    class_bytes = (root / class_source['filename']).read_bytes()
    lecture_bytes = (root / lecture_source['filename']).read_bytes()
    if sha256(class_bytes) != class_source['sha256']:
        raise AssertionError('IZH-M3 class SHA mismatch')
    if sha256(lecture_bytes) != lecture_source['sha256']:
        raise AssertionError('IZH-M3 lecture SHA mismatch')

    cycle_path = root / 'medicine3-legacy-cycle.json'
    if not cycle_path.exists():
        raise AssertionError('IZH-M3 legacy cycle parse must run before composite check')
    cycle = json.loads(cycle_path.read_text(encoding='utf-8'))
    if cycle.get('stats', {}).get('groupPairCount') != 13 or cycle.get('stats', {}).get('dateColumnCount') != 92:
        raise AssertionError(f'IZH-M3 cycle invariant changed: {cycle.get("stats")}')

    lecture = parse_lecture_workbook(root / lecture_source['filename'])
    group_pairs = [build_group_pair(pair, lecture) for pair in cycle['groupPairs']]
    if [item['groupSpan'] for item in group_pairs] != [f'{number}-{number + 1}' for number in range(301, 327, 2)]:
        raise AssertionError('IZH-M3 group pair sequence changed')

    resolved_groups = [group for item in group_pairs if item['sourceResolved'] for group in item['groups']]
    blocked_groups = [group for item in group_pairs if not item['sourceResolved'] for group in item['groups']]
    unexpected_blockers = [
        blocker for item in group_pairs for blocker in item['blockers']
        if blocker['kind'] != 'medicine3_lecture_weekday_mismatch'
    ]
    if unexpected_blockers:
        raise AssertionError(f'IZH-M3 unexpected blockers: {unexpected_blockers}')

    output = {
        'profile': 'IZH-MEDICINE3-CYCLE-LECTURE-COMPOSITE',
        'version': 1,
        'classSource': {
            'filename': class_source['filename'],
            'sha256': class_source['sha256'],
        },
        'lectureSource': {
            'filename': lecture_source['filename'],
            'sha256': lecture_source['sha256'],
            'sheet': lecture['sheet'],
        },
        'policy': {
            'practiceTime': 'primary unless same-cycle discipline has an exact 08:30-10:05 lecture on that source date; then use parenthesized 10:15 start variant',
            'lectureAudience': 'Russian group pair receives core lecture only when exact lecture discipline/date matches its current cycle discipline/date',
            'physicalEducationAudience': 'explicit source stream ranges only',
            'productionAuthorized': False,
        },
        'lectureDiagnostics': {
            'invalidWeekday': lecture['invalidWeekday'],
            'ignoredEnglishPhysicalEducationCount': len(lecture['ignoredEnglishPhysicalEducation']),
        },
        'groupPairs': group_pairs,
        'stats': {
            'groupPairCount': len(group_pairs),
            'groupCount': sum(len(item['groups']) for item in group_pairs),
            'sourceResolvedGroupCount': len(resolved_groups),
            'blockedGroupCount': len(blocked_groups),
            'resolvedGroups': resolved_groups,
            'blockedGroups': blocked_groups,
            'productionAuthorized': False,
        },
    }

    destination = Path(args.output) if args.output else root / 'medicine3-composite-diagnostic.json'
    destination.write_text(json.dumps(output, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    print('IZHGMU_MEDICINE3_COMPOSITE', json.dumps(output['stats'], ensure_ascii=False))
    for item in group_pairs:
        print('IZHGMU_MEDICINE3_GROUP_PAIR', json.dumps({
            'groupSpan': item['groupSpan'],
            'sourceResolved': item['sourceResolved'],
            **item['stats'],
            'blockers': item['blockers'],
        }, ensure_ascii=False))


if __name__ == '__main__':
    main()
